"""技术演化域数据导入（阶段二）：技术主数据 + 里程碑 + 算法知识层 → 统一库。

数据来源（自 embodied-job-evolution-lab 复制至 data/raw/）：
- 技术词主数据_20260727.xlsx：L1/L2/L3/L4 技术实体体系 → technologies 表
- milestones_export_6fields.xlsx：460 条技术里程碑 → milestones 表（附 MILESTONE_BRIDGE 桥接）
- 算法知识层种子（原 lab data_importer.py KNOWLEDGE_* 常量）→ capabilities / tasks / role_titles

幂等：全部 INSERT OR REPLACE，可重复执行。
用法：python3 -m pipeline.import_emerging_data
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from . import config, db

TAXONOMY_FILE = "技术词主数据_20260727.xlsx"
MILESTONES_FILE = "milestones_export_6fields.xlsx"

# 里程碑类目 → 技术编码桥接映射（移植自 lab data_importer.MILESTONE_BRIDGE）
MILESTONE_BRIDGE: dict[str, list[tuple[str, float]]] = {
    "VLA与端到端模型": [("T1.01", 1.0)],
    "世界模型": [("T1.02", 1.0)],
    "机器人控制与运动控制": [("T1.03", 1.0)],
    "强化学习": [("T1.05", 1.0)],
    "具身感知": [("T1.06", 0.7), ("T2.01", 0.3)],
    "灵巧操作与触觉": [("T1.09", 0.65), ("T2.02", 0.35)],
    "仿真平台与数据工程": [("T4.03", 0.55), ("T4.01", 0.25), ("T4.05", 0.20)],
    "Sim-to-Real迁移": [("T4.04", 1.0)],
    "硬件本体与零部件": [("T3", 1.0)],
    "产业政策与标准": [("T6.03", 0.75)],
    "大模型与认知架构": [("T1.01", 0.6), ("T1.07", 0.4)],
}

# 算法知识层种子（移植自 lab data_importer KNOWLEDGE_*；
# technology_code 为根编码，'fallback' 为兜底模板，{技术}为运行时占位符）
KNOWLEDGE_CAPABILITIES: dict[str, list[dict[str, str]]] = {
    "T1.02": [
        {"name": "预测机器人环境未来状态", "object": "环境状态", "scenario": "动作规划"},
        {"name": "学习环境动力学与物理规律", "object": "环境动力学", "scenario": "模型训练"},
        {"name": "在执行前模拟动作后果", "object": "动作结果", "scenario": "安全决策"},
        {"name": "生成可供规划使用的潜在状态", "object": "潜在状态", "scenario": "模型驱动规划"},
    ],
    "T1.01": [
        {"name": "联合理解视觉语言与动作", "object": "多模态指令", "scenario": "机器人控制"},
        {"name": "从多模态输入生成连续动作", "object": "动作序列", "scenario": "端到端控制"},
        {"name": "跨任务迁移机器人策略", "object": "通用策略", "scenario": "多任务学习"},
        {"name": "利用示范数据优化策略", "object": "机器人示范", "scenario": "策略训练"},
    ],
    "T4.04": [
        {"name": "将仿真策略迁移到真实机器人", "object": "控制策略", "scenario": "真实部署"},
        {"name": "缩小仿真与现实分布差异", "object": "域差异", "scenario": "模型训练"},
        {"name": "评估策略在真实环境中的鲁棒性", "object": "策略鲁棒性", "scenario": "迁移验证"},
    ],
}

KNOWLEDGE_TASKS: dict[str, list[dict[str, Any]]] = {
    "T1.02": [
        {"name": "构建机器人时序交互与环境状态数据集", "group": "data", "keywords": ["数据集", "数据采集", "时序", "轨迹", "标注"], "relevance": 0.92},
        {"name": "开发环境动力学与状态预测模型", "group": "model", "keywords": ["世界模型", "环境模型", "动力学", "状态预测", "视频预测"], "relevance": 0.98},
        {"name": "设计世界模型预测准确性评测指标", "group": "evaluation", "keywords": ["评测", "指标", "准确性", "基准", "benchmark"], "relevance": 0.96},
        {"name": "构建模型驱动的机器人动作规划流程", "group": "planning", "keywords": ["动作规划", "模型预测控制", "规划", "决策", "推演"], "relevance": 0.95},
        {"name": "优化世界模型推理性能与部署效率", "group": "deployment", "keywords": ["推理", "部署", "加速", "性能优化", "量化"], "relevance": 0.86},
        {"name": "验证模型预测失效模式与安全边界", "group": "evaluation", "keywords": ["失效", "安全", "鲁棒性", "测试", "验证"], "relevance": 0.91},
    ],
    "T1.01": [
        {"name": "构建视觉语言动作多模态训练数据", "group": "data", "keywords": ["多模态", "数据集", "示范", "轨迹", "标注"], "relevance": 0.96},
        {"name": "开发视觉语言动作联合建模架构", "group": "model", "keywords": ["VLA", "多模态", "动作生成", "端到端", "模型架构"], "relevance": 0.99},
        {"name": "训练跨任务机器人通用操作策略", "group": "model", "keywords": ["策略训练", "模仿学习", "强化学习", "行为克隆"], "relevance": 0.94},
        {"name": "设计VLA指令跟随与操作评测体系", "group": "evaluation", "keywords": ["评测", "指令跟随", "成功率", "基准"], "relevance": 0.92},
        {"name": "优化VLA实时推理与机器人部署", "group": "deployment", "keywords": ["推理", "部署", "实时", "延迟", "端侧"], "relevance": 0.88},
    ],
    "T4.04": [
        {"name": "构建仿真与真实环境对齐数据", "group": "data", "keywords": ["仿真数据", "真实数据", "对齐", "采集"], "relevance": 0.90},
        {"name": "开发域随机化与策略迁移方法", "group": "model", "keywords": ["域随机化", "迁移学习", "sim2real", "策略迁移"], "relevance": 0.98},
        {"name": "设计仿真到现实迁移评测流程", "group": "evaluation", "keywords": ["评测", "迁移", "鲁棒性", "验证"], "relevance": 0.95},
        {"name": "执行真实机器人回归测试与失效分析", "group": "evaluation", "keywords": ["回归测试", "失效", "真机", "测试"], "relevance": 0.92},
        {"name": "维护仿真资产与真实部署流水线", "group": "deployment", "keywords": ["仿真平台", "部署", "流水线", "资产"], "relevance": 0.85},
    ],
}

KNOWLEDGE_ROLE_TITLES: dict[str, dict[str, str]] = {
    "T1.02": {
        "data": "世界模型数据工程师",
        "model": "机器人世界模型工程师",
        "evaluation": "机器人世界模型评测工程师",
        "planning": "模型驱动规划工程师",
        "deployment": "具身模型推理部署工程师",
    },
    "T1.01": {
        "data": "VLA多模态数据工程师",
        "model": "具身VLA算法工程师",
        "evaluation": "VLA操作评测工程师",
        "deployment": "VLA推理优化工程师",
    },
    "T4.04": {
        "data": "Sim-to-Real数据工程师",
        "model": "机器人策略迁移工程师",
        "evaluation": "Sim-to-Real验证工程师",
        "deployment": "仿真部署流水线工程师",
    },
}


def _clean(value: Any) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def _sheet_records(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    headers = [_clean(v) for v in next(rows)]
    result = [dict(zip(headers, row)) for row in rows]
    workbook.close()
    return result


def import_taxonomy(conn, data_dir: Path) -> int:
    """技术主数据 → technologies：L2 实体 + L3 实体（别名为子级 L4 技术词，供 JD 召回）。"""
    path = data_dir / TAXONOMY_FILE
    l2_records = _sheet_records(path, "L2技术类")
    l3_records = _sheet_records(path, "L3技术点")
    l4_records = _sheet_records(path, "L4技术词")

    l4_by_l3: dict[str, list[str]] = {}
    for row in l4_records:
        l4_by_l3.setdefault(_clean(row.get("挂载L3编码")), []).append(_clean(row.get("技术词")))
    l3_by_l2: dict[str, list[dict[str, Any]]] = {}
    for row in l3_records:
        l3_by_l2.setdefault(_clean(row.get("L2编码")), []).append(row)

    entities: list[tuple] = []
    for row in l2_records:
        code = _clean(row.get("L2编码"))
        aliases: list[str] = []
        for child in l3_by_l2.get(code, []):
            aliases.extend(l4_by_l3.get(_clean(child.get("L3编码")), [])[:16])
        l1 = _clean(row.get("所属L1"))
        entities.append((
            code, _clean(row.get("技术类")), "L2", _clean(row.get("所属技术域")),
            _clean(row.get("定义")), "", json.dumps(aliases, ensure_ascii=False),
            _clean(row.get("归类别名正则")), l1.split(".")[0] if l1 else "",
        ))
    for row in l3_records:
        code = _clean(row.get("L3编码"))
        l1 = _clean(row.get("L1编码"))
        entities.append((
            code, _clean(row.get("L3标准名")), "L3", _clean(row.get("L1技术域")),
            _clean(row.get("备注")), _clean(row.get("L2编码")),
            json.dumps(l4_by_l3.get(code, [])[:24], ensure_ascii=False), "",
            l1.split(".")[0] if l1 else "",
        ))
    conn.executemany(
        "INSERT OR REPLACE INTO technologies"
        " (technology_id, standard_name, level, domain, definition, parent_id, aliases_json, regex, mapped_l1_code)"
        " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        entities,
    )
    return len(entities)


def import_milestones(conn, data_dir: Path) -> int:
    """里程碑 xlsx → milestones（附 MILESTONE_BRIDGE 桥接映射）。"""
    records = _sheet_records(data_dir / MILESTONES_FILE, "技术里程碑")
    rows = []
    for index, row in enumerate(records, start=1):
        raw_date = row.get("发生时间")
        if isinstance(raw_date, datetime):
            event_date = raw_date.date().isoformat()
        else:
            event_date = _clean(raw_date)
        category = _clean(row.get("涉及技术"))
        rows.append((
            f"EVENT-{index:04d}",
            _clean(row.get("事件名称")),
            _clean(row.get("事件描述")),
            event_date,
            _clean(row.get("来源")),
            category,
            _clean(row.get("事件类型")),
            json.dumps(MILESTONE_BRIDGE.get(category, []), ensure_ascii=False),
        ))
    conn.executemany(
        "INSERT OR REPLACE INTO milestones"
        " (event_id, name, description, event_date, source, technology_category, event_type, technology_links)"
        " VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        rows,
    )
    return len(rows)


def import_knowledge(conn) -> dict[str, int]:
    """算法知识层种子 → capabilities / tasks / role_titles（含 fallback 兜底模板）。"""
    ph = "{技术}"
    capability_rows = [(code, i["name"], i["object"], i["scenario"])
                       for code, items in KNOWLEDGE_CAPABILITIES.items() for i in items]
    capability_rows += [
        ("fallback", f"应用{ph}形成新的机器人能力", ph, "产业应用"),
        ("fallback", f"评估{ph}的可靠性", "技术能力", "测试评测"),
    ]
    conn.executemany(
        "INSERT OR REPLACE INTO capabilities (technology_code, name, object, scenario) VALUES (?, ?, ?, ?)",
        capability_rows,
    )

    task_rows = [(code, i["name"], i["group"], json.dumps(i["keywords"], ensure_ascii=False), i["relevance"])
                 for code, items in KNOWLEDGE_TASKS.items() for i in items]
    task_rows += [
        ("fallback", f"构建{ph}相关数据", "data", json.dumps(["数据", "采集", "标注"], ensure_ascii=False), 0.85),
        ("fallback", f"开发{ph}核心方法", "model", json.dumps(["开发", "算法", "模型"], ensure_ascii=False), 0.94),
        ("fallback", f"设计{ph}评测体系", "evaluation", json.dumps(["评测", "测试", "验证"], ensure_ascii=False), 0.90),
    ]
    conn.executemany(
        "INSERT OR REPLACE INTO tasks (technology_code, name, task_group, keywords_json, relevance) VALUES (?, ?, ?, ?, ?)",
        task_rows,
    )

    title_rows = [(code, group, title)
                  for code, mapping in KNOWLEDGE_ROLE_TITLES.items() for group, title in mapping.items()]
    conn.executemany(
        "INSERT OR REPLACE INTO role_titles (technology_code, task_group, title) VALUES (?, ?, ?)",
        title_rows,
    )
    return {"capabilities": len(capability_rows), "tasks": len(task_rows), "role_titles": len(title_rows)}


def run_import(conn, data_dir: Path | None = None) -> dict[str, Any]:
    data_dir = Path(data_dir or (config.ROOT / "data" / "raw"))
    db.init_db(conn)
    tech = import_taxonomy(conn, data_dir)
    milestones = import_milestones(conn, data_dir)
    knowledge = import_knowledge(conn)
    db.set_meta(conn, "emerging_data_imported_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    conn.commit()
    return {"technologies": tech, "milestones": milestones, "knowledge": knowledge}


def main() -> None:
    parser = argparse.ArgumentParser(description="技术演化域数据导入（幂等）")
    parser.add_argument("--data-dir", default=None, help="数据目录（默认 data/raw）")
    args = parser.parse_args()
    conn = db.connect()
    stats = run_import(conn, args.data_dir)
    conn.close()
    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
