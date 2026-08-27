from __future__ import annotations

import json
import hashlib
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
V4 = ROOT / "data/source/20260826/core/岗位信息v4_企业增强分析.xlsx"
SUMMARY = ROOT / "data/source/20260810/derived/job_keyword_summary_5bd0aecd.xlsx"
MATCHES = ROOT / "data/source/20260810/derived/job_keyword_matches_0d1c88e4_(1).xlsx"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize(value: object) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", clean(value).lower())


def content_hash(value: object) -> str:
    return hashlib.sha1(normalize(value).encode("utf-8")).hexdigest() if clean(value) else ""


def rows(path: Path, sheet_name: str | None = None):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    values = sheet.iter_rows(values_only=True)
    header = [clean(value) for value in next(values)]
    result = [dict(zip(header, row, strict=False)) for row in values]
    workbook.close()
    return result


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    v4_rows = rows(V4, "岗位信息v4_企业增强")
    summary_rows = rows(SUMMARY)
    summary_by_key: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    summary_by_jd: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in summary_rows:
        summary_by_key[(normalize(row.get("公司")), normalize(row.get("岗位")))].append(row)
        summary_by_jd[content_hash(row.get("JD描述"))].append(row)

    linked_summary_ids: dict[tuple[str, str], set[int]] = defaultdict(set)
    linked_v4 = set()
    ambiguous_v4 = set()
    exact_jd_v4 = set()
    exact_jd_summary_ids: dict[tuple[str, str], set[int]] = defaultdict(set)
    for index, row in enumerate(v4_rows):
        candidates = summary_by_key.get((normalize(row.get("公司")), normalize(row.get("岗位"))), [])
        if candidates:
            linked_v4.add(index)
            if len(candidates) > 1:
                ambiguous_v4.add(index)
            for candidate in candidates:
                linked_summary_ids[(clean(candidate.get("source_file")), clean(candidate.get("job_id")))].add(index)
        exact_jd_candidates = summary_by_jd.get(content_hash(row.get("清洗JD描述")), [])
        if exact_jd_candidates:
            exact_jd_v4.add(index)
            for candidate in exact_jd_candidates:
                exact_jd_summary_ids[(clean(candidate.get("source_file")), clean(candidate.get("job_id")))].add(index)

    matched_relations = Counter()
    mapped_v4_jobs = set()
    mapped_v4_l1: dict[int, set[str]] = defaultdict(set)
    mapped_v4_l2: dict[int, set[str]] = defaultdict(set)
    mapped_v4_l3: dict[int, set[str]] = defaultdict(set)
    mapped_v4_l4: dict[int, set[str]] = defaultdict(set)
    exact_jd_mapped_v4_jobs = set()
    match_workbook = load_workbook(MATCHES, read_only=True, data_only=True)
    sheet = match_workbook.active
    values = sheet.iter_rows(values_only=True)
    header = [clean(value) for value in next(values)]
    for values_row in values:
        row = dict(zip(header, values_row, strict=False))
        key = (clean(row.get("source_file")), clean(row.get("job_id")))
        targets = linked_summary_ids.get(key)
        exact_jd_targets = exact_jd_summary_ids.get(key)
        if not targets:
            targets = set()
        if exact_jd_targets:
            exact_jd_mapped_v4_jobs.update(exact_jd_targets)
        if not targets and not exact_jd_targets:
            continue
        for index in targets:
            mapped_v4_jobs.add(index)
            mapped_v4_l1[index].add(clean(row.get("L1编码")))
            mapped_v4_l2[index].add(clean(row.get("L2技术类")))
            mapped_v4_l3[index].add(clean(row.get("挂载L3名称")))
            mapped_v4_l4[index].add(clean(row.get("技术词(原始)")))
            matched_relations["relations"] += 1
    match_workbook.close()

    payload = {
        "v4_rows": len(v4_rows),
        "source_batches": Counter(clean(row.get("数据来源批次")) for row in v4_rows).most_common(),
        "original_occ_prefix": Counter(clean(row.get("原始occ_id"))[:1] for row in v4_rows).most_common(),
        "summary_rows": len(summary_rows),
        "summary_sources": Counter(clean(row.get("source_file")) for row in summary_rows).most_common(),
        "exact_company_title_linked_v4": len(linked_v4),
        "exact_company_title_rate": round(len(linked_v4) / max(1, len(v4_rows)), 4),
        "ambiguous_v4": len(ambiguous_v4),
        "exact_jd_linked_v4": len(exact_jd_v4),
        "exact_jd_linked_rate": round(len(exact_jd_v4) / max(1, len(v4_rows)), 4),
        "exact_jd_mapped_v4_jobs": len(exact_jd_mapped_v4_jobs),
        "mapped_v4_jobs": len(mapped_v4_jobs),
        "mapped_v4_rate": round(len(mapped_v4_jobs) / max(1, len(v4_rows)), 4),
        "mapped_relations": matched_relations["relations"],
        "average_l1_per_mapped_job": round(sum(map(len, mapped_v4_l1.values())) / max(1, len(mapped_v4_jobs)), 2),
        "average_l2_per_mapped_job": round(sum(map(len, mapped_v4_l2.values())) / max(1, len(mapped_v4_jobs)), 2),
        "average_l3_per_mapped_job": round(sum(map(len, mapped_v4_l3.values())) / max(1, len(mapped_v4_jobs)), 2),
        "average_l4_per_mapped_job": round(sum(map(len, mapped_v4_l4.values())) / max(1, len(mapped_v4_jobs)), 2),
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
