from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FILES = [
    ROOT / "data/source/20260810/core/技术词主数据_20260727.xlsx",
    ROOT / "data/source/20260810/derived/具身智能岗位_技术规范聚类分析_v2.xlsx",
    ROOT / "data/source/20260810/derived/job_keyword_matches_0d1c88e4_(1).xlsx",
    ROOT / "data/source/20260810/derived/job_keyword_summary_5bd0aecd.xlsx",
    ROOT / "data/source/20260826/core/岗位信息v4_企业增强分析.xlsx",
]


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def inspect_workbook(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        first_rows = []
        for _ in range(5):
            try:
                first_rows.append([clean(value) for value in next(rows)])
            except StopIteration:
                break
        header = first_rows[0] if first_rows else []
        samples = first_rows[1:4]
        sheets.append(
            {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "header": header,
                "samples": samples,
            }
        )
    workbook.close()
    return {"file": path.name, "sheets": sheets}


def inspect_runtime_db(path: Path) -> dict[str, object]:
    if not path.exists():
        return {"file": str(path), "exists": False}
    connection = sqlite3.connect(path)
    try:
        tables = [row[0] for row in connection.execute("select name from sqlite_master where type='table' order by name")]
        result: dict[str, object] = {"file": str(path), "exists": True, "tables": tables}
        candidates = [table for table in tables if "taxonomy" in table or "technology" in table or "job" in table]
        counts = {}
        schemas = {}
        for table in candidates:
            counts[table] = connection.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
            schemas[table] = [row[1] for row in connection.execute(f'PRAGMA table_info("{table}")')]
        result["counts"] = counts
        result["schemas"] = schemas
        return result
    finally:
        connection.close()


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    payload = {
        "workbooks": [inspect_workbook(path) for path in FILES if path.exists()],
        "runtime_db": inspect_runtime_db(ROOT / "data/runtime/job-capability-graph-runtime.db"),
        "dev_db": inspect_runtime_db(ROOT / "backend/.local/dev.db"),
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
