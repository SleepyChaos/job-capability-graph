# -*- coding: utf-8 -*-
"""
将权威的「技术词主数据 v1.1」+「科研助理 L2L3 分类结果」合并进
能力图谱（技术—岗位图谱 / 图谱二）的构建产物 data/processed/graph_build。

产出增量：
  1) L4 表面词(1872) 建为 technology 节点(level=L4)，belongs_to 到对应 L3。
  2) 把岗位集从 3718 扩展到 Coze Drive 全量 4655（新增 937 个 job 节点，
     公司能精确命中现有 org 时才补 posts_job 边）。
  3) 给全部 4655 岗位加 classified_to_L2 / classified_to_L3 直接分类边，
     边带 method(匹配方式) + hit_words(命中词) 证据，满足「每条边可回查 JD」。
  4) 原有 capability_family / capability 桥接边全部保留。

运行：python work/transform_tech_job_graph.py
（请在 git-bash 之外用原生 Windows python 调用，路径用 C:/... 形式）
"""
import json
import collections
import openpyxl

GB_DIR = r"C:/Users/10741/WorkBuddy/2026-08-14-11-44-53/job-capability-graph/data/processed/graph_build"
MASTER = r"C:/Users/10741/Desktop/新建文件夹 (2)/技术词/技术词主数据_20260727.xlsx"
CD = r"C:/Users/10741/Desktop/新建文件夹 (2)/岗位/技术分层/_Coze_Drive_科研助理_岗位_L2L3分类结果.xlsx"

nodes = json.load(open(f"{GB_DIR}/nodes.json", encoding="utf-8"))
edges = json.load(open(f"{GB_DIR}/edges.json", encoding="utf-8"))

code2id = {n.get("code"): n["id"] for n in nodes
           if n.get("type") == "technology" and n.get("code")}
existing_job_ids = {n["id"] for n in nodes if n.get("type") == "job"}
org_label2id = {o.get("label"): o["id"] for o in nodes
                if o.get("type") == "organization"}

new_nodes, new_edges = [], []

# ---------- 1) L4 表面词 ----------
wb = openpyxl.load_workbook(MASTER, read_only=True, data_only=True)
ws = wb["L4技术词"]
rows = list(ws.iter_rows(values_only=True))
mi = {h: i for i, h in enumerate(rows[0])}
l4_count = 0
for r in rows[1:]:
    word = r[mi["技术词"]]
    parent_l3 = r[mi["挂载L3编码"]]
    if not word or not parent_l3:
        continue
    parent_id = code2id.get(parent_l3)
    if not parent_id:
        continue
    nid = f"tech:L4:{l4_count:04d}"
    new_nodes.append({
        "id": nid, "type": "technology", "level": "L4",
        "label": str(word), "code": None,
        "l4_type": r[mi["L4类型"]],
        "parent_l3_code": parent_l3,
        "parent_l3_name": r[mi["挂载L3名称"]],
        "l2_code": r[mi["L2编码"]],
        "l2_name": r[mi["L2技术类"]],
        "l1_code": r[mi["L1编码"]],
        "orig_level": r[mi["原层级(留痕)"]],
        "cross_domain": r[mi["跨域调整"]],
        "hit_source": r[mi["命中来源"]],
    })
    new_edges.append({"source": nid, "target": parent_id,
                       "type": "belongs_to", "layer": "L4"})
    l4_count += 1

# ---------- 2) + 3) 扩展岗位 + 直接分类边 ----------
wb2 = openpyxl.load_workbook(CD, read_only=True, data_only=True)
ws2 = wb2["岗位明细"]
rows2 = list(ws2.iter_rows(values_only=True))
ci = {h: i for i, h in enumerate(rows2[0])}
data = rows2[1:]

added_jobs = classified_l2 = classified_l3 = 0
for r in data:
    occ = r[ci["occ_id"]]
    jid = f"job:{occ}"
    company = r[ci["公司"]]
    title = r[ci["岗位"]]
    l2code = r[ci["匹配L2编码"]]
    l3code = r[ci["匹配L3编码"]]
    method = r[ci["匹配方式"]]
    hit_words = r[ci["命中词"]]
    all_l2 = r[ci["所有命中L2"]]

    if jid not in existing_job_ids:
        new_nodes.append({
            "id": jid, "type": "job",
            "label": title, "title": title,
            "company": company,
            "career_dir": r[ci["职业方向"]],
            "career_kind": r[ci["职业种类"]],
            "skill_tags": r[ci["技能标签"]],
            "source": "CozeDrive_L2L3",
        })
        existing_job_ids.add(jid)
        added_jobs += 1
        if company and company in org_label2id:
            new_edges.append({"source": org_label2id[company],
                              "target": jid, "type": "posts_job"})

    l2set = set()
    if all_l2:
        for c in str(all_l2).split(";"):
            c = c.strip()
            if c:
                l2set.add(c)
    elif l2code:
        l2set.add(l2code)
    for c in l2set:
        tid = code2id.get(c)
        if tid:
            new_edges.append({"source": jid, "target": tid,
                               "type": "classified_to_L2",
                               "method": method, "hit_words": hit_words,
                               "is_primary": (c == l2code)})
            classified_l2 += 1
    if l3code:
        tid = code2id.get(l3code)
        if tid:
            new_edges.append({"source": jid, "target": tid,
                               "type": "classified_to_L3",
                               "method": method, "hit_words": hit_words})
            classified_l3 += 1

nodes.extend(new_nodes)
edges.extend(new_edges)

json.dump(nodes, open(f"{GB_DIR}/nodes.json", "w", encoding="utf-8"),
          ensure_ascii=False, indent=1)
json.dump(edges, open(f"{GB_DIR}/edges.json", "w", encoding="utf-8"),
          ensure_ascii=False, indent=1)

print("added L4 nodes:", l4_count)
print("added job nodes:", added_jobs)
print("classified_to_L2 edges:", classified_l2)
print("classified_to_L3 edges:", classified_l3)
print("TOTAL nodes:", len(nodes), " TOTAL edges:", len(edges))
print("node types:", dict(collections.Counter(x.get("type") for x in nodes)))
print("edge types:", dict(collections.Counter(e.get("type") for e in edges)))
