from __future__ import annotations

import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
V4 = ROOT / "data/source/20260826/core/岗位信息v4_企业增强分析.xlsx"
TECH = ROOT / "data/source/20260810/core/技术词主数据_20260727.xlsx"


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize(value: object) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", clean(value).lower())


def split_skills(value: object) -> list[str]:
    return [item.strip() for item in re.split(r"[;,，；|/]+", clean(value)) if item.strip()]


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    tech_workbook = load_workbook(TECH, read_only=True, data_only=True)
    l4_sheet = tech_workbook["L4技术词"]
    l4_rows = l4_sheet.iter_rows(values_only=True)
    l4_header = [clean(value) for value in next(l4_rows)]
    tech_by_norm = {
        normalize(row.get("技术词")): row
        for row in (dict(zip(l4_header, values, strict=False)) for values in l4_rows)
        if normalize(row.get("技术词"))
    }
    tech_workbook.close()

    workbook = load_workbook(V4, read_only=True, data_only=True)
    payload = {"sheets": []}
    for sheet in workbook.worksheets:
        values = sheet.iter_rows(values_only=True)
        header = [clean(value) for value in next(values)]
        if sheet.title not in {"岗位信息v4_企业增强", "岗位信息v4"}:
            payload["sheets"].append({"name": sheet.title, "rows": max(0, (sheet.max_row or 1) - 1), "header": header})
            continue
        rows = [dict(zip(header, row, strict=False)) for row in values]
        fields = [
            "职业方向", "职业种类", "岗位", "公司", "清洗JD描述", "技能标签", "occ_id",
            "源记录ID", "数据来源", "产业链层级", "产业链12类", "公司细分领域", "融资轮次",
            "公司所属地区", "公司总部城市", "企业库标准名称", "企业匹配方式", "企业匹配置信度",
            "企业属性补全状态",
        ]
        unique_skills = sorted({skill for row in rows for skill in split_skills(row.get("技能标签"))})
        matched_skills = [skill for skill in unique_skills if normalize(skill) in tech_by_norm]
        mapped_jobs = sum(
            any(normalize(skill) in tech_by_norm for skill in split_skills(row.get("技能标签")))
            for row in rows
        )
        payload["sheets"].append(
            {
                "name": sheet.title,
                "rows": len(rows),
                "header": header,
                "completeness": {
                    field: sum(bool(clean(row.get(field))) for row in rows)
                    for field in fields
                    if field in header
                },
                "source_distribution": Counter(clean(row.get("数据来源")) for row in rows).most_common(),
                "enterprise_status": Counter(clean(row.get("企业属性补全状态")) for row in rows).most_common(),
                "category_distribution": Counter(clean(row.get("职业种类")) for row in rows).most_common(),
                "technology_mapping": {
                    "unique_skill_tags": len(unique_skills),
                    "exact_l4_skill_tags": len(matched_skills),
                    "exact_l4_skill_rate": round(len(matched_skills) / max(1, len(unique_skills)), 4),
                    "jobs_with_exact_l4": mapped_jobs,
                    "job_rate": round(mapped_jobs / max(1, len(rows)), 4),
                    "examples": [
                        {
                            "skill": skill,
                            "l1": clean(tech_by_norm[normalize(skill)].get("L1编码")),
                            "l2": clean(tech_by_norm[normalize(skill)].get("L2技术类")),
                            "l3": clean(tech_by_norm[normalize(skill)].get("挂载L3名称")),
                        }
                        for skill in matched_skills[:20]
                    ],
                    "unmatched_examples": [skill for skill in unique_skills if normalize(skill) not in tech_by_norm][:30],
                },
                "first_row": {field: clean(rows[0].get(field)) for field in fields if field in header},
            }
        )
    workbook.close()
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
