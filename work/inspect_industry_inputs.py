from pathlib import Path
from collections import Counter
import json
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
FILES = [
    Path('C:/Users/10741/Documents/Codex/2026-08-26/17-ai-6-17-ai-markdown/outputs/岗位信息v4.xlsx'),
    Path('C:/Users/10741/Desktop/新建文件夹 (2)/企业库/具身智能企业数据_整合去重_完整.xlsx'),
    Path('C:/Users/10741/Desktop/新建文件夹 (2)/岗位/岗位信息v4_企业增强分析与图谱.xlsx'),
]
for path in FILES:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(json.dumps({'file': path.name, 'sheets': [(s.title, s.max_row, s.max_column) for s in wb]}, ensure_ascii=False))
    for sheet in wb:
        iterator = sheet.iter_rows(values_only=True)
        preview = [next(iterator, ()) for _ in range(3)]
        print(json.dumps({'sheet': sheet.title, 'header': preview[0], 'preview': [[str(v or '')[:90] for v in row] for row in preview[1:]] if '企业数据' in path.name else []}, ensure_ascii=False, default=str))
        if '企业数据' in path.name:
            headers = list(preview[0])
            rows = [dict(zip(headers, values)) for values in [*preview[1:], *iterator]]
            for field in headers:
                if any(token in str(field) for token in ['产业', '地区', '总部', '招聘', '融资']):
                    counts = Counter(str(r.get(field) or '') for r in rows)
                    print(json.dumps({'field': field, 'filled': sum(v for k, v in counts.items() if k), 'values': counts.most_common(45) if field not in ['官网岗位招聘链接', '猎聘招聘链接（https://www.liepin.com/company-jobs/xxx/）', '其他第三方招聘链接（boss；智联；脉脉…）'] else counts.most_common(3)}, ensure_ascii=False))
            for field in ['企业名称', '层级', '省份', '城市', '国家', '在聘岗位数量']:
                counts = Counter(str(r.get(field) or '') for r in rows)
                print(json.dumps({'check': field, 'unique': len(counts), 'blank': counts.get('', 0), 'values': counts.most_common(80) if field != '企业名称' else [x for x in counts.items() if x[1] > 1]}, ensure_ascii=False))
    wb.close()
