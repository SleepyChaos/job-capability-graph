from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SENSITIVE_PATH_PARTS = {"restricted"}


@dataclass(frozen=True)
class ColumnProfile:
    index: int
    header: str
    non_empty_count: int
    empty_count: int
    unique_count: int
    type_counts: dict[str, int]
    formula_count: int


@dataclass(frozen=True)
class SheetProfile:
    name: str
    state: str
    row_count: int
    column_count: int
    merged_range_count: int
    duplicate_data_row_count: int
    columns: list[ColumnProfile]


@dataclass(frozen=True)
class WorkbookProfile:
    relative_path: str
    file_name: str
    size_bytes: int
    sha256: str
    modified_at: str
    access_classification: str
    schema_fingerprint: str
    sheets: list[SheetProfile]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def value_type(value: Any) -> str:
    if value is None or value == "":
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    return "text"


def stable_value(value: Any) -> str:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value).strip()


def profile_sheet(worksheet: Any) -> SheetProfile:
    rows = list(worksheet.iter_rows(values_only=False))
    if not rows:
        return SheetProfile(worksheet.title, worksheet.sheet_state, 0, 0, 0, 0, [])

    max_column = max((len(row) for row in rows), default=0)
    header_cells = rows[0]
    headers = []
    for index in range(max_column):
        value = header_cells[index].value if index < len(header_cells) else None
        headers.append(stable_value(value) if value not in (None, "") else f"__column_{index + 1}")

    column_profiles: list[ColumnProfile] = []
    for index, header in enumerate(headers):
        types: Counter[str] = Counter()
        unique_values: set[str] = set()
        formula_count = 0
        for row in rows[1:]:
            cell = row[index] if index < len(row) else None
            value = cell.value if cell is not None else None
            kind = value_type(value)
            types[kind] += 1
            if kind != "empty":
                unique_values.add(stable_value(value))
            if cell is not None and cell.data_type == "f":
                formula_count += 1
        non_empty = sum(count for kind, count in types.items() if kind != "empty")
        empty = types.get("empty", 0)
        column_profiles.append(
            ColumnProfile(
                index=index + 1,
                header=header,
                non_empty_count=non_empty,
                empty_count=empty,
                unique_count=len(unique_values),
                type_counts=dict(sorted(types.items())),
                formula_count=formula_count,
            )
        )

    row_hashes: Counter[str] = Counter()
    for row in rows[1:]:
        values = [stable_value(cell.value) if cell.value not in (None, "") else "" for cell in row]
        if any(values):
            serialized_values = json.dumps(values, ensure_ascii=False).encode()
            fingerprint = hashlib.sha256(serialized_values).hexdigest()
            row_hashes[fingerprint] += 1
    duplicate_count = sum(count - 1 for count in row_hashes.values() if count > 1)

    return SheetProfile(
        name=worksheet.title,
        state=worksheet.sheet_state,
        row_count=max(len(rows) - 1, 0),
        column_count=max_column,
        merged_range_count=len(worksheet.merged_cells.ranges),
        duplicate_data_row_count=duplicate_count,
        columns=column_profiles,
    )


def profile_workbook(path: Path, source_root: Path) -> WorkbookProfile:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        sheets = [profile_sheet(sheet) for sheet in workbook.worksheets]
    finally:
        workbook.close()

    relative_path = path.relative_to(source_root).as_posix()
    schema_payload = [
        {
            "name": sheet.name,
            "row_count": sheet.row_count,
            "headers": [column.header for column in sheet.columns],
        }
        for sheet in sheets
    ]
    schema_fingerprint = hashlib.sha256(
        json.dumps(schema_payload, ensure_ascii=False, sort_keys=True).encode()
    ).hexdigest()
    classification = (
        "restricted" if SENSITIVE_PATH_PARTS.intersection(path.parts) else "project-internal"
    )
    return WorkbookProfile(
        relative_path=relative_path,
        file_name=path.name,
        size_bytes=path.stat().st_size,
        sha256=sha256_file(path),
        modified_at=datetime.fromtimestamp(path.stat().st_mtime).astimezone().isoformat(),
        access_classification=classification,
        schema_fingerprint=schema_fingerprint,
        sheets=sheets,
    )


def render_markdown(profiles: list[WorkbookProfile], source_root: Path) -> str:
    lines = [
        "# 原始 XLSX 数据准入画像",
        "",
        "> 此报告由只读画像工具生成，不包含单元格样例值或个人信息。",
        "",
        f"- 数据批次目录：`{portable_source_path(source_root)}`",
        f"- 工作簿数量：{len(profiles)}",
        f"- 工作表数量：{sum(len(profile.sheets) for profile in profiles)}",
        "",
        "## 文件清单",
        "",
        "| 文件 | 分级 | 大小（字节） | SHA-256 | 工作表数 |",
        "| --- | --- | ---: | --- | ---: |",
    ]
    for profile in profiles:
        lines.append(
            f"| `{profile.relative_path}` | {profile.access_classification} | "
            f"{profile.size_bytes} | `{profile.sha256}` | {len(profile.sheets)} |"
        )

    lines.extend(["", "## 工作表画像", ""])
    for profile in profiles:
        lines.extend([f"### {profile.file_name}", ""])
        lines.append(
            f"Schema 指纹：`{profile.schema_fingerprint}`；访问级别："
            f"`{profile.access_classification}`。"
        )
        lines.append("")
        lines.append("| 工作表 | 状态 | 数据行 | 列 | 合并区域 | 重复数据行 |")
        lines.append("| --- | --- | ---: | ---: | ---: | ---: |")
        for sheet in profile.sheets:
            lines.append(
                f"| {sheet.name} | {sheet.state} | {sheet.row_count} | {sheet.column_count} | "
                f"{sheet.merged_range_count} | {sheet.duplicate_data_row_count} |"
            )
        lines.append("")
        for sheet in profile.sheets:
            lines.extend([f"#### {sheet.name} 字段", ""])
            lines.append("| # | 字段 | 非空 | 空值 | 唯一值 | 类型分布 | 公式 |")
            lines.append("| ---: | --- | ---: | ---: | ---: | --- | ---: |")
            for column in sheet.columns:
                types = ", ".join(f"{key}:{value}" for key, value in column.type_counts.items())
                safe_header = column.header.replace("|", "\\|")
                lines.append(
                    f"| {column.index} | {safe_header} | {column.non_empty_count} | "
                    f"{column.empty_count} | {column.unique_count} | {types} | "
                    f"{column.formula_count} |"
                )
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def portable_source_path(source_root: Path) -> str:
    parts = source_root.parts
    if "data" in parts:
        return Path(*parts[parts.index("data") :]).as_posix()
    return source_root.name


def run(source_root: Path, output_dir: Path) -> list[WorkbookProfile]:
    workbook_paths = sorted(source_root.rglob("*.xlsx"))
    profiles = [profile_workbook(path, source_root) for path in workbook_paths]
    output_dir.mkdir(parents=True, exist_ok=True)
    json_payload = {
        "schema_version": "1.0",
        "generated_at": datetime.now().astimezone().isoformat(),
        "source_root": portable_source_path(source_root),
        "workbooks": [asdict(profile) for profile in profiles],
    }
    (output_dir / "xlsx_profile.json").write_text(
        json.dumps(json_payload, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (output_dir / "xlsx_profile.md").write_text(
        render_markdown(profiles, source_root), encoding="utf-8"
    )
    return profiles


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="只读分析项目原始 XLSX 的结构和质量。")
    parser.add_argument("--source-root", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    profiles = run(args.source_root.resolve(), args.output_dir.resolve())
    print(
        json.dumps(
            {
                "workbook_count": len(profiles),
                "sheet_count": sum(len(profile.sheets) for profile in profiles),
                "output_dir": str(args.output_dir.resolve()),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
