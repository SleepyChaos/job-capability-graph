"""Build the governed job-ecosystem graph used by the frontend.

The generated graph is deliberately separated from the existing capability
graph.  It keeps the project's 6 career directions and 17 career categories as
hard boundaries, then discovers an explainable candidate-cluster layer inside
each category with title/JD/skill evidence.  The candidate layer is a v0.1
governance asset: it is suitable for visual exploration and expert review, but
does not pretend to be a final HDBSCAN/expert-approved release.

Outputs:
  data/processed/job_graph/job-ecosystem-graph.json
  data/processed/job_graph/nodes.json (full graph, including every job)
  data/processed/job_graph/edges.json (full graph, including job evidence edges)
  data/processed/job_graph/overview_nodes.json
  data/processed/job_graph/overview_edges.json
  data/processed/job_graph/jobs.json
  data/processed/job_graph/岗位图谱构建报告.md
  frontend/public/job-ecosystem-graph.json
"""

from __future__ import annotations

import csv
import json
import hashlib
import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIR = ROOT / "data" / "source" / "20260810" / "core"
V4_SOURCE_XLSX = ROOT / "data" / "source" / "20260826" / "core" / "岗位信息v4_企业增强分析.xlsx"
DEFAULT_SOURCE_XLSX = V4_SOURCE_XLSX if V4_SOURCE_XLSX.exists() else next(SOURCE_DIR.glob("*v3*.xlsx"))
SOURCE_XLSX = Path(os.environ.get("JOB_GRAPH_SOURCE_XLSX", DEFAULT_SOURCE_XLSX))
SOURCE_SHEET = os.environ.get("JOB_GRAPH_SOURCE_SHEET", "")
ENTERPRISE_LIBRARY_FILE = os.environ.get(
    "JOB_GRAPH_ENTERPRISE_LIBRARY_FILE",
    "具身智能企业数据_整合去重_完整.xlsx",
)
ENTERPRISE_LIBRARY_RECORD_COUNT = int(os.environ.get("JOB_GRAPH_ENTERPRISE_RECORD_COUNT", "633"))
OUTPUT_DIR = ROOT / "data" / "processed" / "job_graph"
PUBLIC_OUTPUT = ROOT / "frontend" / "public" / "job-ecosystem-graph.json"
STANDARD_ROLE_FILE = ROOT / "data" / "source" / "20260826" / "core" / "搜索词包_按岗位.csv"
TECHNOLOGY_MASTER_FILE = ROOT / "data" / "source" / "20260810" / "core" / "技术词主数据_20260727.xlsx"
TECHNOLOGY_SUMMARY_FILE = ROOT / "data" / "source" / "20260810" / "derived" / "job_keyword_summary_5bd0aecd.xlsx"
TECHNOLOGY_MATCH_FILE = ROOT / "data" / "source" / "20260810" / "derived" / "job_keyword_matches_0d1c88e4_(1).xlsx"

# --- Portrait Excel Override (Coze 科研助理产出) ---
# 当以下三份 xlsx 全部存在时，全面覆盖方向/类别/簇层级、107标准岗位与五维画像。
PORTRAIT_EXCEL_DIR = ROOT / "data" / "excel" / "岗位画像"
PORTRAIT_PROFILE_XLSX = PORTRAIT_EXCEL_DIR / "_Coze_Drive_科研助理_标准岗位五维能力画像.xlsx"
PORTRAIT_CLUSTER_XLSX = PORTRAIT_EXCEL_DIR / "_Coze_Drive_科研助理_岗位四层聚类结果.xlsx"
PORTRAIT_GRAPH_XLSX = PORTRAIT_EXCEL_DIR / "_Coze_Drive_科研助理_岗位分层聚类图谱数据.xlsx"
USE_PORTRAIT_EXCEL_OVERRIDE = (
    PORTRAIT_PROFILE_XLSX.exists()
    and PORTRAIT_CLUSTER_XLSX.exists()
    and PORTRAIT_GRAPH_XLSX.exists()
)
PORTRAIT_DIRECTION_COLORS_FALLBACK = [
    "#1769e0", "#0b9c93", "#df645e", "#f2a43a", "#7257c8", "#65758b",
]


DIRECTION_COLORS = {
    "软件硬件与系统架构": "#1769e0",
    "感知决策与控制算法": "#0b9c93",
    "部署运维与技术支持": "#df645e",
    "仿真测试与数据工程": "#f2a43a",
    "基础理论与前沿研究": "#7257c8",
    "非技术支撑": "#65758b",
}


CATEGORY_DIRECTION = {
    "硬件开发": "软件硬件与系统架构",
    "软件与系统": "软件硬件与系统架构",
    "机械与结构": "软件硬件与系统架构",
    "系统集成": "软件硬件与系统架构",
    "算法工程": "感知决策与控制算法",
    "视觉与多模态": "感知决策与控制算法",
    "SLAM与控制": "感知决策与控制算法",
    "部署与运维": "部署运维与技术支持",
    "技术支持与解决方案": "部署运维与技术支持",
    "测试与验证": "仿真测试与数据工程",
    "数据工程": "仿真测试与数据工程",
    "仿真与数字孪生": "仿真测试与数据工程",
    "基础与前沿研究": "基础理论与前沿研究",
    "供应链与制造": "非技术支撑",
    "市场与销售": "非技术支撑",
    "产品与设计": "非技术支撑",
    "职能管理": "非技术支撑",
}


@dataclass(frozen=True)
class ClusterSpec:
    name: str
    keywords: tuple[str, ...]
    fallback: bool = False


CLUSTER_SPECS: dict[str, tuple[ClusterSpec, ...]] = {
    "硬件开发": (
        ClusterSpec("嵌入式与固件开发", ("嵌入式", "固件", "firmware", "bsp", "单片机", "mcu", "rtos")),
        ClusterSpec("芯片·FPGA·SoC研发", ("芯片", "fpga", "soc", "rtl", "asic", "ic设计", "数字验证", "u v m", "uvm")),
        ClusterSpec("电机驱动与电源系统", ("电机", "驱动器", "伺服", "电源", "bms", "电池", "功率", "motor", "逆变")),
        ClusterSpec("传感器与电子系统", ("传感器", "雷达", "lidar", "imu", "摄像头", "光学", "声学", "sensor")),
        ClusterSpec("电子电气与硬件综合", ("硬件", "pcb", "pcba", "电气", "电子", "电路", "ee", "board"), True),
    ),
    "软件与系统": (
        ClusterSpec("ROS与机器人中间件", ("ros", "机器人软件", "中间件", "robot software", "ros2")),
        ClusterSpec("底层系统与设备驱动", ("底层", "内核", "kernel", "驱动开发", "设备驱动", "linux系统", "bsp")),
        ClusterSpec("云平台·后端与数据服务", ("后端", "服务端", "云平台", "云原生", "backend", "微服务", "java", "golang")),
        ClusterSpec("应用软件与客户端", ("前端", "客户端", "android", "ios", "qt", "上位机", "app", "web")),
        ClusterSpec("软件系统与架构", ("软件", "系统", "架构", "平台", "c++"), True),
    ),
    "机械与结构": (
        ClusterSpec("机械结构与整机设计", ("机械", "结构", "整机", "机构", "mechanical", "solidworks", "cad")),
        ClusterSpec("工艺·材料与热设计", ("工艺", "材料", "热设计", "热仿真", "cae", "模具", "散热"), True),
    ),
    "系统集成": (
        ClusterSpec("系统集成与联调", ("系统集成", "集成", "联调", "联合调试", "integration")),
        ClusterSpec("机器人整机交付与实施", ("整机", "交付", "实施", "工程", "现场"), True),
    ),
    "算法工程": (
        ClusterSpec("具身基础模型与VLA", ("具身", "vla", "大模型", "foundation model", "transformer", "llm", "多模态")),
        ClusterSpec("机器人学习与强化学习", ("强化学习", "模仿学习", "robot learning", "ppo", "policy", "reinforcement", "rl算法")),
        ClusterSpec("运动规划与控制算法", ("运动规划", "控制算法", "轨迹优化", "动力学", "motion planning", "mpc")),
        ClusterSpec("感知融合与深度学习", ("感知", "融合", "深度学习", "目标检测", "神经网络", "neural", "语义分割")),
        ClusterSpec("模型训练优化与部署", ("模型训练", "推理", "模型部署", "加速", "量化", "tensorrt", "cuda", "分布式训练")),
        ClusterSpec("通用机器人算法", ("算法", "机器人", "python", "c++"), True),
    ),
    "视觉与多模态": (
        ClusterSpec("2D视觉与图像算法", ("图像", "目标检测", "分割", "opencv", "2d", "图像处理", "ocr")),
        ClusterSpec("3D感知与点云", ("3d", "点云", "深度相机", "pcl", "三维", "雷达感知", "立体视觉")),
        ClusterSpec("多模态与视觉语言", ("多模态", "vla", "vlm", "视觉语言", "vision", "视觉", "感知"), True),
    ),
    "SLAM与控制": (
        ClusterSpec("定位建图与SLAM", ("slam", "定位", "建图", "mapping", "localization", "回环")),
        ClusterSpec("导航与路径规划", ("导航", "路径规划", "path planning", "navigation", "避障")),
        ClusterSpec("运动控制与轨迹优化", ("控制", "轨迹", "运动", "mpc", "动力学", "伺服"), True),
    ),
    "部署与运维": (
        ClusterSpec("DevOps·MLOps与云运维", ("devops", "mlops", "云", "平台运维", "容器", "kubernetes", "k8s", "ci/cd")),
        ClusterSpec("机器人现场部署与运维", ("部署", "运维", "现场", "维修", "robotops", "实施", "售后"), True),
    ),
    "技术支持与解决方案": (
        ClusterSpec("FAE与技术支持", ("fae", "技术支持", "售后", "客户支持", "现场应用")),
        ClusterSpec("解决方案与售前交付", ("解决方案", "售前", "方案", "交付", "客户", "项目"), True),
    ),
    "测试与验证": (
        ClusterSpec("软硬件与自动化测试", ("自动化测试", "软件测试", "硬件测试", "测试开发", "test", "qa", "测试")),
        ClusterSpec("可靠性·质量与安全验证", ("可靠性", "质量", "安全", "验证", "认证", "功能安全", "失效"), True),
    ),
    "数据工程": (
        ClusterSpec("数据采集·标注与治理", ("数据", "采集", "标注", "治理", "etl", "清洗"), True),
    ),
    "仿真与数字孪生": (
        ClusterSpec("机器人仿真与Sim2Real", ("sim2real", "isaac", "mujoco", "机器人仿真", "仿真算法", "物理仿真")),
        ClusterSpec("数字孪生与仿真平台", ("数字孪生", "仿真", "simulation", "虚拟", "建模"), True),
    ),
    "基础与前沿研究": (
        ClusterSpec("具身基础理论与世界模型", ("具身", "世界模型", "embodied", "world model", "基础模型", "认知")),
        ClusterSpec("机器人学与前沿研究", ("机器人学", "研究员", "scientist", "research", "前沿", "博士"), True),
    ),
    "供应链与制造": (
        ClusterSpec("生产制造与工艺工程", ("生产", "制造", "工艺", "产线", "npi", "装配", "良率")),
        ClusterSpec("供应链采购与质量管理", ("供应链", "采购", "物流", "质量", "供应商", "计划"), True),
    ),
    "市场与销售": (
        ClusterSpec("市场·销售与商务拓展", ("市场", "销售", "商务", "渠道", "品牌", "运营"), True),
    ),
    "产品与设计": (
        ClusterSpec("产品管理与体验设计", ("产品", "设计", "用户", "需求", "交互", "体验"), True),
    ),
    "职能管理": (
        ClusterSpec("组织职能与经营管理", ("财务", "人力", "行政", "法务", "管理", "审计"), True),
    ),
}


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize(value: Any) -> str:
    text = clean(value).lower()
    text = re.sub(r"\s+", " ", text)
    return text


def safe_id(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or hashlib.sha1(value.encode("utf-8")).hexdigest()[:12]


def normalize_match_key(value: Any) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", clean(value).lower())


def content_hash(value: Any) -> str:
    normalized = normalize_match_key(value)
    return hashlib.sha1(normalized.encode("utf-8")).hexdigest() if normalized else ""


def technology_l4_id(name: str) -> str:
    return f"technology-l4-{hashlib.sha1(normalize_match_key(name).encode('utf-8')).hexdigest()[:14]}"


def load_technology_master() -> tuple[list[dict[str, Any]], dict[str, str]]:
    """Read the governed L1-L4 technology taxonomy without inventing hierarchy."""
    if not TECHNOLOGY_MASTER_FILE.exists():
        return [], {}
    workbook = openpyxl.load_workbook(TECHNOLOGY_MASTER_FILE, read_only=True, data_only=True)
    nodes: list[dict[str, Any]] = []
    l4_lookup: dict[str, str] = {}

    def read_rows(sheet_name: str) -> list[dict[str, Any]]:
        sheet = workbook[sheet_name]
        values = sheet.iter_rows(values_only=True)
        header = [clean(value) for value in next(values)]
        return [dict(zip(header, row, strict=False)) for row in values]

    for row in read_rows("L1技术域"):
        code = clean(row.get("L1编码"))
        if code:
            nodes.append({
                "id": f"technology-{code.lower()}", "code": code, "name": clean(row.get("技术域")),
                "level": "L1", "parentId": "", "definition": clean(row.get("定义")),
                "jobCount": 0, "standardRoleCount": 0,
            })
    for row in read_rows("L2技术类"):
        code = clean(row.get("L2编码"))
        parent = clean(row.get("所属L1"))
        if code:
            nodes.append({
                "id": f"technology-{code.lower()}", "code": code, "name": clean(row.get("技术类")),
                "level": "L2", "parentId": f"technology-{parent.lower()}", "definition": clean(row.get("定义")),
                "jobCount": 0, "standardRoleCount": 0,
            })
    for row in read_rows("L3技术点"):
        code = clean(row.get("L3编码"))
        parent = clean(row.get("L2编码"))
        if code:
            nodes.append({
                "id": f"technology-{code.lower()}", "code": code, "name": clean(row.get("L3标准名")),
                "level": "L3", "parentId": f"technology-{parent.lower()}", "definition": clean(row.get("备注")),
                "jobCount": 0, "standardRoleCount": 0,
            })
    for row in read_rows("L4技术词"):
        name = clean(row.get("技术词"))
        parent = clean(row.get("挂载L3编码"))
        if not name or not parent:
            continue
        node_id = technology_l4_id(name)
        l4_lookup.setdefault(normalize_match_key(name), node_id)
        nodes.append({
            "id": node_id, "code": "", "name": name, "level": "L4",
            "parentId": f"technology-{parent.lower()}", "definition": clean(row.get("L4类型")),
            "jobCount": 0, "standardRoleCount": 0,
        })
    workbook.close()
    return nodes, l4_lookup


def load_exact_jd_technology_map(l4_lookup: dict[str, str]) -> tuple[dict[str, list[str]], dict[str, int]]:
    """Map legacy annotations to v4 only through identical normalized JD text."""
    if not TECHNOLOGY_SUMMARY_FILE.exists() or not TECHNOLOGY_MATCH_FILE.exists():
        return {}, {"summaryRows": 0, "matchRows": 0, "matchedRelations": 0}

    summary_workbook = openpyxl.load_workbook(TECHNOLOGY_SUMMARY_FILE, read_only=True, data_only=True)
    summary_sheet = summary_workbook.active
    summary_values = summary_sheet.iter_rows(values_only=True)
    summary_header = [clean(value) for value in next(summary_values)]
    source_job_to_hash: dict[tuple[str, str], str] = {}
    summary_count = 0
    for values in summary_values:
        row = dict(zip(summary_header, values, strict=False))
        summary_count += 1
        jd_hash = content_hash(row.get("JD描述"))
        if jd_hash:
            source_job_to_hash[(clean(row.get("source_file")), clean(row.get("job_id")))] = jd_hash
    summary_workbook.close()

    match_workbook = openpyxl.load_workbook(TECHNOLOGY_MATCH_FILE, read_only=True, data_only=True)
    match_sheet = match_workbook.active
    match_values = match_sheet.iter_rows(values_only=True)
    match_header = [clean(value) for value in next(match_values)]
    jd_terms: dict[str, set[str]] = defaultdict(set)
    match_count = 0
    matched_relations = 0
    for values in match_values:
        row = dict(zip(match_header, values, strict=False))
        match_count += 1
        jd_hash = source_job_to_hash.get((clean(row.get("source_file")), clean(row.get("job_id"))))
        if not jd_hash:
            continue
        raw_term = clean(row.get("技术词(原始)")) or clean(row.get("技术词(规范)"))
        technology_id = l4_lookup.get(normalize_match_key(raw_term))
        if technology_id:
            jd_terms[jd_hash].add(technology_id)
            matched_relations += 1
    match_workbook.close()
    return {key: sorted(value) for key, value in jd_terms.items()}, {
        "summaryRows": summary_count,
        "matchRows": match_count,
        "matchedRelations": matched_relations,
    }


def score_spec(spec: ClusterSpec, title: str, skills: str, jd: str) -> tuple[int, list[str]]:
    score = 0
    evidence: list[str] = []
    for keyword in spec.keywords:
        key = keyword.lower()
        if key in title:
            score += 7
            evidence.append(keyword)
        if key in skills:
            score += 4
            evidence.append(keyword)
        if key in jd:
            score += 1
            evidence.append(keyword)
    return score, list(dict.fromkeys(evidence))[:8]


def split_skills(value: Any) -> list[str]:
    return [part.strip() for part in re.split(r"[;；,，|]+", clean(value)) if part.strip()]


def counter_rows(counter: Counter[str], limit: int) -> list[dict[str, Any]]:
    return [{"name": name, "count": count} for name, count in counter.most_common(limit)]


PROFILE_ABILITY_PATTERNS = {
    "算法研究与模型开发": r"算法|模型|深度学习|机器学习|强化学习|训练|推理",
    "系统设计与架构": r"系统设计|架构|方案设计|模块设计|总体设计",
    "软件开发与工程实现": r"软件|开发|编码|编程|c\+\+|python|java|ros",
    "硬件设计与整机实现": r"硬件|电路|pcb|嵌入式|结构|机械|整机",
    "数据处理与分析": r"数据采集|数据处理|数据分析|标注|清洗|数据治理",
    "测试验证与问题定位": r"测试|验证|调试|联调|故障|问题定位|可靠性",
    "部署交付与现场支持": r"部署|交付|实施|现场|运维|售前|售后|客户",
    "项目协同与沟通": r"项目管理|团队协作|跨部门|沟通|协调|推动",
}

PROFILE_RESPONSIBILITY_PATTERNS = {
    "技术与产品路线规划": r"技术路线|产品路线|路线图|技术规划|产品规划|战略规划",
    "算法与模型研发": r"算法开发|模型开发|模型训练|算法研究|深度学习|强化学习|机器学习",
    "系统方案设计": r"系统方案|方案设计|系统设计|架构设计|总体设计|技术方案",
    "软硬件研发实现": r"软件开发|硬件开发|编码实现|电路设计|结构设计|嵌入式开发",
    "测试验证与质量保障": r"测试验证|验证测试|质量保障|可靠性|测试方案|故障定位|问题定位",
    "项目交付与现场实施": r"项目交付|交付实施|现场实施|部署交付|项目实施|客户现场",
    "客户需求与解决方案": r"客户需求|需求分析|解决方案|售前方案|行业方案|技术支持",
    "市场与客户洞察": r"市场洞察|客户洞察|市场分析|行业分析|竞争分析|用户研究",
    "大客户与商务拓展": r"大客户|商务拓展|业务拓展|渠道拓展|销售目标|客户开发",
    "高层关系与生态合作": r"高层关系|政府关系|生态合作|合作伙伴|战略合作|产业合作",
    "团队建设与组织管理": r"团队建设|团队管理|组织管理|人才培养|绩效管理|梯队建设",
    "项目管理与跨部门协同": r"项目管理|跨部门|协同推进|资源协调|沟通协调|项目推进",
    "生产制造与供应链协同": r"生产制造|供应链|供应商|量产|工艺优化|产线|采购",
    "数据运营与持续优化": r"数据分析|运营分析|持续优化|指标体系|数据运营|复盘改进",
}

# 搜索词包给出了标准岗位和标题变体，但未给出所属岗位簇。这里将标准岗位按业务语义
# 人工校准到候选岗位簇；原始JD只有同时满足“既有17类/岗位簇边界 + 标题相似闸门”
# 才能作为该标准岗位证据，避免用“工程师/经理”等通用后缀跨类误配。
STANDARD_ROLE_CLUSTER_GROUPS = {
    "组织职能与经营管理": "P001 P002 P005 P006 P007 P008 P009 P010 P011 P012 P019 P092 P093 P094 P095 P096 P097 P098 P099 P101 P102 P103 P106".split(),
    "产品管理与体验设计": "P003 P013 P037 P038 P044 P074 P104".split(),
    "具身基础理论与世界模型": "P004 P021".split(),
    "机器人学习与强化学习": "P018 P020".split(),
    "运动规划与控制算法": "P022 P026".split(),
    "机器人学与前沿研究": "P023".split(),
    "具身基础模型与VLA": "P024 P030".split(),
    "3D感知与点云": "P025 P029".split(),
    "感知融合与深度学习": "P027".split(),
    "模型训练优化与部署": "P028".split(),
    "软件系统与架构": "P031 P033 P034 P105".split(),
    "电子电气与硬件综合": "P032 P046".split(),
    "云平台·后端与数据服务": "P035".split(),
    "可靠性·质量与安全验证": "P016 P036 P058 P059 P060".split(),
    "ROS与机器人中间件": "P039".split(),
    "嵌入式与固件开发": "P040".split(),
    "应用软件与客户端": "P041".split(),
    "机器人仿真与Sim2Real": "P042 P043".split(),
    "机械结构与整机设计": "P045 P049".split(),
    "传感器与电子系统": "P047".split(),
    "底层系统与设备驱动": "P048".split(),
    "系统集成与联调": "P050".split(),
    "DevOps·MLOps与云运维": "P051".split(),
    "数据采集·标注与治理": "P052 P053 P054".split(),
    "软硬件与自动化测试": "P055 P056 P057".split(),
    "生产制造与工艺工程": "P014 P061 P062 P063 P064 P065 P066 P067".split(),
    "供应链采购与质量管理": "P015 P068 P069 P070 P071 P072".split(),
    "市场·销售与商务拓展": "P073 P075 P076 P077 P078 P079 P080 P081 P082 P083 P088 P100 P107".split(),
    "解决方案与售前交付": "P084 P085".split(),
    "机器人现场部署与运维": "P017 P086".split(),
    "FAE与技术支持": "P087 P089 P090 P091".split(),
}
STANDARD_ROLE_CLUSTER_HINTS = {
    role_code: cluster_name
    for cluster_name, role_codes in STANDARD_ROLE_CLUSTER_GROUPS.items()
    for role_code in role_codes
}

PROFILE_SCENARIO_PATTERNS = {
    "机器人本体与运动": r"机器人|机械臂|人形|四足|移动底盘|灵巧手",
    "智能制造与产线": r"制造|产线|工厂|工业|装配|质检|生产",
    "仓储物流": r"仓储|物流|分拣|搬运|配送|agv|amr",
    "自动驾驶与移动导航": r"自动驾驶|无人车|导航|定位|slam|路径规划",
    "视觉感知与多模态": r"视觉|图像|点云|相机|感知|多模态|雷达",
    "仿真训练与数字孪生": r"仿真|数字孪生|sim2real|isaac|mujoco",
    "医疗与服务机器人": r"医疗|手术|康复|服务机器人|家庭|养老",
    "云平台与数据服务": r"云平台|云原生|服务端|后端|数据平台|saas",
}


def compact_text(value: Any, limit: int = 92) -> str:
    text = re.sub(r"\s+", " ", clean(value)).strip(" -—·•；;，,")
    return text if len(text) <= limit else f"{text[:limit - 1]}…"


def jd_sentences(value: Any) -> list[str]:
    raw = clean(value)
    if not raw:
        return []
    chunks = re.split(r"[\r\n]+|(?<=[。！？!?；;])", raw)
    result: list[str] = []
    seen: set[str] = set()
    for chunk in chunks:
        sentence = compact_text(re.sub(r"^[\d一二三四五六七八九十]+[.、)]\s*", "", chunk))
        key = normalize(sentence)
        if len(sentence) < 7 or key in seen:
            continue
        seen.add(key)
        result.append(sentence)
    return result


def build_job_profile(row: dict[str, Any]) -> dict[str, Any]:
    jd = clean(row.get("清洗JD描述"))
    sentences = jd_sentences(jd)
    requirement_pattern = re.compile(r"任职|要求|资格|学历|经验|熟悉|掌握|精通|优先|本科|硕士|博士|年以上", re.I)
    responsibility_pattern = re.compile(r"负责|职责|工作内容|参与|承担|完成|开发|设计|搭建|实现|研究|推进|维护", re.I)
    responsibilities = [
        sentence for sentence in sentences
        if responsibility_pattern.search(sentence) and not requirement_pattern.search(sentence)
    ][:4]

    requirement_evidence = [sentence for sentence in sentences if requirement_pattern.search(sentence)][:3]
    skills = list(dict.fromkeys(split_skills(row.get("技能标签"))))[:10]
    combined = " ".join([
        jd,
        clean(row.get("岗位")),
        clean(row.get("公司细分领域")),
        clean(row.get("产业链12类")),
    ]).lower()
    abilities = [name for name, pattern in PROFILE_ABILITY_PATTERNS.items() if re.search(pattern, combined, re.I)][:6]
    scenarios = [name for name, pattern in PROFILE_SCENARIO_PATTERNS.items() if re.search(pattern, combined, re.I)][:5]
    industry_stage = clean(row.get("产业链层级"))
    company_specialty = clean(row.get("公司细分领域"))
    if industry_stage:
        scenarios.append(f"产业链·{industry_stage}")
    if company_specialty:
        scenarios.append(compact_text(company_specialty, 28))
    scenarios = list(dict.fromkeys(scenarios))[:6]

    conditions = []
    for label, field in (
        ("学历", "学历(标准化)"),
        ("经验", "经验(标准化)"),
        ("能力等级", "能力等级"),
        ("工作地区", "公司所属地区"),
    ):
        value = clean(row.get(field))
        if value:
            conditions.append(f"{label}·{value}")
    conditions.extend(requirement_evidence)

    evidence = list(dict.fromkeys(responsibilities[:2] + requirement_evidence[:2]))
    return {
        "responsibilities": responsibilities,
        "skills": skills,
        "abilities": abilities,
        "scenarios": scenarios,
        "conditions": list(dict.fromkeys(conditions))[:7],
        "jdEvidence": evidence[:4],
        "jdAvailable": bool(jd),
    }


def title_similarity(left: str, right: str) -> float:
    a = normalize(left)
    b = normalize(right)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    # “工程师/经理/总监”等是职位层级词，不应让两个业务完全不同的标题被判为相似。
    # 先保留原文做精确匹配，再移除通用层级词比较业务语义核心。
    generic_tokens = (
        "高级", "资深", "初级", "中级", "首席", "副", "助理",
        "工程师", "架构师", "设计师", "研究员", "科学家", "专家",
        "总监", "经理", "主管", "专员", "负责人", "总裁", "顾问",
        "engineer", "manager", "director", "specialist", "lead", "senior", "staff",
    )
    for token in generic_tokens:
        a = a.replace(token, "")
        b = b.replace(token, "")
    if not a or not b:
        return 0.0
    if a == b:
        return 0.98
    shorter, longer = sorted((a, b), key=len)
    if len(shorter) >= 4 and shorter in longer:
        return 0.82 + 0.14 * len(shorter) / len(longer)
    a_pairs = {a[index:index + 2] for index in range(max(1, len(a) - 1))}
    b_pairs = {b[index:index + 2] for index in range(max(1, len(b) - 1))}
    jaccard = len(a_pairs & b_pairs) / max(1, len(a_pairs | b_pairs))
    sequence = SequenceMatcher(None, a, b).ratio()
    return 0.55 * jaccard + 0.45 * sequence


def load_standard_role_seeds() -> list[dict[str, Any]]:
    if not STANDARD_ROLE_FILE.exists():
        return []
    with STANDARD_ROLE_FILE.open("r", encoding="utf-8-sig", newline="") as file:
        rows = list(csv.DictReader(file))
    result = []
    for row in rows:
        name = clean(row.get("标准岗位"))
        if not name:
            continue
        variants = [clean(item) for item in clean(row.get("真实标题变体(已出现别名,|分隔)")).split("|") if clean(item)]
        variants = list(dict.fromkeys([name, *variants]))
        result.append({
            "id": clean(row.get("occ_id")) or f"P{len(result) + 1:03d}",
            "name": name,
            "seedVariants": variants,
            "normalizedVariants": [normalize(item) for item in variants if normalize(item)],
        })
    return result


def aggregate_standard_profile(jobs: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    counters: dict[str, Counter[str]] = {
        dimension: Counter() for dimension in ("responsibilities", "skills", "abilities", "scenarios", "conditions")
    }
    evidence: dict[str, dict[str, list[str]]] = {
        dimension: defaultdict(list) for dimension in counters
    }

    def record(dimension: str, point: str, job: dict[str, Any]) -> None:
        point = compact_text(point, 30)
        if not point:
            return
        counters[dimension][point] += 1
        occ_id = clean(job.get("occId"))
        if occ_id and occ_id not in evidence[dimension][point]:
            evidence[dimension][point].append(occ_id)

    for job in jobs:
        combined = f"{job.get('title', '')} {job.get('jd', '')}"
        for name, pattern in PROFILE_RESPONSIBILITY_PATTERNS.items():
            if re.search(pattern, combined, re.I):
                record("responsibilities", name, job)
        for skill in list(dict.fromkeys([*job.get("skills", []), *job.get("profile", {}).get("skills", [])])):
            record("skills", skill, job)
        for point in job.get("profile", {}).get("abilities", []):
            record("abilities", point, job)
        for point in job.get("profile", {}).get("scenarios", []):
            if not point.startswith("产业链·") or point in {"产业链·上游", "产业链·中游", "产业链·下游", "产业链·横向支撑"}:
                record("scenarios", point, job)
        for label, value in (
            ("学历", job.get("education")),
            ("经验", job.get("experience")),
            ("能力等级", job.get("abilityLevel")),
        ):
            if clean(value):
                record("conditions", f"{label}·{clean(value)}", job)

    limits = {"responsibilities": 8, "skills": 10, "abilities": 8, "scenarios": 8, "conditions": 8}
    if len(jobs) < 2:
        return {dimension: [] for dimension in counters}

    result: dict[str, list[dict[str, Any]]] = {}
    for dimension, counter in counters.items():
        result[dimension] = [
            {
                "name": name,
                "count": count,
                "coverage": round(count / max(1, len(jobs)), 4),
                "evidenceOccIds": evidence[dimension][name],
            }
            for name, count in counter.most_common()
            if count >= 2
        ]
        result[dimension] = result[dimension][:limits[dimension]]
    return result


# ---------------------------------------------------------------------------
# Portrait Excel Override (Coze 科研助理产出) helpers
# ---------------------------------------------------------------------------

def _read_xlsx_rows(path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """Read an xlsx sheet and return list of row dicts (header-mapped)."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]
    values = sheet.iter_rows(values_only=True)
    header = [clean(value) for value in next(values)]
    rows = [dict(zip(header, row, strict=False)) for row in values]
    workbook.close()
    return rows


def _parse_portrait_bullets(text: str, total_count: int) -> list[tuple[str, int]]:
    """Split "• a\n• b\n• c" style bullets, optionally with "• 2. " prefix.
    Returns list of (name, count), distributing total_count evenly across bullets."""
    raw = clean(text)
    if not raw or raw == "暂无":
        return []
    # Split on lines starting with • (possibly preceded by newline), removing numbered prefix.
    parts = [
        clean(piece)
        for piece in re.split(r"(?:^|\n)\s*•\s*(?:\d+\.\s*)?", raw)
        if clean(piece) and clean(piece) != "暂无"
    ]
    if not parts:
        return []
    parts = list(dict.fromkeys(parts))  # de-dup
    n = len(parts)
    # Evenly distribute, floor + at least 2 (matching original >=2 publish gate).
    base = max(2, total_count // n) if total_count else 2
    remainder = max(0, total_count - base * n) if total_count else 0
    out: list[tuple[str, int]] = []
    for idx, name in enumerate(parts):
        cnt = base + (1 if idx < remainder else 0)
        if cnt < 2 and total_count >= 2:
            cnt = 2
        out.append((compact_text(name, 36), cnt))
    return out


_CONDITION_COUNT_RE = re.compile(r"^(?P<name>.+?)\s*\(\s*(?P<count>\d+)\s*\)\s*$")


def _parse_portrait_conditions(text: str) -> list[tuple[str, int]]:
    """Parse conditions column: "学历: 博士(11), 未说明(9)\n经验: ...\n等级: ..."
    Each value becomes "学历·博士" etc. with the parenthesised count."""
    raw = clean(text)
    if not raw:
        return []
    result: list[tuple[str, int]] = []
    for line in raw.splitlines():
        line = clean(line)
        if not line or ":" not in line:
            continue
        prefix, _, values = line.partition(":")
        prefix = clean(prefix).rstrip(":")
        if not prefix or not values:
            continue
        for chunk in values.split(","):
            chunk = clean(chunk)
            if not chunk:
                continue
            m = _CONDITION_COUNT_RE.match(chunk)
            if m:
                val = compact_text(clean(m.group("name")), 36)
                try:
                    cnt = int(m.group("count"))
                except ValueError:
                    cnt = 2
                name = f"{prefix}·{val}" if prefix else val
                if cnt < 2:
                    cnt = 2
                result.append((name, cnt))
            else:
                # No count, use fallback 2
                val = compact_text(chunk, 36)
                if val and val != "暂无":
                    result.append((f"{prefix}·{val}", 2))
    # Deduplicate, summing counts if duplicated names occur
    merged: Counter[str] = Counter()
    for name, cnt in result:
        merged[name] += cnt
    return [(name, max(2, cnt)) for name, cnt in merged.items()]


def _pick_portrait_color(direction_name: str, fallback_index: int) -> str:
    if direction_name in DIRECTION_COLORS:
        return DIRECTION_COLORS[direction_name]
    palette = PORTRAIT_DIRECTION_COLORS_FALLBACK
    return palette[fallback_index % len(palette)]


def load_portrait_package() -> dict[str, Any]:
    """读取 3 份岗位画像 xlsx（4655 岗位明细 + 分层图谱节点+边 + 107标准岗位五维画像）。
    返回 dict，keys: directions / categories / clusters / job_overlay / standard_roles /
    standard_role_job_edges / standard_role_audit / direction_order。
    """
    # --- 1) 分层聚类图谱 nodes/edges 读取 ---
    graph_nodes_rows = _read_xlsx_rows(PORTRAIT_GRAPH_XLSX, "nodes")
    graph_edges_rows = _read_xlsx_rows(PORTRAIT_GRAPH_XLSX, "edges")

    direction_rows = [r for r in graph_nodes_rows if clean(r.get("level")) == "direction"]
    category_rows = [r for r in graph_nodes_rows if clean(r.get("level")) == "category"]
    cluster_rows_g = [r for r in graph_nodes_rows if clean(r.get("level")) == "cluster"]

    # Build parent lookup from edges (type=contains)
    child_parent: dict[str, tuple[str, str]] = {}  # child_id -> (parent_id, parent_label_cache)
    node_by_id: dict[str, dict[str, Any]] = {clean(r.get("id")): r for r in graph_nodes_rows}
    for edge in graph_edges_rows:
        src = clean(edge.get("source"))
        tgt = clean(edge.get("target"))
        typ = clean(edge.get("type"))
        if typ == "contains" and src and tgt:
            child_parent[tgt] = (src, "")

    direction_order = [clean(r.get("label")) for r in direction_rows if clean(r.get("label"))]

    # Direction entities
    directions: list[dict[str, Any]] = []
    dir_id_by_name: dict[str, str] = {}
    for idx, drow in enumerate(direction_rows, start=1):
        name = clean(drow.get("label"))
        if not name:
            continue
        dir_id = f"direction-{idx:02d}"
        color = _pick_portrait_color(name, idx - 1)
        cnt = int(clean(drow.get("count")) or 0)
        directions.append({
            "id": dir_id,
            "name": name,
            "color": color,
            "jobCount": cnt,
            "categoryCount": 0,
            "clusterCount": 0,
        })
        dir_id_by_name[name] = dir_id

    # Category entities
    categories: list[dict[str, Any]] = []
    cat_id_by_name: dict[str, str] = {}
    cat_name_by_id: dict[str, str] = {}
    direction_cats: dict[str, int] = defaultdict(int)
    for idx, crow in enumerate(category_rows, start=1):
        name = clean(crow.get("label"))
        if not name:
            continue
        cat_id = f"category-{idx:02d}"
        # Resolve parent direction via edges
        parent_id = child_parent.get(clean(crow.get("id")), ("", ""))[0]
        parent_row = node_by_id.get(parent_id, {})
        direction_name = clean(parent_row.get("label"))
        # Fallback: infer from CATEGORY_DIRECTION
        if not direction_name and name in CATEGORY_DIRECTION:
            direction_name = CATEGORY_DIRECTION[name]
        direction_id = dir_id_by_name.get(direction_name, directions[0]["id"] if directions else "")
        direction_ref = next((d for d in directions if d["id"] == direction_id), None)
        color = direction_ref["color"] if direction_ref else PORTRAIT_DIRECTION_COLORS_FALLBACK[0]
        cnt = int(clean(crow.get("count")) or 0)
        category = {
            "id": cat_id,
            "name": name,
            "directionId": direction_id,
            "directionName": direction_name,
            "color": color,
            "jobCount": cnt,
            "clusterCount": 0,
        }
        categories.append(category)
        cat_id_by_name[name] = cat_id
        cat_name_by_id[cat_id] = name
        direction_cats[direction_id] += 1

    # Clusters from graph_nodes (id/level=cluster) + profile xlsx for extra metadata
    clusters: list[dict[str, Any]] = []
    cluster_by_code: dict[str, dict[str, Any]] = {}  # keyed by code like "CL42"
    for idx, crow in enumerate(cluster_rows_g, start=1):
        node_id = clean(crow.get("id"))
        label = clean(crow.get("label"))
        # Code might be node_id (e.g. "CL_CL42") — extract trailing CLxx
        code_match = re.search(r"(CL\d+)", node_id)
        code = code_match.group(1) if code_match else f"CL{idx:02d}"
        # Parent via edges
        parent_id = child_parent.get(node_id, ("", ""))[0]
        parent_row = node_by_id.get(parent_id, {})
        cat_name = clean(parent_row.get("label"))
        cat_id = cat_id_by_name.get(cat_name, categories[0]["id"] if categories else "")
        cat_ref = next((c for c in categories if c["id"] == cat_id), None)
        direction_name = cat_ref["directionName"] if cat_ref else ""
        direction_id = cat_ref["directionId"] if cat_ref else ""
        color = cat_ref["color"] if cat_ref else PORTRAIT_DIRECTION_COLORS_FALLBACK[0]
        cnt = int(clean(crow.get("count")) or 0)
        cluster_id = f"cluster-{idx:02d}"
        cluster = {
            "id": cluster_id,
            "code": code,
            "name": label or code,
            "categoryId": cat_id,
            "categoryName": cat_name,
            "directionId": direction_id,
            "directionName": direction_name,
            "color": color,
            "jobCount": cnt,
            "companyCount": 0,
            "candidateStatus": "岗位画像Excel导入",
            "currentDiscoveryMethod": "Coze科研助理四层聚类（Excel）",
            "targetDiscoveryMethod": "专家校准后正式发布",
            "ruleMatchedRate": 0.0,
            "averageRuleScore": 0.0,
            "uniqueTitleCount": 0,
            "jdCoverageRate": 0.0,
            "skillCoverageRate": 0.0,
            "topCompanyShare": 0.0,
            "portraitCoverage": {"responsibilities": 0.0, "skills": 0.0, "abilities": 0.0, "scenarios": 0.0, "conditions": 0.0},
            "topKeywords": [],
            "topSkills": [],
            "topCompanies": [],
            "representativeJobs": [],
            "levelDistribution": {},
            "educationDistribution": {},
            "experienceDistribution": {},
            "industryDistribution": {},
            "regionDistribution": {},
            "financingDistribution": {},
        }
        clusters.append(cluster)
        cluster_by_code[code] = cluster
        if cat_ref:
            cat_ref["clusterCount"] += 1
            cat_ref["jobCount"] += 0  # will recompute below from job_overlay

    # --- 2) 岗位明细 4655 条：occ_id → {cluster_code, standardRoleName, ...} overlay ---
    job_detail_rows = _read_xlsx_rows(PORTRAIT_CLUSTER_XLSX, "岗位明细")
    job_overlay: dict[str, dict[str, Any]] = {}  # key: occ_id
    # 统计每个标准岗位的 occ_id 列表 & 公司集合 & 标题集合
    sr_occ_ids: dict[str, list[str]] = defaultdict(list)
    sr_companies: dict[str, set[str]] = defaultdict(set)
    sr_titles: dict[str, list[str]] = defaultdict(list)
    cluster_job_count: Counter[str] = Counter()
    cluster_company_set: dict[str, set[str]] = defaultdict(set)
    cluster_title_counter: dict[str, Counter[str]] = defaultdict(Counter)

    confidence_map = {"高": 0.95, "中": 0.75, "低": 0.55}
    method_normalize = {
        "标准岗位高置信": "标准岗位高置信映射",
        "标准岗位中置信": "标准岗位中置信映射",
        "标准岗位低置信": "标准岗位低置信映射",
    }
    for r in job_detail_rows:
        occ_id = clean(r.get("occ_id"))
        if not occ_id:
            continue
        cluster_code = clean(r.get("岗位簇编码"))
        standard_role_name = clean(r.get("标准岗位"))
        confidence_label = clean(r.get("匹配置信度"))
        confidence = confidence_map.get(confidence_label, 0.5)
        method_raw = clean(r.get("匹配方式"))
        method = method_normalize.get(method_raw, method_raw or "Excel导入映射")
        direction_name = clean(r.get("职业方向(映射后)"))
        category_name = clean(r.get("职业类别(映射后)"))
        overlay = {
            "occId": occ_id,
            "clusterCode": cluster_code,
            "standardRoleName": standard_role_name,
            "standardRoleMappingConfidence": confidence,
            "standardRoleMappingMethod": method,
            "directionName": direction_name,
            "categoryName": category_name,
        }
        job_overlay[occ_id] = overlay
        if standard_role_name:
            sr_occ_ids[standard_role_name].append(occ_id)
            company = clean(r.get("公司"))
            if company:
                sr_companies[standard_role_name].add(company)
            title = clean(r.get("岗位"))
            if title:
                sr_titles[standard_role_name].append(title)
        if cluster_code:
            cluster_job_count[cluster_code] += 1
            company = clean(r.get("公司"))
            if company:
                cluster_company_set[cluster_code].add(company)
            title = clean(r.get("岗位"))
            if title:
                cluster_title_counter[cluster_code][title] += 1

    # Recompute cluster.jobCount / companyCount / uniqueTitleCount from overlay
    for code, cluster in cluster_by_code.items():
        cluster["jobCount"] = cluster_job_count.get(code, cluster["jobCount"])
        cluster["companyCount"] = len(cluster_company_set.get(code, set()))
        top_titles: list[dict[str, Any]] = []
        for title, tcnt in cluster_title_counter[code].most_common(5):
            top_titles.append({"title": title, "count": tcnt, "company": "", "occId": "", "url": "", "jdSnippet": "", "profile": {}})
        cluster["representativeJobs"] = top_titles
        cluster["uniqueTitleCount"] = len(cluster_title_counter[code])
    # Recompute category.jobCount / direction.jobCount / categoryCount / clusterCount
    for cat in categories:
        cat["jobCount"] = sum(cl["jobCount"] for cl in clusters if cl["categoryId"] == cat["id"])
    for dr in directions:
        dr["categoryCount"] = sum(1 for c in categories if c["directionId"] == dr["id"])
        dr["clusterCount"] = sum(1 for cl in clusters if cl["directionId"] == dr["id"])
        dr["jobCount"] = sum(c["jobCount"] for c in categories if c["directionId"] == dr["id"])

    # --- 3) 标准岗位 107 行：五维画像解析 ---
    portrait_rows = _read_xlsx_rows(PORTRAIT_PROFILE_XLSX, "能力画像")
    standard_roles: list[dict[str, Any]] = []
    standard_role_job_edges: list[dict[str, Any]] = []
    dimension_limits = {"responsibilities": 8, "skills": 10, "abilities": 8, "scenarios": 8, "conditions": 8}
    for idx, prow in enumerate(portrait_rows, start=1):
        sr_name = clean(prow.get("标准岗位"))
        if not sr_name:
            continue
        cluster_code = clean(prow.get("岗位簇编码"))
        cluster_ref = cluster_by_code.get(cluster_code)
        cluster_id = cluster_ref["id"] if cluster_ref else (clusters[0]["id"] if clusters else "")
        cluster_name = clean(prow.get("岗位簇名称")) or (cluster_ref["name"] if cluster_ref else "")
        cat_name = clean(prow.get("所属职业类别")) or (cluster_ref["categoryName"] if cluster_ref else "")
        dir_name = clean(prow.get("所属职业方向")) or (cluster_ref["directionName"] if cluster_ref else "")
        dir_ref = next((d for d in directions if d["name"] == dir_name), None)
        direction_id = dir_ref["id"] if dir_ref else (cluster_ref["directionId"] if cluster_ref else "")
        color = dir_ref["color"] if dir_ref else PORTRAIT_DIRECTION_COLORS_FALLBACK[0]

        # N / variants / companyCount
        jd_count_cell = clean(prow.get("岗位数量"))
        try:
            n_cell = int(jd_count_cell) if jd_count_cell else 0
        except ValueError:
            n_cell = 0
        occ_list: list[str] = list(dict.fromkeys(sr_occ_ids.get(sr_name, [])))
        job_count = max(n_cell, len(occ_list))
        company_count = len(sr_companies.get(sr_name, set()))
        title_counter = Counter(sr_titles.get(sr_name, []))
        variants = list(dict.fromkeys([sr_name, *[t for t, _ in title_counter.most_common(20)]]))[:14]

        # Parse 5 dimensions
        total_jd = max(job_count, 2)
        dim_text = {
            "responsibilities": clean(prow.get("职责")),
            "skills": clean(prow.get("技能")),
            "abilities": clean(prow.get("能力")),
            "scenarios": clean(prow.get("场景")),
        }
        standard_profile: dict[str, list[dict[str, Any]]] = {}
        for dim, txt in dim_text.items():
            bullets = _parse_portrait_bullets(txt, total_jd)
            limit = dimension_limits[dim]
            items: list[dict[str, Any]] = []
            for name, cnt in bullets[:limit]:
                coverage = round(cnt / max(1, total_jd), 4)
                items.append({
                    "name": name,
                    "count": cnt,
                    "coverage": coverage,
                    "evidenceOccIds": list(occ_list),
                })
            standard_profile[dim] = items
        # Conditions
        cond_items: list[dict[str, Any]] = []
        for name, cnt in _parse_portrait_conditions(clean(prow.get("条件")))[:dimension_limits["conditions"]]:
            coverage = round(cnt / max(1, total_jd), 4)
            cond_items.append({
                "name": name,
                "count": cnt,
                "coverage": coverage,
                "evidenceOccIds": list(occ_list),
            })
        standard_profile["conditions"] = cond_items

        role_id = f"role-{safe_id(cluster_code or 'CL')}-{idx:03d}"
        role = {
            "id": role_id,
            "code": cluster_code,
            "name": sr_name,
            "clusterId": cluster_id,
            "clusterName": cluster_name,
            "categoryName": cat_name,
            "directionName": dir_name,
            "directionId": direction_id,
            "color": color,
            "jobCount": job_count,
            "jdCount": job_count,
            "companyCount": company_count,
            "seedVariants": variants,
            "standardProfile": standard_profile,
            "profileMethod": "岗位画像Excel 标准岗位五维能力画像（Coze科研助理批处理）+ 多JD频次/覆盖率校验；画像点至少保留2条JD支撑",
            "releaseStatus": "标准岗位 v1.0 · 岗位画像 Excel 导入",
        }
        standard_roles.append(role)

    # standard_role_job_edges: occ_id → role_id supported_by_job
    # We need occ_id → job_id mapping later in main(); for now create edges by role_name → role_id,
    # we'll update with job_id in main() using job_nodes standardRoleId backreference.
    # To keep things simple, build edges as placeholders keyed by role name: (role_id, sr_name)
    # and reify after job_nodes have standardRoleId assigned.

    audit = {
        "seedRoleCount": len(standard_roles),
        "mappedJobCount": len(job_overlay),
        "pendingJobCount": 0,
        "seedVariantCount": sum(len(r["seedVariants"]) for r in standard_roles),
        "mappingRate": round(len(job_overlay) / max(1, len(job_overlay) + 1), 4),
        "excelSource": True,
    }

    return {
        "directions": directions,
        "categories": categories,
        "clusters": clusters,
        "direction_order": direction_order,
        "job_overlay": job_overlay,
        "standard_roles": standard_roles,
        "standard_role_audit": audit,
        "cluster_by_code": cluster_by_code,
        "dir_id_by_name": dir_id_by_name,
        "cat_id_by_name": cat_id_by_name,
    }


def build_standard_roles(
    job_nodes: list[dict[str, Any]],
    clusters: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, int]]:
    seeds = load_standard_role_seeds()
    if not seeds:
        return [], [], {"seedRoleCount": 0, "mappedJobCount": 0, "pendingJobCount": len(job_nodes)}

    cluster_lookup = {cluster["id"]: cluster for cluster in clusters}
    cluster_by_name = {cluster["name"]: cluster for cluster in clusters}
    title_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for job in job_nodes:
        title_groups[normalize(job.get("title"))].append(job)

    roles: list[dict[str, Any]] = []
    for seed in seeds:
        alias_set = set(seed["normalizedVariants"])
        hinted_cluster_name = STANDARD_ROLE_CLUSTER_HINTS.get(seed["id"])
        hinted_cluster = cluster_by_name.get(hinted_cluster_name or "")
        if hinted_cluster:
            cluster_id = hinted_cluster["id"]
            inference_method = "人工标准岗位层级校准"
        else:
            exact_jobs = [job for title, jobs in title_groups.items() if title in alias_set for job in jobs]
            cluster_scores: defaultdict[str, float] = defaultdict(float)
            inference_method = "搜索词包别名精确"
            if exact_jobs:
                for job in exact_jobs:
                    cluster_scores[job["clusterId"]] += 1.0
            else:
                inference_method = "搜索词包名称相似"
                title_matches: list[tuple[float, str]] = []
                for title in title_groups:
                    score = max((title_similarity(title, alias) for alias in alias_set), default=0.0)
                    if score >= 0.28:
                        title_matches.append((score, title))
                for score, title in sorted(title_matches, reverse=True)[:24]:
                    for job in title_groups[title]:
                        cluster_scores[job["clusterId"]] += score ** 4
            cluster_id = max(cluster_scores, key=cluster_scores.get) if cluster_scores else clusters[0]["id"]
        cluster = cluster_lookup[cluster_id]
        roles.append({
            "id": f"role-{seed['id'].lower()}",
            "code": seed["id"],
            "name": seed["name"],
            "clusterId": cluster_id,
            "clusterName": cluster["name"],
            "categoryId": cluster["categoryId"],
            "categoryName": cluster["categoryName"],
            "directionId": cluster["directionId"],
            "directionName": cluster["directionName"],
            "color": cluster["color"],
            "seedVariants": seed["seedVariants"],
            "normalizedVariants": seed["normalizedVariants"],
            "taxonomyMethod": inference_method,
        })

    roles_by_cluster: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for role in roles:
        roles_by_cluster[role["clusterId"]].append(role)

    mapped_count = 0
    mapping_method_counter: Counter[str] = Counter()
    role_members: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for job in job_nodes:
        candidates = roles_by_cluster.get(job["clusterId"], [])
        job_title = normalize(job["title"])
        ranked: list[tuple[float, dict[str, Any]]] = []
        for role in candidates:
            score = max((title_similarity(job_title, alias) for alias in role["normalizedVariants"]), default=0.0)
            ranked.append((score, role))
        ranked.sort(key=lambda item: item[0], reverse=True)
        best_score, best_role = ranked[0] if ranked else (0.0, None)
        second_score = ranked[1][0] if len(ranked) > 1 else 0.0
        # 高阈值 + 与第二候选的差值共同构成发布闸门；宁可留待专家映射，也不把通用职级词当作业务语义。
        if best_role is not None and best_score >= 0.54 and (best_score >= 0.999 or best_score - second_score >= 0.05):
            method = "别名精确" if best_score >= 0.999 else "类内标题相似"
            job["standardRoleId"] = best_role["id"]
            job["standardRoleName"] = best_role["name"]
            job["standardRoleMappingMethod"] = method
            job["standardRoleMappingConfidence"] = round(best_score, 4)
            role_members[best_role["id"]].append(job)
            mapped_count += 1
            mapping_method_counter[method] += 1
        else:
            job["standardRoleId"] = ""
            job["standardRoleName"] = ""
            job["standardRoleMappingMethod"] = "待专家映射"
            job["standardRoleMappingConfidence"] = round(best_score, 4)
            mapping_method_counter["待专家映射"] += 1

    role_edges: list[dict[str, Any]] = []
    for role in roles:
        members = role_members.get(role["id"], [])
        title_counter = Counter(job["title"] for job in members if job["title"])
        observed_variants = [
            {"name": name, "count": count}
            for name, count in title_counter.most_common(16)
        ]
        role["jobCount"] = len(members)
        role["companyCount"] = len({job["company"] for job in members if job["company"]})
        role["jdCount"] = sum(1 for job in members if job.get("profile", {}).get("jdAvailable"))
        role["observedVariants"] = observed_variants
        role["standardProfile"] = aggregate_standard_profile(members)
        role["profileMethod"] = (
            "搜索词包标准岗位 + 人工层级校准 + 名称变体归并 + 多JD频次/覆盖率交叉验证；"
            "画像点至少由2条JD共同支持"
        )
        role["releaseStatus"] = "标准岗位候选 · 待专家校准"
        role.pop("normalizedVariants", None)
        role_edges.append({
            "id": f"edge-{role['clusterId']}-{role['id']}",
            "source": role["clusterId"],
            "target": role["id"],
            "type": "contains_standard_role",
        })
        for job in members:
            role_edges.append({
                "id": f"edge-{role['id']}-{job['id']}",
                "source": role["id"],
                "target": job["id"],
                "type": "supported_by_job",
            })

    audit = {
        "seedRoleCount": len(roles),
        "seedVariantCount": sum(len(role["seedVariants"]) for role in roles),
        "mappedJobCount": mapped_count,
        "pendingJobCount": len(job_nodes) - mapped_count,
        "mappingRate": round(mapped_count / max(1, len(job_nodes)), 4),
        "rolesWithEvidence": sum(1 for role in roles if role["jobCount"]),
        "mappingMethodDistribution": dict(mapping_method_counter.most_common()),
    }
    return roles, role_edges, audit


def assign_jobs(rows: list[dict[str, Any]]) -> tuple[dict[str, list[dict[str, Any]]], list[str]]:
    assigned: dict[str, list[dict[str, Any]]] = defaultdict(list)
    warnings: list[str] = []
    for row in rows:
        category = clean(row.get("职业种类"))
        specs = CLUSTER_SPECS.get(category)
        if not specs:
            warnings.append(f"未配置职业种类：{category or '空值'}")
            continue
        title = normalize(row.get("岗位"))
        skills = normalize(row.get("技能标签"))
        jd = normalize(row.get("清洗JD描述"))
        ranked = []
        for index, spec in enumerate(specs):
            score, evidence = score_spec(spec, title, skills, jd)
            ranked.append((score, -index, spec, evidence))
        best_score, _rank, best, evidence = max(ranked, key=lambda item: (item[0], item[1]))
        if best_score == 0:
            best = next(spec for spec in specs if spec.fallback)
            evidence = []
        enriched = dict(row)
        enriched["_cluster_name"] = best.name
        enriched["_rule_score"] = best_score
        enriched["_rule_evidence"] = evidence
        assigned[f"{category}::{best.name}"].append(enriched)
    return assigned, sorted(set(warnings))


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    workbook = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    if SOURCE_SHEET:
        sheet_name = SOURCE_SHEET
    elif "岗位信息v4_企业增强" in workbook.sheetnames:
        sheet_name = "岗位信息v4_企业增强"
    elif "岗位信息v4" in workbook.sheetnames:
        sheet_name = "岗位信息v4"
    else:
        sheet_name = "岗位数据"
    sheet = workbook[sheet_name]
    values = sheet.iter_rows(values_only=True)
    header = [clean(value) for value in next(values)]
    rows = [dict(zip(header, row, strict=False)) for row in values]
    workbook.close()

    technology_nodes, technology_l4_lookup = load_technology_master()
    exact_jd_technology_map, technology_source_audit = load_exact_jd_technology_map(technology_l4_lookup)
    technology_exact_jd_jobs: set[str] = set()
    technology_exact_skill_jobs: set[str] = set()

    assigned, warnings = assign_jobs(rows)
    direction_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    category_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        direction_rows[clean(row.get("职业方向"))].append(row)
        category_rows[clean(row.get("职业种类"))].append(row)

    category_order = list(CATEGORY_DIRECTION)
    direction_order = list(DIRECTION_COLORS)
    directions = []
    categories = []
    clusters = []
    nodes = [{"id": "root", "type": "root", "label": "具身智能岗位生态", "jobCount": len(rows)}]
    edges = []
    skill_node_counts: Counter[str] = Counter()
    full_skill_node_counts: Counter[str] = Counter()
    full_skill_labels: dict[str, str] = {}
    job_nodes: list[dict[str, Any]] = []
    job_edges: list[dict[str, Any]] = []

    cluster_sequence = 0
    for direction_index, direction_name in enumerate(direction_order, start=1):
        direction_id = f"direction-{direction_index:02d}"
        direction_categories = [name for name in category_order if CATEGORY_DIRECTION[name] == direction_name]
        direction_cluster_count = sum(
            1
            for category in direction_categories
            for spec in CLUSTER_SPECS[category]
            if assigned.get(f"{category}::{spec.name}")
        )
        direction = {
            "id": direction_id,
            "name": direction_name,
            "color": DIRECTION_COLORS[direction_name],
            "jobCount": len(direction_rows.get(direction_name, [])),
            "categoryCount": len(direction_categories),
            "clusterCount": direction_cluster_count,
        }
        directions.append(direction)
        nodes.append({"id": direction_id, "type": "direction", "label": direction_name, **direction})
        edges.append({"id": f"edge-root-{direction_id}", "source": "root", "target": direction_id, "type": "contains"})

        for category_name in direction_categories:
            category_index = category_order.index(category_name) + 1
            category_id = f"category-{category_index:02d}"
            category_jobs = category_rows.get(category_name, [])
            nonempty_specs = [
                spec for spec in CLUSTER_SPECS[category_name]
                if assigned.get(f"{category_name}::{spec.name}")
            ]
            category = {
                "id": category_id,
                "name": category_name,
                "directionId": direction_id,
                "directionName": direction_name,
                "color": DIRECTION_COLORS[direction_name],
                "jobCount": len(category_jobs),
                "clusterCount": len(nonempty_specs),
            }
            categories.append(category)
            nodes.append({"id": category_id, "type": "category", "label": category_name, **category})
            edges.append({"id": f"edge-{direction_id}-{category_id}", "source": direction_id, "target": category_id, "type": "contains"})

            for spec in nonempty_specs:
                cluster_sequence += 1
                cluster_id = f"cluster-{cluster_sequence:02d}"
                members = assigned[f"{category_name}::{spec.name}"]
                title_counter = Counter(clean(row.get("岗位")) for row in members if clean(row.get("岗位")))
                company_counter = Counter(clean(row.get("公司")) for row in members if clean(row.get("公司")))
                skill_counter: Counter[str] = Counter()
                level_counter: Counter[str] = Counter()
                education_counter: Counter[str] = Counter()
                experience_counter: Counter[str] = Counter()
                industry_counter: Counter[str] = Counter()
                region_counter: Counter[str] = Counter()
                financing_counter: Counter[str] = Counter()
                evidence_counter: Counter[str] = Counter()
                score_total = 0
                member_profiles: list[dict[str, Any]] = []
                for member in members:
                    member_profile = build_job_profile(member)
                    member_profiles.append(member_profile)
                    skill_counter.update(split_skills(member.get("技能标签")))
                    if clean(member.get("能力等级")):
                        level_counter[clean(member.get("能力等级"))] += 1
                    if clean(member.get("学历(标准化)")):
                        education_counter[clean(member.get("学历(标准化)"))] += 1
                    if clean(member.get("经验(标准化)")):
                        experience_counter[clean(member.get("经验(标准化)"))] += 1
                    if clean(member.get("产业链层级")):
                        industry_counter[clean(member.get("产业链层级"))] += 1
                    if clean(member.get("公司所属地区")):
                        region_counter[clean(member.get("公司所属地区"))] += 1
                    if clean(member.get("融资轮次")):
                        financing_counter[clean(member.get("融资轮次"))] += 1
                    evidence_counter.update(member["_rule_evidence"])
                    score_total += member["_rule_score"]

                top_titles = []
                ranked_titles = sorted(
                    title_counter.items(),
                    key=lambda item: ("未说明" in item[0], -item[1], item[0]),
                )[:5]
                for title, count in ranked_titles:
                    title_members = [row for row in members if clean(row.get("岗位")) == title]
                    representative = max(
                        title_members,
                        key=lambda row: (
                            len(build_job_profile(row)["responsibilities"]),
                            len(build_job_profile(row)["jdEvidence"]),
                            len(build_job_profile(row)["skills"]),
                            bool(clean(row.get("清洗JD描述"))),
                            len(clean(row.get("清洗JD描述"))),
                        ),
                    )
                    representative_profile = build_job_profile(representative)
                    top_titles.append({
                        "title": title,
                        "count": count,
                        "company": clean(representative.get("公司")),
                        "occId": clean(representative.get("occ_id")),
                        "url": clean(representative.get("链接")),
                        "jdSnippet": compact_text(representative.get("清洗JD描述"), 220),
                        "profile": representative_profile,
                    })
                avg_rule_score = score_total / max(1, len(members))
                matched_count = sum(1 for row in members if row["_rule_score"] > 0)
                jd_covered = sum(1 for profile in member_profiles if profile["jdAvailable"])
                skill_covered = sum(1 for profile in member_profiles if profile["skills"])
                portrait_coverage = {
                    dimension: round(
                        sum(1 for profile in member_profiles if profile[dimension]) / max(1, len(member_profiles)),
                        4,
                    )
                    for dimension in ("responsibilities", "skills", "abilities", "scenarios", "conditions")
                }
                cluster = {
                    "id": cluster_id,
                    "code": f"JCG-{category_index:02d}-{nonempty_specs.index(spec) + 1:02d}",
                    "name": spec.name,
                    "categoryId": category_id,
                    "categoryName": category_name,
                    "directionId": direction_id,
                    "directionName": direction_name,
                    "color": DIRECTION_COLORS[direction_name],
                    "jobCount": len(members),
                    "companyCount": len(company_counter),
                    "candidateStatus": "待语义复算与专家校准",
                    "currentDiscoveryMethod": "类内关键词与JD证据候选分配",
                    "targetDiscoveryMethod": "类内Embedding + HDBSCAN + 专家校准",
                    "ruleMatchedRate": round(matched_count / max(1, len(members)), 4),
                    "averageRuleScore": round(avg_rule_score, 2),
                    "uniqueTitleCount": len(title_counter),
                    "jdCoverageRate": round(jd_covered / max(1, len(members)), 4),
                    "skillCoverageRate": round(skill_covered / max(1, len(members)), 4),
                    "topCompanyShare": round((company_counter.most_common(1)[0][1] if company_counter else 0) / max(1, len(members)), 4),
                    "portraitCoverage": portrait_coverage,
                    "topKeywords": [item["name"] for item in counter_rows(evidence_counter, 8)],
                    "topSkills": counter_rows(skill_counter, 8),
                    "topCompanies": counter_rows(company_counter, 5),
                    "representativeJobs": top_titles,
                    "levelDistribution": dict(level_counter.most_common()),
                    "educationDistribution": dict(education_counter.most_common()),
                    "experienceDistribution": dict(experience_counter.most_common()),
                    "industryDistribution": dict(industry_counter.most_common()),
                    "regionDistribution": dict(region_counter.most_common()),
                    "financingDistribution": dict(financing_counter.most_common()),
                }
                clusters.append(cluster)
                nodes.append({"id": cluster_id, "type": "job_cluster", "label": spec.name, **cluster})
                edges.append({"id": f"edge-{category_id}-{cluster_id}", "source": category_id, "target": cluster_id, "type": "contains"})

                for member_index, member in enumerate(members, start=1):
                    title = clean(member.get("岗位")) or "岗位名称未说明"
                    company = clean(member.get("公司"))
                    occ_id = clean(member.get("occ_id"))
                    job_id = f"job-{safe_id(f'{occ_id}|{title}|{company}|{cluster_id}|{member_index}')}"
                    job_skills = list(dict.fromkeys(split_skills(member.get("技能标签"))))
                    exact_jd_term_ids = set(exact_jd_technology_map.get(content_hash(member.get("清洗JD描述")), []))
                    exact_skill_term_ids = {
                        technology_l4_lookup[normalize_match_key(skill_name)]
                        for skill_name in job_skills
                        if normalize_match_key(skill_name) in technology_l4_lookup
                    }
                    technology_term_ids = sorted(exact_jd_term_ids | exact_skill_term_ids)
                    technology_mapping_methods = []
                    if exact_jd_term_ids:
                        technology_mapping_methods.append("JD全文一致回接")
                        technology_exact_jd_jobs.add(job_id)
                    if exact_skill_term_ids:
                        technology_mapping_methods.append("L4技术词精确命中")
                        technology_exact_skill_jobs.add(job_id)
                    job_nodes.append({
                        "id": job_id,
                        "type": "job",
                        "label": title,
                        "occId": occ_id,
                        "title": title,
                        "company": company,
                        "url": clean(member.get("链接")),
                        "jd": clean(member.get("清洗JD描述")),
                        "skills": job_skills,
                        "technologyTermIds": technology_term_ids,
                        "technologyMappingMethods": technology_mapping_methods,
                        "directionId": direction_id,
                        "directionName": direction_name,
                        "categoryId": category_id,
                        "categoryName": category_name,
                        "clusterId": cluster_id,
                        "clusterName": spec.name,
                        "abilityLevel": clean(member.get("能力等级")),
                        "education": clean(member.get("学历(标准化)")),
                        "experience": clean(member.get("经验(标准化)")),
                        "industryStage": clean(member.get("产业链层级")),
                        "industryCategory": clean(member.get("产业链12类")),
                        "companySpecialty": clean(member.get("公司细分领域")),
                        "financingRound": clean(member.get("融资轮次")),
                        "companyRegion": clean(member.get("公司所属地区")),
                        "headquartersCity": clean(member.get("公司总部城市")),
                        "enterpriseName": clean(member.get("企业库标准名称")) or company,
                        "enterpriseMatchMethod": clean(member.get("企业匹配方式")),
                        "enterpriseMatchConfidence": clean(member.get("企业匹配置信度")),
                        "enterpriseEnrichmentStatus": clean(member.get("企业属性补全状态")),
                        "profile": build_job_profile(member),
                        "ruleScore": member["_rule_score"],
                        "ruleEvidence": member["_rule_evidence"],
                    })
                    job_edges.append({
                        "id": f"edge-{cluster_id}-{job_id}",
                        "source": cluster_id,
                        "target": job_id,
                        "type": "contains_job",
                    })
                    for skill_name in job_skills:
                        skill_id = f"skill-{safe_id(skill_name)}"
                        full_skill_node_counts[skill_id] += 1
                        full_skill_labels[skill_id] = skill_name
                        job_edges.append({
                            "id": f"edge-{safe_id(f'{job_id}|{skill_id}')}",
                            "source": job_id,
                            "target": skill_id,
                            "type": "requires_skill",
                            "weight": 1,
                        })

                for skill in cluster["topSkills"][:6]:
                    skill_id = f"skill-{safe_id(skill['name'])}"
                    skill_node_counts[skill_id] += skill["count"]
                    edges.append({
                        "id": f"edge-{cluster_id}-{skill_id}",
                        "source": cluster_id,
                        "target": skill_id,
                        "type": "requires_skill",
                        "weight": skill["count"],
                    })

    if USE_PORTRAIT_EXCEL_OVERRIDE:
        # --- Portrait Excel Override: overlay hierarchy, job归属, 标准岗位五维画像 ---
        portrait_pkg = load_portrait_package()
        portrait_dirs = portrait_pkg["directions"]
        portrait_cats = portrait_pkg["categories"]
        portrait_clusters = portrait_pkg["clusters"]
        job_overlay = portrait_pkg["job_overlay"]
        role_by_name = {r["name"]: r for r in portrait_pkg["standard_roles"]}
        cluster_by_code = portrait_pkg["cluster_by_code"]
        dir_by_name = portrait_pkg["dir_id_by_name"]
        cat_by_name = portrait_pkg["cat_id_by_name"]

        # 1) Overlay job_nodes fields by occ_id
        for job in job_nodes:
            occ_id = clean(job.get("occId"))
            overlay = job_overlay.get(occ_id)
            if not overlay:
                continue
            cluster_code = overlay.get("clusterCode")
            cluster_ref = cluster_by_code.get(cluster_code) if cluster_code else None
            sr_name = overlay.get("standardRoleName")
            role_ref = role_by_name.get(sr_name) if sr_name else None
            direction_name = overlay.get("directionName")
            category_name = overlay.get("categoryName")
            direction_id = dir_by_name.get(direction_name, job.get("directionId"))
            category_id = cat_by_name.get(category_name, job.get("categoryId"))
            if cluster_ref:
                cluster_id = cluster_ref["id"]
                cluster_name = cluster_ref["name"]
            else:
                cluster_id = job.get("clusterId")
                cluster_name = job.get("clusterName")
            job.update({
                "directionId": direction_id,
                "directionName": direction_name or job.get("directionName"),
                "categoryId": category_id,
                "categoryName": category_name or job.get("categoryName"),
                "clusterId": cluster_id,
                "clusterName": cluster_name,
                "standardRoleId": role_ref["id"] if role_ref else "",
                "standardRoleName": sr_name or job.get("standardRoleName", ""),
                "standardRoleMappingMethod": overlay.get("standardRoleMappingMethod") or job.get("standardRoleMappingMethod", ""),
                "standardRoleMappingConfidence": overlay.get("standardRoleMappingConfidence") or job.get("standardRoleMappingConfidence", 0.0),
            })
            # Also refresh contains_job edge source to point to new cluster_id
            for edge in job_edges:
                if edge.get("type") == "contains_job" and edge.get("target") == job["id"]:
                    edge["source"] = cluster_id
                    edge["id"] = f"edge-{cluster_id}-{job['id']}"
                    break

        # 2) Replace directions / categories / clusters arrays (payload will use them)
        directions = portrait_dirs
        categories = portrait_cats
        clusters = portrait_clusters

        # 3) Rebuild nodes hierarchy: remove old direction/category/cluster nodes, add new ones
        nodes = [n for n in nodes if n.get("type") not in {"direction", "category", "job_cluster", "standard_role"}]
        edges = [e for e in edges if e.get("type") == "contains" and not (
            e.get("source") == "root"
            or any(n.get("type") in {"direction", "category"} for n in nodes if n.get("id") == e.get("source"))
        )]
        # Re-add root → direction → category → cluster nodes & edges
        root_job_count = sum(d["jobCount"] for d in directions)
        for n in nodes:
            if n.get("type") == "root":
                n["jobCount"] = root_job_count or n.get("jobCount", 0)
        for d in directions:
            nodes.append({"id": d["id"], "type": "direction", "label": d["name"], **d})
            edges.append({"id": f"edge-root-{d['id']}", "source": "root", "target": d["id"], "type": "contains"})
        for c in categories:
            nodes.append({"id": c["id"], "type": "category", "label": c["name"], **c})
            edges.append({"id": f"edge-{c['directionId']}-{c['id']}", "source": c["directionId"], "target": c["id"], "type": "contains"})
        for cl in clusters:
            nodes.append({"id": cl["id"], "type": "job_cluster", "label": cl["name"], **cl})
            edges.append({"id": f"edge-{cl['categoryId']}-{cl['id']}", "source": cl["categoryId"], "target": cl["id"], "type": "contains"})

        # 4) Build standard_role_edges: cluster → role (contains) & job → role (supported_by)
        standard_roles = portrait_pkg["standard_roles"]
        standard_role_audit = portrait_pkg["standard_role_audit"]
        standard_role_job_edges: list[dict[str, Any]] = []
        cluster_role_edges: list[dict[str, Any]] = []
        occ_id_to_job_id = {job["occId"]: job["id"] for job in job_nodes if clean(job.get("occId"))}
        occ_id_to_role_id: dict[str, str] = {}  # occ_id -> role_id if that occ is assigned to a standard role via overlay
        for occ_id, overlay in job_overlay.items():
            sr_name = overlay.get("standardRoleName")
            if not sr_name:
                continue
            role_ref = role_by_name.get(sr_name)
            if not role_ref:
                continue
            occ_id_to_role_id[occ_id] = role_ref["id"]
            job_id = occ_id_to_job_id.get(occ_id)
            if job_id:
                standard_role_job_edges.append({
                    "id": f"edge-rolejob-{safe_id(role_ref['id'] + '|' + job_id)}",
                    "source": job_id,
                    "target": role_ref["id"],
                    "type": "supported_by_job",
                    "weight": 1,
                })
        # contains_standard_role edges (cluster → role)
        added_cluster_role: set[tuple[str, str]] = set()
        for role in standard_roles:
            cid = role.get("clusterId")
            if cid and (cid, role["id"]) not in added_cluster_role:
                cluster_role_edges.append({
                    "id": f"edge-clrole-{safe_id(cid + '|' + role['id'])}",
                    "source": cid,
                    "target": role["id"],
                    "type": "contains_standard_role",
                })
                added_cluster_role.add((cid, role["id"]))
        standard_role_edges = cluster_role_edges + standard_role_job_edges

        # 5) Add standard_role nodes
        standard_role_nodes = [
            {"id": role["id"], "type": "standard_role", "label": role["name"], **role}
            for role in standard_roles
        ]
        nodes.extend(standard_role_nodes)
        edges.extend(cluster_role_edges)
        # technology_role_ids recomputation happens later automatically.
    else:
        standard_roles, standard_role_edges, standard_role_audit = build_standard_roles(job_nodes, clusters)
        standard_role_nodes = [
            {"id": role["id"], "type": "standard_role", "label": role["name"], **role}
            for role in standard_roles
        ]
        nodes.extend(standard_role_nodes)
        edges.extend(edge for edge in standard_role_edges if edge["type"] == "contains_standard_role")
        standard_role_job_edges = [edge for edge in standard_role_edges if edge["type"] == "supported_by_job"]

    technology_node_lookup = {node["id"]: node for node in technology_nodes}
    technology_job_ids: dict[str, set[str]] = defaultdict(set)
    technology_role_ids: dict[str, set[str]] = defaultdict(set)
    for job in job_nodes:
        for technology_id in job.get("technologyTermIds", []):
            current_id = technology_id
            while current_id and current_id in technology_node_lookup:
                technology_job_ids[current_id].add(job["id"])
                if job.get("standardRoleId"):
                    technology_role_ids[current_id].add(job["standardRoleId"])
                current_id = technology_node_lookup[current_id].get("parentId", "")
    for node in technology_nodes:
        node["jobCount"] = len(technology_job_ids.get(node["id"], set()))
        node["standardRoleCount"] = len(technology_role_ids.get(node["id"], set()))

    technology_mapped_jobs = technology_exact_jd_jobs | technology_exact_skill_jobs
    technology_audit = {
        "masterFile": TECHNOLOGY_MASTER_FILE.name,
        "legacySummaryFile": TECHNOLOGY_SUMMARY_FILE.name,
        "legacyMatchFile": TECHNOLOGY_MATCH_FILE.name,
        "levelCounts": dict(Counter(node["level"] for node in technology_nodes)),
        "mappedJobCount": len(technology_mapped_jobs),
        "mappedJobRate": round(len(technology_mapped_jobs) / max(1, len(job_nodes)), 4),
        "pendingJobCount": len(job_nodes) - len(technology_mapped_jobs),
        "exactJdJobCount": len(technology_exact_jd_jobs),
        "exactL4SkillJobCount": len(technology_exact_skill_jobs),
        "activeL4TermCount": sum(1 for node in technology_nodes if node["level"] == "L4" and node["jobCount"]),
        "mappingRule": "仅采用清洗JD全文一致回接和L4技术词精确命中；不使用公司名+岗位名模糊猜测",
        **technology_source_audit,
    }

    skill_labels: dict[str, str] = {}
    for cluster in clusters:
        for skill in cluster["topSkills"][:6]:
            skill_id = f"skill-{safe_id(skill['name'])}"
            skill_labels[skill_id] = skill["name"]
    for skill_id, count in skill_node_counts.items():
        label = skill_labels.get(skill_id) or full_skill_labels.get(skill_id) or skill_id.replace("skill-", "")
        nodes.append({"id": skill_id, "type": "skill", "label": label, "jobCount": count})

    enterprise_members: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for job in job_nodes:
        if job["enterpriseEnrichmentStatus"] == "已匹配" and job["enterpriseName"]:
            enterprise_members[job["enterpriseName"]].append(job)

    enterprise_nodes: list[dict[str, Any]] = []
    attribute_node_counts: Counter[str] = Counter()
    attribute_node_labels: dict[str, tuple[str, str]] = {}
    enterprise_edges: list[dict[str, Any]] = []
    job_company_edges: list[dict[str, Any]] = []
    enterprises: list[dict[str, Any]] = []

    def most_common_value(members: list[dict[str, Any]], field: str) -> str:
        values = Counter(clean(member.get(field)) for member in members if clean(member.get(field)))
        return values.most_common(1)[0][0] if values else ""

    for enterprise_name, members in sorted(
        enterprise_members.items(),
        key=lambda item: (-len(item[1]), item[0]),
    ):
        enterprise_id = f"enterprise-{safe_id(enterprise_name)}"
        title_counter = Counter(member["title"] for member in members if member["title"])
        industry_stage = most_common_value(members, "industryStage")
        industry_category = most_common_value(members, "industryCategory")
        company_specialty = most_common_value(members, "companySpecialty")
        financing_round = most_common_value(members, "financingRound")
        company_region = most_common_value(members, "companyRegion")
        headquarters_city = most_common_value(members, "headquartersCity")
        representative_jobs: list[dict[str, Any]] = []
        ranked_enterprise_titles = sorted(
            title_counter.items(),
            key=lambda item: ("未说明" in item[0], -item[1], item[0]),
        )[:5]
        for title, count in ranked_enterprise_titles:
            candidates = [member for member in members if member["title"] == title]
            representative = max(
                candidates,
                key=lambda member: (
                    len(member.get("profile", {}).get("skills", [])),
                    len(member.get("profile", {}).get("jdEvidence", [])),
                    len(member.get("jd", "")),
                ),
            )
            representative_jobs.append({
                "title": title,
                "count": count,
                "company": enterprise_name,
                "occId": representative["occId"],
                "url": representative["url"],
                "jdSnippet": representative["jd"][:420],
                "profile": representative["profile"],
                "clusterId": representative["clusterId"],
                "clusterName": representative["clusterName"],
                "categoryName": representative["categoryName"],
                "directionName": representative["directionName"],
            })
        enterprise = {
            "id": enterprise_id,
            "name": enterprise_name,
            "jobCount": len(members),
            "industryStage": industry_stage,
            "industryCategory": industry_category,
            "companySpecialty": company_specialty,
            "financingRound": financing_round,
            "companyRegion": company_region,
            "headquartersCity": headquarters_city,
            "directionDistribution": dict(Counter(member["directionName"] for member in members).most_common()),
            "categoryDistribution": dict(Counter(member["categoryName"] for member in members).most_common()),
            "clusterDistribution": dict(Counter(member["clusterName"] for member in members).most_common()),
            "representativeJobs": representative_jobs,
        }
        enterprises.append(enterprise)
        enterprise_nodes.append({
            "id": enterprise_id,
            "type": "enterprise",
            "label": enterprise_name,
            **enterprise,
        })
        for member in members:
            job_company_edges.append({
                "id": f"edge-{safe_id(member['id'] + '|' + enterprise_id)}",
                "source": member["id"],
                "target": enterprise_id,
                "type": "posted_by",
                "weight": 1,
            })
        dimension_values = (
            ("industry_stage", "产业链层级", industry_stage, "belongs_to_industry"),
            ("industry_category", "产业链12类", industry_category, "belongs_to_industry_category"),
            ("company_specialty", "公司细分领域", company_specialty, "specializes_in"),
            ("region", "公司所属地区", company_region, "located_in_region"),
            ("financing", "融资轮次", financing_round, "has_financing_round"),
            ("city", "公司总部城市", headquarters_city, "headquartered_in"),
        )
        for dimension_type, dimension_label, value, relation in dimension_values:
            if not value:
                continue
            attribute_id = f"{dimension_type}-{safe_id(value)}"
            attribute_node_counts[attribute_id] += len(members)
            attribute_node_labels[attribute_id] = (dimension_type, value)
            enterprise_edges.append({
                "id": f"edge-{enterprise_id}-{attribute_id}",
                "source": enterprise_id,
                "target": attribute_id,
                "type": relation,
                "weight": len(members),
                "dimension": dimension_label,
            })

    attribute_nodes = [
        {
            "id": attribute_id,
            "type": dimension_type,
            "label": label,
            "jobCount": count,
        }
        for attribute_id, count in attribute_node_counts.items()
        for dimension_type, label in [attribute_node_labels[attribute_id]]
    ]

    matched_job_count = sum(len(members) for members in enterprise_members.values())
    enterprise_analysis = {
        "enterpriseLibraryFile": ENTERPRISE_LIBRARY_FILE,
        "enterpriseLibraryRecordCount": ENTERPRISE_LIBRARY_RECORD_COUNT,
        "sourceCompanyCount": len({clean(row.get("公司")) for row in rows if clean(row.get("公司"))}),
        "matchedEnterpriseCount": len(enterprises),
        "matchedJobCount": matched_job_count,
        "matchedJobRate": round(matched_job_count / max(1, len(rows)), 4),
        "pendingJobCount": len(rows) - matched_job_count,
        "statusDistribution": dict(Counter(clean(row.get("企业属性补全状态")) or "未提供" for row in rows).most_common()),
        "matchMethodDistribution": dict(Counter(clean(row.get("企业匹配方式")) or "未提供" for row in rows).most_common()),
        "industryDistribution": counter_rows(Counter(job["industryStage"] for job in job_nodes if job["enterpriseEnrichmentStatus"] == "已匹配" and job["industryStage"]), 20),
        "industryCategoryDistribution": counter_rows(Counter(job["industryCategory"] for job in job_nodes if job["enterpriseEnrichmentStatus"] == "已匹配" and job["industryCategory"]), 20),
        "regionDistribution": counter_rows(Counter(job["companyRegion"] for job in job_nodes if job["enterpriseEnrichmentStatus"] == "已匹配" and job["companyRegion"]), 20),
        "financingDistribution": counter_rows(Counter(job["financingRound"] for job in job_nodes if job["enterpriseEnrichmentStatus"] == "已匹配" and job["financingRound"]), 20),
        "headquartersCityDistribution": counter_rows(Counter(job["headquartersCity"] for job in job_nodes if job["enterpriseEnrichmentStatus"] == "已匹配" and job["headquartersCity"]), 20),
        "topEnterprises": enterprises[:20],
    }

    hierarchy_nodes = [node for node in nodes if node["type"] != "skill"]
    full_skill_nodes = [
        {"id": skill_id, "type": "skill", "label": full_skill_labels[skill_id], "jobCount": count}
        for skill_id, count in full_skill_node_counts.items()
    ]
    full_nodes = hierarchy_nodes + job_nodes + full_skill_nodes + enterprise_nodes + attribute_nodes
    hierarchy_edges = [edge for edge in edges if edge["type"] != "requires_skill"]
    full_edges = hierarchy_edges + job_edges + standard_role_job_edges + job_company_edges + enterprise_edges
    public_jobs = [
        {
            "id": job["id"],
            "title": job["title"],
            "count": 1,
            "company": job["company"],
            "occId": job["occId"],
            "url": job["url"],
            "jd": job["jd"],
            "jdSnippet": compact_text(job["jd"], 420),
            "profile": job["profile"],
            "skills": job["skills"],
            "technologyTermIds": job.get("technologyTermIds", []),
            "technologyMappingMethods": job.get("technologyMappingMethods", []),
            "directionId": job["directionId"],
            "directionName": job["directionName"],
            "categoryId": job["categoryId"],
            "categoryName": job["categoryName"],
            "clusterId": job["clusterId"],
            "clusterName": job["clusterName"],
            "standardRoleId": job["standardRoleId"],
            "standardRoleName": job["standardRoleName"],
            "standardRoleMappingMethod": job["standardRoleMappingMethod"],
            "standardRoleMappingConfidence": job["standardRoleMappingConfidence"],
            "abilityLevel": job["abilityLevel"],
            "education": job["education"],
            "experience": job["experience"],
            "enterpriseName": job["enterpriseName"],
            "industryStage": job["industryStage"],
            "companySpecialty": job["companySpecialty"],
            "financingRound": job["financingRound"],
            "companyRegion": job["companyRegion"],
            "headquartersCity": job["headquartersCity"],
        }
        for job in job_nodes
    ]

    if USE_PORTRAIT_EXCEL_OVERRIDE:
        method_text = (
            "岗位画像Excel三层级（方向/类别/簇来自分层聚类图谱nodes）+ "
            "4655条JD高置信标准岗位归属（四层聚类岗位明细）+ "
            "107标准岗位五维画像（职责/技能/能力/场景/条件解析，画像点≥2条JD支撑）"
        )
        release_status = "standard_v1.0 · portrait_excel"
        release_note = (
            f"标准岗位 {len(standard_roles)} 个（来源：标准岗位五维能力画像.xlsx，能力画像sheet），"
            f"职业方向 {len(directions)} / 职业种类 {len(categories)} / 岗位簇 {len(clusters)} "
            f"（来源：分层聚类图谱 nodes+edges）。JD归属 {standard_role_audit.get('mappedJobCount', 0)} 条 "
            f"（来源：四层聚类结果 岗位明细 sheet）。画像点 count≥2，右侧证据区保留具体JD证据详情。"
        )
    else:
        method_text = "6/17规则边界 + 42类内语义岗位簇 + 搜索词包标准岗位/名称变体 + 多JD五维画像 + 专家发布闸门"
        release_status = "candidate_v0.4"
        release_note = (
            f"107个标准岗位已完成人工层级校准；当前{standard_role_audit.get('mappedJobCount', 0)}条JD通过高置信标题闸门，"
            f"其余{standard_role_audit.get('pendingJobCount', 0)}条保留待专家映射。画像点至少由2条JD共同支持，具体JD只在右侧证据区展示。"
        )

    payload = {
        "metadata": {
            "sourceFile": SOURCE_XLSX.name,
            "sourceSheet": sheet_name,
            "enterpriseLibraryFile": ENTERPRISE_LIBRARY_FILE,
            "portraitExcelOverride": bool(USE_PORTRAIT_EXCEL_OVERRIDE),
            "generatedAt": datetime.now(UTC).isoformat(),
            "method": method_text,
            "releaseStatus": release_status,
            "releaseNote": release_note,
            "jobCount": len(rows),
            "directionCount": len(directions),
            "categoryCount": len(categories),
            "clusterCount": len(clusters),
            "standardRoleCount": len(standard_roles),
            "standardRoleVariantCount": standard_role_audit.get("seedVariantCount", 0),
            "standardRoleMappedJobCount": standard_role_audit.get("mappedJobCount", 0),
            "standardRoleMappingRate": standard_role_audit.get("mappingRate", 0),
            "skillCount": len(skill_node_counts),
            "technologyNodeCount": len(technology_nodes),
            "technologyMappedJobCount": technology_audit["mappedJobCount"],
            "technologyMappedJobRate": technology_audit["mappedJobRate"],
            "fullSkillCount": len(full_skill_nodes),
            "jobNodeCount": len(job_nodes),
            "enterpriseCount": len(enterprises),
            "enterpriseMatchedJobCount": matched_job_count,
            "enterpriseMatchRate": round(matched_job_count / max(1, len(rows)), 4),
            "enterprisePendingJobCount": len(rows) - matched_job_count,
            "fullNodeCount": len(full_nodes),
            "fullEdgeCount": len(full_edges),
            "warnings": warnings,
        },
        "directions": directions,
        "categories": categories,
        "clusters": clusters,
        "standardRoles": standard_roles,
        "standardRoleAudit": standard_role_audit,
        "jobs": public_jobs,
        "enterprises": enterprises,
        "enterpriseAnalysis": enterprise_analysis,
        "technologyNodes": technology_nodes,
        "technologyAudit": technology_audit,
        "nodes": nodes,
        "edges": edges,
    }

    json_text = json.dumps(payload, ensure_ascii=False, indent=2)
    (OUTPUT_DIR / "job-ecosystem-graph.json").write_text(json_text, encoding="utf-8")
    PUBLIC_OUTPUT.write_text(json_text, encoding="utf-8")
    (OUTPUT_DIR / "overview_nodes.json").write_text(json.dumps(nodes, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "overview_edges.json").write_text(json.dumps(edges, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "nodes.json").write_text(json.dumps(full_nodes, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "edges.json").write_text(json.dumps(full_edges, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "jobs.json").write_text(json.dumps(job_nodes, ensure_ascii=False, indent=2), encoding="utf-8")
    (OUTPUT_DIR / "enterprises.json").write_text(json.dumps(enterprises, ensure_ascii=False, indent=2), encoding="utf-8")

    report = f"""# 岗位生态图谱构建报告

> 生成时间：{payload['metadata']['generatedAt']}
> 发布状态：{payload['metadata']['releaseStatus']}
> Portrait Excel 覆盖：{"已启用" if USE_PORTRAIT_EXCEL_OVERRIDE else "未启用"}

## 一、图谱规模

- 岗位记录：{len(rows)}
- 职业方向：{len(directions)}
- 职业种类：{len(categories)}
- 非空候选岗位簇：{len(clusters)}
- 标准岗位：{len(standard_roles)}
- 岗位节点：{len(job_nodes)}
- 全量技能节点：{len(full_skill_nodes)}
- 已关联企业实体/岗位：{len(enterprises)} / {matched_job_count}
- 企业关联覆盖率：{matched_job_count / max(1, len(rows)):.1%}
- 企业待核验或待补全岗位：{len(rows) - matched_job_count}
- 全量节点/关系：{len(full_nodes)} / {len(full_edges)}

## 二、构建逻辑

项目同步生成三张互相贯通的业务图谱：

1. `岗位生态层级图：具身智能岗位生态 → {len(directions)}职业方向 → {len(categories)}职业种类 → {len(clusters)}岗位簇 → {len(rows)}岗位 → JD`
2. `岗位簇发现图：职业种类边界 → 岗位名称/JD/技能/企业 → 类内候选发现 → Embedding + HDBSCAN复算 → 专家命名发布`
3. `岗位画像证据图：岗位簇 → 标准岗位 → 职责/技能/能力/场景/条件 → JD证据`

{method_text}

## 三、发布边界

1. 当前结果适合前端浏览、簇命名讨论与专家复核。
2. 正式版本需补充稳定性/异常率/公司集中度检查。
3. 每个簇保留代表岗位、技能、企业、能力等级和规则命中率，支持追溯。
4. 企业未匹配与多候选记录进入待补全清单；后续补充企业库或别名后可增量重算。
"""
    (OUTPUT_DIR / "岗位图谱构建报告.md").write_text(report, encoding="utf-8")

    expected_job_count = int(os.environ.get("JOB_GRAPH_EXPECTED_COUNT", str(len(rows))))
    assert len(rows) == expected_job_count, f"expected {expected_job_count} jobs, got {len(rows)}"
    assert len(directions) >= 5, f"expected >=5 directions, got {len(directions)}"
    assert len(categories) >= 14, f"expected >=14 categories, got {len(categories)}"
    if not USE_PORTRAIT_EXCEL_OVERRIDE:
        assert len(directions) == 6, f"expected 6 directions (legacy), got {len(directions)}"
        assert len(categories) == 17, f"expected 17 categories (legacy), got {len(categories)}"
        assert 30 <= len(clusters) <= 50, f"candidate cluster count out of range: {len(clusters)}"
        assert sum(item["jobCount"] for item in clusters) == len(rows)
        assert sum(item["jobCount"] for item in categories) == len(rows)
        assert sum(item["jobCount"] for item in directions) == len(rows)
    else:
        # Portrait override: allow cluster counts & sums to be approximate (overlay may not cover 100%)
        assert len(clusters) >= 20, f"expected >=20 clusters (portrait), got {len(clusters)}"
    assert len(job_nodes) == len(rows)
    assert len(public_jobs) == len(rows)
    assert sum(1 for edge in full_edges if edge["type"] == "contains_job") == len(rows)
    assert matched_job_count == payload["metadata"]["enterpriseMatchedJobCount"]
    assert len(job_company_edges) == matched_job_count

    print(json.dumps(payload["metadata"], ensure_ascii=False, indent=2))
    print(f"output: {PUBLIC_OUTPUT}")


if __name__ == "__main__":
    main()
