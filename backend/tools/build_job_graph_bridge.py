from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    return [dict(zip(headers, row, strict=False)) for row in rows if any(row)]


def _text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).replace("\u200b", "").strip()
    return result or None


def _items(value: Any) -> list[str]:
    text = _text(value)
    if not text or text == "暂无":
        return []
    values = []
    for raw in re.split(r"[\r\n]+", text):
        item = re.sub(r"^[\s•·▪◦*-]+", "", raw).strip()
        if item and item != "暂无" and item not in values:
            values.append(item)
    return values


def _terms(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return list(dict.fromkeys(item.strip() for item in re.split(r"[;；|]", text) if item.strip()))


def _role_code(name: str) -> str:
    digest = hashlib.sha1(name.encode("utf-8")).hexdigest()[:12].upper()
    return f"SR-{digest}"


def build(source_root: Path) -> dict[str, Any]:
    portrait_dir = source_root / "岗位画像"
    technology_dir = source_root / "技术分层"
    hierarchy_path = portrait_dir / "_Coze_Drive_科研助理_岗位四层聚类结果.xlsx"
    portrait_path = portrait_dir / "_Coze_Drive_科研助理_标准岗位五维能力画像.xlsx"
    technology_path = technology_dir / "_Coze_Drive_科研助理_岗位_L2L3分类结果.xlsx"
    technology_graph_path = technology_dir / "_Coze_Drive_科研助理_岗位_L2L3图谱数据.xlsx"

    labels: dict[str, str] = {}
    for row in _records(technology_graph_path, "nodes"):
        code = _text(row.get("编码"))
        label = _text(row.get("label"))
        if code and label:
            labels.setdefault(code, label)

    roles: dict[str, dict[str, Any]] = {}
    role_code_by_name: dict[str, str] = {}
    for row in _records(portrait_path, "能力画像"):
        name = _text(row.get("标准岗位"))
        if not name:
            continue
        code = _role_code(name)
        role_code_by_name[name] = code
        roles[code] = {
            "name": name,
            "cluster_code": _text(row.get("岗位簇编码")),
            "cluster_name": _text(row.get("岗位簇名称")),
            "category": _text(row.get("所属职业类别")),
            "direction": _text(row.get("所属职业方向")),
            "job_count": int(row.get("岗位数量") or 0),
            "portrait": {
                "responsibilities": _items(row.get("职责")),
                "skills": _items(row.get("技能")),
                "capabilities": _items(row.get("能力")),
                "scenarios": _items(row.get("场景")),
                "conditions": _items(row.get("条件")),
            },
        }

    jobs: dict[str, dict[str, Any]] = {}
    for row in _records(hierarchy_path, "岗位明细"):
        occ_id = _text(row.get("occ_id"))
        role_name = _text(row.get("标准岗位"))
        if not occ_id or not role_name:
            continue
        role_code = role_code_by_name.setdefault(role_name, _role_code(role_name))
        if role_code not in roles:
            roles[role_code] = {
                "name": role_name,
                "cluster_code": _text(row.get("岗位簇编码")),
                "cluster_name": _text(row.get("岗位簇名称")),
                "category": _text(row.get("职业类别(映射后)")),
                "direction": _text(row.get("职业方向(映射后)")),
                "job_count": 0,
                "portrait": {
                    "responsibilities": [],
                    "skills": [],
                    "capabilities": [],
                    "scenarios": [],
                    "conditions": [],
                },
            }
        jobs[occ_id] = {
            "title": _text(row.get("岗位")),
            "company": _text(row.get("公司")),
            "role_code": role_code,
            "role_name": role_name,
            "direction": _text(row.get("职业方向(映射后)")),
            "category": _text(row.get("职业类别(映射后)")),
            "cluster_code": _text(row.get("岗位簇编码")),
            "cluster_name": _text(row.get("岗位簇名称")),
            "match_confidence": _text(row.get("匹配置信度")),
            "match_method": _text(row.get("匹配方式")),
            "technology_paths": [],
        }

    for row in _records(technology_path, "岗位明细"):
        occ_id = _text(row.get("occ_id"))
        job = jobs.get(occ_id or "")
        if not job:
            continue
        l2_code = _text(row.get("匹配L2编码"))
        l3_code = _text(row.get("匹配L3编码"))
        if not l2_code:
            continue
        l1_code = l2_code.split(".", 1)[0]
        path = [
            {"level": "L1", "code": l1_code, "name": labels.get(l1_code, l1_code)},
            {
                "level": "L2",
                "code": l2_code,
                "name": _text(row.get("匹配L2技术类")) or labels.get(l2_code, l2_code),
            },
        ]
        if l3_code:
            path.append(
                {
                    "level": "L3",
                    "code": l3_code,
                    "name": _text(row.get("匹配L3名称")) or labels.get(l3_code, l3_code),
                }
            )
        method = _text(row.get("匹配方式"))
        job["technology_paths"].append(
            {
                "path": path,
                "match_method": method,
                "evidence_grade": method == "L4精确匹配",
                "hit_terms": _terms(row.get("命中词")),
            }
        )

    return {
        "schema_version": "job_graph_bridge_v1",
        "source_version": "v4-2026-08-27",
        "metadata": {
            "job_count": len(jobs),
            "standard_role_count": len(roles),
            "join_key": "occ_id",
            "hierarchy": ["职业方向", "职业类别", "岗位簇", "标准岗位", "具体JD"],
        },
        "roles": roles,
        "jobs": jobs,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the person-job matching graph bridge.")
    parser.add_argument("source_root", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "data" / "job_graph_bridge.json",
    )
    args = parser.parse_args()
    payload = build(args.source_root)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    print(
        f"wrote {args.output}: {payload['metadata']['job_count']} jobs, "
        f"{payload['metadata']['standard_role_count']} roles"
    )


if __name__ == "__main__":
    main()
