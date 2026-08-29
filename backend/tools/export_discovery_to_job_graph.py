"""把新岗位发现的候选导出到岗位图谱的两个消费口。

产出两份，内容同源：

1. `data/source/20260826/core/岗位信息v4_企业增强分析.xlsx` 里新增 `新岗位发现`
   工作表——图谱的数据底座是这份 xlsx，候选写进去才算进了共享资产；
2. `frontend/public/new-role-discovery.json`——技术—岗位图谱按技术编码把候选
   挂到 L1–L4 节点上，读的是这份。

技术编码是两侧唯一的连接键：候选的技术点与图谱的 technologyNodes 用的是同一套
`T1.03.02` 体系。挂不上的编码逐条列在 `unmatchedTechnologyCodes` 里，不做近似
匹配——匹配不上通常意味着该词条晚于图谱快照，猜一个只会把错误藏起来。
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
from sqlalchemy import select, text

from app.db.session import SessionLocal
from app.modules.clustering.models import JobRole  # noqa: F401  # 候选表外键需要
from app.modules.discovery.models import CandidateTechnology, EmergingRoleCandidate
from app.modules.taxonomy.models import TechnologyNode

ROOT = Path(__file__).resolve().parents[2]
SOURCE_XLSX = ROOT / "data" / "source" / "20260826" / "core" / "岗位信息v4_企业增强分析.xlsx"
GRAPH_JSON = ROOT / "frontend" / "public" / "job-ecosystem-graph.json"
PUBLIC_JSON = ROOT / "frontend" / "public" / "new-role-discovery.json"
SHEET_NAME = "新岗位发现"

CLASSIFICATION_LABELS = {
    "existing_role": "已被覆盖",
    "role_evolution": "岗位演化",
    "library_gap": "岗位库缺失",
    "potential_new_role": "潜在新岗位",
    "upstream_signal": "研究侧领先信号",
    "milestone_signal": "产业里程碑信号",
}
MATURITY_LABELS = {
    "potential": "潜在",
    "emerging": "涌现中",
    "confirmed": "已确认",
    "mature": "成熟",
}
COLUMNS = [
    "候选编码",
    "岗位名称",
    "分类",
    "分类编码",
    "成熟度",
    "综合评分",
    "支撑JD数",
    "独立企业数",
    "缺口等级",
    "技术点编码",
    "技术点名称",
    "一句话定义",
]


# 候选 → 支撑 JD 的企业与产业链层级。走的是发现流程自己写下的证据链
# （task_ids → rel_industry_task_evidence → 招聘文本 → 企业），不做名称近似。
ENTERPRISE_SQL = text("""
SELECT c.candidate_code, o.canonical_name,
       JSON_UNQUOTE(JSON_EXTRACT(p.source_metadata_json,'$.industry_chain_level'))
FROM biz_emerging_role_candidate c
JOIN rel_industry_task_evidence e
  ON JSON_CONTAINS(JSON_EXTRACT(c.mechanical_card_json,'$.task_ids'),
                   CAST(e.industry_task_id AS JSON))
JOIN biz_job_posting p ON p.job_posting_id = e.job_posting_id
JOIN md_organization o ON o.organization_id = p.organization_id
GROUP BY c.candidate_code, o.canonical_name,
         JSON_UNQUOTE(JSON_EXTRACT(p.source_metadata_json,'$.industry_chain_level'))
""")


def collect(db) -> list[dict]:
    codes_by_candidate: dict[int, list[tuple[str, str]]] = defaultdict(list)
    for cid, code, name in db.execute(
        select(
            CandidateTechnology.emerging_role_candidate_id,
            TechnologyNode.technology_code,
            TechnologyNode.technology_name,
        ).join(
            TechnologyNode,
            TechnologyNode.technology_node_id == CandidateTechnology.technology_node_id,
        )
    ):
        codes_by_candidate[cid].append((code, name))

    enterprises: dict[str, set[str]] = defaultdict(set)
    stages: dict[str, Counter] = defaultdict(Counter)
    for code, org, stage in db.execute(ENTERPRISE_SQL):
        enterprises[code].add(org)
        if stage:
            stages[code][stage] += 1

    rows = []
    ordered = select(EmergingRoleCandidate).order_by(EmergingRoleCandidate.candidate_score.desc())
    for c in db.scalars(ordered):
        card = c.mechanical_card_json or {}
        expression = c.expression_json or {}
        pairs = sorted(set(codes_by_candidate.get(c.emerging_role_candidate_id, [])))
        rows.append(
            {
                "candidateCode": c.candidate_code,
                "name": c.proposed_name,
                "classification": CLASSIFICATION_LABELS.get(
                    c.classification_code, c.classification_code
                ),
                "classificationCode": c.classification_code,
                "maturity": MATURITY_LABELS.get(c.maturity_stage_code, c.maturity_stage_code),
                "score": float(c.candidate_score or 0),
                "supportJobCount": int(card.get("job_count", 0) or 0),
                "organizationCount": int(card.get("organization_count", 0) or 0),
                "gapGrade": card.get("gap_grade") or "",
                "technologyCodes": [code for code, _ in pairs],
                "technologyNames": [name for _, name in pairs],
                "definition": expression.get("one_line_definition") or "",
                "enterprises": sorted(enterprises.get(c.candidate_code, ())),
                "industryStages": [
                    {"stage": k, "jobCount": v}
                    for k, v in stages.get(c.candidate_code, Counter()).most_common()
                ],
            }
        )
    return rows


def write_sheet(rows: list[dict]) -> None:
    wb = openpyxl.load_workbook(SOURCE_XLSX)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    ws.append(COLUMNS)
    for row in rows:
        ws.append(
            [
                row["candidateCode"],
                row["name"],
                row["classification"],
                row["classificationCode"],
                row["maturity"],
                row["score"],
                row["supportJobCount"],
                row["organizationCount"],
                row["gapGrade"],
                "、".join(row["technologyCodes"]),
                "、".join(row["technologyNames"]),
                row["definition"],
            ]
        )
    wb.save(SOURCE_XLSX)


# 库里的产业链层级是自由文本（"中游-整机+算法（软硬全栈）"），图谱只用四档。
# 按前缀归并，归不进去的记为空，不猜。
GRAPH_STAGES = ("上游", "中游", "下游", "横向支撑")


def normalize_stage(raw: str | None) -> str:
    if not raw or raw == "null":
        return ""
    for stage in GRAPH_STAGES:
        if raw.startswith(stage):
            return stage
    return ""


def technology_descendants(graph: dict) -> dict[str, set[str]]:
    children: dict[str, list[str]] = defaultdict(list)
    for node in graph["technologyNodes"]:
        children[node.get("parentId") or ""].append(node["id"])

    def walk(root: str) -> set[str]:
        out: set[str] = set()
        stack = [root]
        while stack:
            current = stack.pop()
            out.add(current)
            stack.extend(children.get(current, []))
        return out

    return {node["id"]: walk(node["id"]) for node in graph["technologyNodes"]}


def attach_portrait(rows: list[dict], graph: dict) -> None:
    """按岗位簇给候选定位，供岗位画像图谱使用。

    判据是**招聘文本必须同时命中候选的每一个技术点**，而不是命中任意一个。
    放宽到任意一个会让结果失去意义：实测那样有 56 条候选被归到
    「市场·销售与商务拓展」——这类 JD 泛泛提及大量技术词，会主导任何宽松匹配。
    收紧后每条候选中位命中 22 条 JD，主导簇变为「机器人学习与强化学习」等
    与其技术构成相符的簇。

    定位不到的候选保留空值，不退回宽松口径。
    """
    descendants = technology_descendants(graph)
    jobs_by_term: dict[str, list[dict]] = defaultdict(list)
    for job in graph["jobs"]:
        for term in job.get("technologyTermIds") or []:
            jobs_by_term[term].append(job)

    for row in rows:
        groups = []
        for node_id in row.get("technologyNodeIds", []):
            reach = descendants.get(node_id, {node_id})
            groups.append({j["id"] for term in reach for j in jobs_by_term.get(term, [])})
        shared = set.intersection(*groups) if groups else set()
        if not shared:
            row["portraitClusterName"] = ""
            row["portraitDirectionName"] = ""
            row["portraitEvidenceJobCount"] = 0
            continue
        hit = [job for job in graph["jobs"] if job["id"] in shared]
        cluster = Counter(j.get("clusterName") or "" for j in hit).most_common(1)[0][0]
        direction = Counter(
            j.get("directionName") or "" for j in hit if (j.get("clusterName") or "") == cluster
        ).most_common(1)[0][0]
        row["portraitClusterName"] = cluster
        row["portraitDirectionName"] = direction
        row["portraitEvidenceJobCount"] = len(shared)


def write_json(rows: list[dict]) -> dict:
    graph = json.loads(GRAPH_JSON.read_text(encoding="utf-8"))
    # 图谱按 technologyNodes 的 code 建索引；候选的技术码挂到同码节点上。
    node_ids_by_code: dict[str, list[str]] = defaultdict(list)
    for node in graph["technologyNodes"]:
        if node.get("code"):
            node_ids_by_code[node["code"]].append(node["id"])

    matched: set[str] = set()
    unmatched: set[str] = set()
    for row in rows:
        ids: list[str] = []
        for code in row["technologyCodes"]:
            hit = node_ids_by_code.get(code)
            if hit:
                ids.extend(hit)
                matched.add(code)
            else:
                unmatched.add(code)
        row["technologyNodeIds"] = sorted(set(ids))
        merged: Counter = Counter()
        for item in row.get("industryStages", []):
            stage = normalize_stage(item["stage"])
            if stage:
                merged[stage] += item["jobCount"]
        row["industryStages"] = [
            {"stage": k, "jobCount": v} for k, v in merged.most_common()
        ]

    attach_portrait(rows, graph)
    placed = sum(1 for row in rows if row.get("portraitClusterName"))
    payload = {
        "metadata": {
            "generatedAt": datetime.now(UTC).isoformat(timespec="seconds"),
            "candidateCount": len(rows),
            "matchedTechnologyCodeCount": len(matched),
            "unmatchedTechnologyCodes": sorted(unmatched),
            "joinKey": "technology_code",
            "enterpriseLinkedCount": sum(1 for row in rows if row.get("enterprises")),
            "portraitPlacedCount": placed,
            "note": "候选为未入库的提议，与图谱中已观测的标准岗位不同级。",
        },
        "candidates": rows,
    }
    PUBLIC_JSON.write_text(
        json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    return payload["metadata"]


def main() -> None:
    parser = argparse.ArgumentParser()
    # 库在容器里、xlsx 与 public 目录在宿主机（data 是只读挂载），因此分两步：
    #   docker compose exec -T backend python -m tools.<本模块> --dump > rows.json
    #   python -m tools.<本模块> --rows rows.json
    parser.add_argument("--dump", action="store_true", help="只从库里取数并打印 JSON")
    parser.add_argument("--rows", help="读取 --dump 的产物，跳过取数直接落盘")
    args = parser.parse_args()

    if args.rows:
        rows = json.loads(Path(args.rows).read_text(encoding="utf-8"))
    else:
        with SessionLocal() as db:
            rows = collect(db)
        if args.dump:
            json.dump(rows, sys.stdout, ensure_ascii=False)
            return
    write_sheet(rows)
    meta = write_json(rows)
    print(f"候选 {meta['candidateCount']} 条已写入 {SHEET_NAME} 工作表与 {PUBLIC_JSON.name}")
    print(f"技术编码对上 {meta['matchedTechnologyCodeCount']} 个")
    print(
        f"有企业足迹 {meta['enterpriseLinkedCount']} 条 · "
        f"能定位到岗位簇 {meta['portraitPlacedCount']} 条"
    )
    if meta["unmatchedTechnologyCodes"]:
        print(f"挂不上图谱的编码 {len(meta['unmatchedTechnologyCodes'])} 个："
              f"{'、'.join(meta['unmatchedTechnologyCodes'])}")


if __name__ == "__main__":
    main()
