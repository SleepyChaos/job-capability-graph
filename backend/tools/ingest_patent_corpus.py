"""把专利与标准工作簿转成与 arXiv 语料同格式的上游语料。

**为什么专利比 arXiv 更适合本项目。** arXiv 语料在缺口分析中被大量滤掉——189 对
被「JD 中提及过少」剔除，根因是英文学术语料与中文具身智能招聘市场谈的不是同一批
技术（cs.RO 含大量无人机与低空研究）。中文专利与中文 JD 的技术分布接近得多，
而且专利是**企业行为**，直接反映谁在投研发，比学术论文更靠近招聘决策。
论文第四章也把专利列为「最早可被系统化采集的结构化信号」。

**只有标题，没有摘要。** 工作簿的摘要列全为「无」，因此抽取只能基于标题。
中文标题平均 18–19 字，技术密度高（「具身智能模型的运动规划的方法」一句含两个
技术点）但覆盖面窄；英文标题平均 66–71 字符，长得多，且词表的英文别名覆盖率已达
91%，因此**两段一并送入匹配**。

英文标题是机器翻译，专有术语会失真（「具身智能」被译成 Self-body intelligent），
但 dexterous hand / humanoid robot / motion planning 这类标准术语译得准确，
净收益为正。失真带来的是漏检而非误检——译错的词匹配不上任何别名，不会造成假阳性。

**申请日而非公开日。** 专利从申请到公开有约 18 个月的审查期，公开日会把信号
系统性推后。申请日才是技术活动真正发生的时间，也是论文对信号锚点的定义。

**只取标注为「相关」的条目。** 关键词检索表里 7,469 条被标为「无关联」——那是
关键词召回的噪声（表里甚至有 1907 年的专利）。人工标注已经做过一轮筛选，
沿用它，不另起判据。

标准另存一份：标准发布意味着技术已进入规模化阶段，是比论文晚、比招聘早的一环，
用途与专利不同，不混进同一个语料池。

用法（backend 目录 / 容器内）：
    python -m tools.ingest_patent_corpus \\
        --workbook /srv/data/upstream/patent/专利+标准_终版.xlsx \\
        --out /srv/data/upstream/patent
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path

import openpyxl

TOOL_VERSION = "patent_corpus_v1"
RELEVANT = "相关"

# 两张专利表的列名不同，但语义一一对应。
PATENT_SHEETS = {
    "专利_关键词检索": {
        "title": "专利标题",
        "title_en": "英文标题",
        "filing_date": "申请日",
        "publication_date": "公开日",
        "applicant": "申请人",
        "ipc": "IPC分类号",
        "relevance": "与具身智能相关性",
        "doc_id": "公开号",
    },
    "专利_高级检索": {
        "title": "标题(中文)",
        "title_en": "标题(英文)",
        "filing_date": "申请日",
        "publication_date": "公开(公告)日",
        "applicant": "申请人(原始)",
        "ipc": "IPC分类号",
        "relevance": "与具身智能相关性",
        "doc_id": "公开(公告)号",
    },
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="导入专利与标准语料。")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument(
        "--include-unrelated",
        action="store_true",
        help="连同标注为「无关联」的条目一起导入。默认不导——那是关键词召回的噪声",
    )
    return parser.parse_args()


def normalise_date(value) -> str | None:
    """工作簿里的日期是 YYYYMMDD 数字或字符串，统一成 ISO。"""
    if value is None:
        return None
    text = str(value).strip().replace("-", "").replace("/", "")
    if len(text) < 8 or not text[:8].isdigit():
        return None
    year = int(text[:4])
    # 关键词检索表里混进了 1907 年这类明显越界的记录，一并挡掉。
    if not 1990 <= year <= 2030:
        return None
    return f"{text[:4]}-{text[4:6]}-{text[6:8]}"


def main() -> None:
    args = parse_args()
    workbook = openpyxl.load_workbook(args.workbook, read_only=True)
    args.out.mkdir(parents=True, exist_ok=True)

    seen: dict[str, dict] = {}
    stats: Counter[str] = Counter()
    for sheet_name, columns in PATENT_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        header = {name: index for index, name in enumerate(next(rows))}
        for row in rows:
            if not row or row[0] is None:
                continue
            stats["读取"] += 1

            def get(key: str, _row=row, _header=header, _columns=columns):
                column = _columns[key]
                return _row[_header[column]] if column in _header else None

            if not args.include_unrelated and str(get("relevance")).strip() != RELEVANT:
                stats["标注为无关联，跳过"] += 1
                continue
            title = str(get("title") or "").strip()
            if not title or title == "无":
                stats["无标题，跳过"] += 1
                continue
            # 申请日优先；专利公开滞后申请约 18 个月，用公开日会把信号系统性推后。
            filing = normalise_date(get("filing_date")) or normalise_date(
                get("publication_date")
            )
            if not filing:
                stats["无可用日期，跳过"] += 1
                continue
            title_en = str(get("title_en") or "").strip()
            if title_en in ("无", "None"):
                title_en = ""
            doc_id = str(get("doc_id") or "").strip()
            # 同一件专利的「申请公开」与「授权公告」是两条记录、两个公开号，
            # 按公开号去重会把它们当成两件。用「标题 + 申请日」作键才能合并。
            key = f"{title}|{filing}"
            if key in seen:
                stats["同一专利的重复公告，合并"] += 1
                continue
            seen[key] = {
                "arxiv_id": doc_id or key,  # 复用下游字段名，避免改动抽取与回测工具
                "published": filing,
                "updated": normalise_date(get("publication_date")) or filing,
                "title": title,
                # 摘要列全为「无」。英文标题放进 abstract 位：下游按
                # title + abstract 拼接后匹配，这样中英两段都参与，无需改动抽取工具。
                "abstract": title_en,
                "categories": [
                    part.strip()
                    for part in str(get("ipc") or "").split(";")
                    if part.strip() and part.strip() != "无"
                ],
                "primary_category": "patent",
                "applicant": str(get("applicant") or "").strip(),
                "source_sheet": sheet_name,
            }
            stats["纳入"] += 1

    papers = sorted(seen.values(), key=lambda item: item["published"], reverse=True)
    by_year: dict[str, list[dict]] = {}
    for paper in papers:
        by_year.setdefault(paper["published"][:4], []).append(paper)
    for year, rows in sorted(by_year.items()):
        path = args.out / f"arxiv_{year}.jsonl"
        with path.open("w", encoding="utf-8") as handle:
            for row in rows:
                handle.write(json.dumps(row, ensure_ascii=False) + "\n")

    counts = {year: len(rows) for year, rows in sorted(by_year.items())}
    (args.out / "_manifest.json").write_text(
        json.dumps(
            {
                "tool_version": TOOL_VERSION,
                "source": str(args.workbook.name),
                "anchor_field": "申请日（公开日仅在申请日缺失时兜底）",
                "abstract_available": False,
                "note": (
                    "工作簿摘要列全为「无」；抽取基于中文标题（约 18–19 字）"
                    "与英文标题（约 66–71 字符，机器翻译）两段"
                ),
                "document_count": len(papers),
                "by_year": counts,
                "stats": dict(stats),
            },
            ensure_ascii=False,
            indent=1,
        ),
        encoding="utf-8",
    )

    print(json.dumps(dict(stats), ensure_ascii=False, indent=2))
    print(f"\n纳入 {len(papers)} 件专利，按年分片写入 {args.out}")
    print("按年份分布：")
    for year, count in counts.items():
        print(f"   {year}: {count}")


if __name__ == "__main__":
    main()
