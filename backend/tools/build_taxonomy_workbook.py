"""由 v1.1 工作簿 + 治理变更集渲染新一版技术词工作簿（窗口 C-7 升版链路）。

词表的唯一权威来源是工作簿（每个节点都要能回溯到某一行电子表格），所以升版不能
直接改数据库，而是「基线工作簿 + 变更集 → 新工作簿 → stage → import」。
变更集是仓库内的 JSON，因此整个升版过程只依赖仓库文件，不依赖数据库状态，
任何人重跑都能得到逐字节相同的工作簿。

变更集的三种原语：
- `retire_terms`：把既有 L4 词的「可匹配」列置为「否」（保留行与谱系，只是不再参与 JD 匹配）
- `new_l3`：新增 L3 技术点
- `new_terms`：新增 L4 表面词

「收窄」不是独立原语——它等于「下线宽词 + 新增若干窄词形」，这样每条改动都能
在工作簿里逐行看到，不需要额外的规则表。

用法：
    python tools/build_taxonomy_workbook.py \
        --base data/source/20260810/core/技术词主数据_20260727.xlsx \
        --changeset data/governance/taxonomy_v1_2_changeset.json \
        --out data/governance/技术词主数据_v1.2.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

MATCHABLE_COLUMN = "可匹配"
TREATMENT_COLUMN = "治理动作"


def normalize_term(value: object) -> str:
    normalized = unicodedata.normalize("NFKC", str(value)).strip().casefold()
    return re.sub(r"\s+", " ", normalized)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="由基线工作簿与治理变更集渲染新版技术词工作簿。")
    parser.add_argument("--base", required=True, type=Path)
    parser.add_argument("--changeset", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    return parser.parse_args()


def sheet_table(workbook: openpyxl.Workbook, name: str) -> tuple[list[str], list[dict]]:
    worksheet = workbook[name]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(cell) for cell in next(rows)]
    records = [
        {key: ("" if value is None else value) for key, value in zip(header, row, strict=False)}
        for row in rows
        if any(cell not in (None, "") for cell in row)
    ]
    return header, records


def apply_changeset(
    base: Path, changeset: dict
) -> tuple[dict[str, list[str]], dict[str, list[dict]]]:
    workbook = openpyxl.load_workbook(base, read_only=True)
    headers: dict[str, list[str]] = {}
    tables: dict[str, list[dict]] = {}
    for name in ("L1技术域", "L2技术类", "L3技术点", "L4技术词"):
        headers[name], tables[name] = sheet_table(workbook, name)

    l2_by_code = {row["L2编码"]: row for row in tables["L2技术类"]}
    l3_by_code = {row["L3编码"]: row for row in tables["L3技术点"]}
    errors: list[str] = []

    # --- 下线：只改「可匹配」列，行与谱系保留 ---
    headers["L4技术词"] = headers["L4技术词"] + [MATCHABLE_COLUMN, TREATMENT_COLUMN]
    l4_index: dict[tuple[str, str], dict] = {}
    for row in tables["L4技术词"]:
        # 「碎片词」在 v1.1 里就不参与匹配，这里显式落到列上，让「可匹配」成为唯一口径。
        fragment = str(row.get("L4类型", "")).strip() == "碎片词"
        row[MATCHABLE_COLUMN] = "否" if fragment else "是"
        row[TREATMENT_COLUMN] = "inherited: 碎片词不参与匹配" if fragment else ""
        l4_index[(row["挂载L3编码"], normalize_term(row["技术词"]))] = row
    for item in changeset["retire_terms"]:
        key = (item["l3_code"], normalize_term(item["term"]))
        row = l4_index.get(key)
        if row is None:
            errors.append(f"下线目标不存在：{item['term']} @ {item['l3_code']}")
            continue
        row[MATCHABLE_COLUMN] = "否"
        row[TREATMENT_COLUMN] = f"retired: {item['reason']}"

    # --- 新增 L3 ---
    for item in changeset["new_l3"]:
        if item["code"] in l3_by_code:
            errors.append(f"新增L3已存在：{item['code']}")
            continue
        l2 = l2_by_code.get(item["l2_code"])
        if l2 is None:
            errors.append(f"新增L3挂载的L2不存在：{item['l2_code']}")
            continue
        row = {
            "L3编码": item["code"],
            "L3标准名": item["name"],
            "L2编码": item["l2_code"],
            "L2技术类": l2["技术类"],
            "L1编码": l2["所属L1"],
            "L1技术域": l2["所属技术域"],
            "L4成员数": 0,
            "簇类型": "治理新增L3",
            "备注": item.get("note", ""),
        }
        tables["L3技术点"].append(row)
        l3_by_code[item["code"]] = row

    # --- 新增 L4 表面词 ---
    for item in changeset["new_terms"]:
        l3 = l3_by_code.get(item["l3_code"])
        if l3 is None:
            errors.append(f"新增表面词挂载的L3不存在：{item['l3_code']}（词：{item['term']}）")
            continue
        key = (item["l3_code"], normalize_term(item["term"]))
        if key in l4_index:
            errors.append(f"新增表面词与既有词重复：{item['term']} @ {item['l3_code']}")
            continue
        row = {
            "技术词": item["term"],
            "L4类型": item.get("type", "细分词"),
            "挂载L3编码": item["l3_code"],
            "挂载L3名称": l3["L3标准名"],
            "L2编码": l3["L2编码"],
            "L2技术类": l3["L2技术类"],
            "L1编码": l3["L1编码"],
            "原层级(留痕)": "L4",
            "跨域调整": "",
            "命中来源": "词表治理",
            "来源明细(留痕)": item.get("reason", changeset["provenance"]["new_terms"]),
            MATCHABLE_COLUMN: "是",
            TREATMENT_COLUMN: f"added: {item.get('reason', '覆盖盲区补齐')}",
        }
        tables["L4技术词"].append(row)
        l4_index[key] = row

    # 同一 normalized 词形挂到多个 L3 会让匹配结果依赖别名 id 顺序，必须挡住。
    by_term: dict[str, list[str]] = {}
    for row in tables["L4技术词"]:
        if row[MATCHABLE_COLUMN] == "是":
            by_term.setdefault(normalize_term(row["技术词"]), []).append(row["挂载L3编码"])
    for term, codes in sorted(by_term.items()):
        if len(set(codes)) > 1:
            errors.append(f"词形跨节点重复：{term} → {sorted(set(codes))}")

    if errors:
        raise SystemExit("变更集校验失败：\n" + "\n".join(f"  - {item}" for item in errors))

    # L3 的 L4 成员数随新增词更新，保持工作簿自洽。
    member_counts: dict[str, int] = {}
    for row in tables["L4技术词"]:
        member_counts[row["挂载L3编码"]] = member_counts.get(row["挂载L3编码"], 0) + 1
    for row in tables["L3技术点"]:
        row["L4成员数"] = member_counts.get(row["L3编码"], 0)

    # 确定性排序：编码 → 词形，保证重跑产出逐字节一致。
    tables["L3技术点"].sort(key=lambda row: row["L3编码"])
    tables["L4技术词"].sort(key=lambda row: (row["挂载L3编码"], normalize_term(row["技术词"])))
    return headers, tables


def write_workbook(out: Path, headers: dict[str, list[str]], tables: dict[str, list[dict]]) -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for name in ("L1技术域", "L2技术类", "L3技术点", "L4技术词"):
        worksheet = workbook.create_sheet(name)
        worksheet.append(headers[name])
        for row in tables[name]:
            worksheet.append([row.get(column, "") for column in headers[name]])
    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out)


def main() -> None:
    args = parse_args()
    changeset = json.loads(args.changeset.read_text(encoding="utf-8"))
    headers, tables = apply_changeset(args.base, changeset)
    write_workbook(args.out, headers, tables)
    matchable = sum(1 for row in tables["L4技术词"] if row[MATCHABLE_COLUMN] == "是")
    print(
        json.dumps(
            {
                "target_version": changeset["target_version"],
                "output": str(args.out),
                "l1_count": len(tables["L1技术域"]),
                "l2_count": len(tables["L2技术类"]),
                "l3_count": len(tables["L3技术点"]),
                "l4_count": len(tables["L4技术词"]),
                "matchable_alias_count": matchable,
                "retired_count": len(changeset["retire_terms"]),
                "new_l3_count": len(changeset["new_l3"]),
                "new_term_count": len(changeset["new_terms"]),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
