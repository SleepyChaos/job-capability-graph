"""能力图谱 + 技术图谱 构建脚本（从数据到图谱的完整过程）。

数据源：
  - 具身智能岗位_清洗后_v3.xlsx  (岗位 3718 + 技能树字典 78 + 职业分类映射)
  - 具身智能企业数据_整合去重_完整.xlsx  (企业 634)
  - DB md_technology_node  (技术词 L1×7 / L2×43 / L3×229 / L4×1872)

图谱本体（两张图 + 一座桥）：
  技术域(技术图谱)：技术点 L3 → L2 → L1  (belongs_to 层级)
  能力域(能力图谱)：岗位 → 能力项 → 技能族  (requires_capability / belongs_to_family)
  桥梁：技能族 → 技术域 L2  (supports_domain，人工规则映射)
  附加：企业 → 岗位  (posts_job，公司名归一匹配)

产出（写入 ../data/processed/graph_build/）：
  nodes.json / edges.json / nodes.csv / edges.csv / stats.json
"""
from __future__ import annotations

import csv
import json
import os
import re
from collections import Counter, defaultdict

import openpyxl

from app.db.session import engine

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        "..", "..", "data", "processed", "graph_build")

JOB_XLSX = r"D:\揭榜挂帅\具身智能岗位_清洗后_v3.xlsx"
ORG_XLSX = r"D:\揭榜挂帅\具身智能企业数据_整合去重_完整.xlsx"

# 技能族 -> 技术域 L2 映射（人工规则，可迭代）
SKILL_FAMILY_TO_L2 = {
    "编程语言": "T5.03",   # 开发与部署工具链
    "AI框架": "T1.05",     # 学习与训练方法
    "机器人框架": "T5.01",  # 机器人操作系统
    "算法-感知": "T2.01",   # 视觉传感
    "算法-控制": "T1.03",   # 运动规划与控制
    "算法-AI": "T1.05",     # 学习与训练方法
    "算法-应用": "T1.07",   # 任务规划与推理
    "软件工程": "T5.03",    # 开发与部署工具链
    "嵌入式/硬件": "T3.05",  # 关节与驱动模组
    "仿真与数据": "T4.03",   # 仿真平台与环境
    "机械设计": "T3.07",    # 结构件与材料
    "其他": "T6.03",        # 检测认证与标准
}


def _norm_org_name(s):
    """归一化公司名（去空格/小写/去公司法务后缀）用于企业↔岗位匹配。"""
    if not s:
        return ""
    s = str(s).strip().lower().replace(" ", "").replace("　", "")
    for suf in ["股份有限公司", "有限责任公司", "有限公司", "股份公司", "集团"]:
        if s.endswith(suf):
            s = s[: -len(suf)]
    if s.endswith("科技"):
        s = s[: -len("科技")]
    elif s.endswith("技术"):
        s = s[: -len("技术")]
    elif s.endswith("机器人"):
        s = s[: -len("机器人")]
    return s


_STOP_NAMES = {
    # 城市
    "杭州", "北京", "上海", "深圳", "广州", "苏州", "无锡", "南京", "成都", "武汉",
    "西安", "天津", "重庆", "东莞", "佛山", "宁波", "青岛", "大连", "厦门", "郑州",
    "济南", "福州", "昆明", "南昌", "长沙", "合肥", "贵阳", "石家庄", "哈尔滨", "长春",
    "沈阳", "珠海", "中山", "惠州", "嘉兴", "绍兴", "温州", "台州", "金华", "常州",
    "南通", "扬州", "徐州", "盐城", "泰州", "镇江", "淮安", "连云港", "宿迁", "湖州",
    "芜湖", "泉州", "漳州", "莆田", "柳州", "桂林", "南宁", "太原", "呼和浩特", "兰州",
    "银川", "西宁", "乌鲁木齐", "拉萨", "海口", "三亚",
    # 省份 / 地区 / 国家
    "浙江", "江苏", "广东", "山东", "河南", "河北", "四川", "湖北", "湖南", "福建",
    "安徽", "江西", "辽宁", "陕西", "云南", "贵州", "山西", "甘肃", "青海", "海南",
    "黑龙江", "吉林", "内蒙古", "广西", "宁夏", "新疆", "西藏",
    "中国", "香港", "澳门", "台湾", "美国", "德国", "英国", "日本", "韩国", "新加坡",
    "法国", "加拿大", "瑞士", "荷兰", "瑞典", "以色列",
    # 通用公司后缀 / 杂词
    "有限公司", "有限责任公司", "股份有限公司", "科技", "技术", "集团", "公司",
    "智能", "机器人", "数字", "信息", "数据",
}


def _alias_keys(s):
    """把一个名称字符串拆成多个匹配 key（去括号/斜杠/顿号，过滤地名与通用词）。"""
    if not s:
        return set()
    parts = re.split(r"[（()）/、,，;；|]+", str(s))
    keys = set()
    for p in parts:
        p = p.strip()
        if not p:
            continue
        low = p.lower().replace(" ", "")
        if low in _STOP_NAMES:
            continue
        nk = _norm_org_name(p)
        if nk and nk not in _STOP_NAMES and len(nk) >= 2:
            keys.add(nk)
    return keys


def load_technologies():
    """从 DB 读技术词节点，返回 nodes(list) + 层级边。"""
    import sqlite3
    con = sqlite3.connect(".local/dev.db")
    cur = con.cursor()
    cur.execute("SELECT technology_node_id, technology_code, technology_name, "
                "parent_technology_node_id, level_code FROM md_technology_node")
    rows = cur.fetchall()
    con.close()
    nodes = {}
    for nid, code, name, pid, level in rows:
        nodes[nid] = {"code": code, "name": name, "level": level, "parent_id": pid}
    return nodes


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    nodes = []
    edges = []

    # ------------------------------------------------------------------
    # 1. 读岗位数据
    # ------------------------------------------------------------------
    wb = openpyxl.load_workbook(JOB_XLSX, read_only=True, data_only=True)
    ws = wb["岗位数据"]
    job_rows = list(ws.iter_rows(values_only=True))
    jheader = job_rows[0]
    jdata = job_rows[1:]
    jidx = {h: i for i, h in enumerate(jheader)}

    # 技能树字典 -> 能力项 / 技能族
    dic_rows = list(wb["技能树字典"].iter_rows(values_only=True))[1:]
    family_map = defaultdict(list)   # 技能族 -> [(能力项, 命中数)]
    for r in dic_rows:
        if r[0] and r[1]:
            family_map[str(r[0]).strip()].append((str(r[1]).strip(), r[2] or 0))

    wb.close()

    # ------------------------------------------------------------------
    # 2. 技术节点（L1-L3 纳入图，L4 仅统计不建节点）
    # ------------------------------------------------------------------
    tech_all = load_technologies()
    tech_by_code = {}
    l4_count = 0
    for nid, t in tech_all.items():
        if t["level"] == "L4":
            l4_count += 1
            continue
        tech_by_code[t["code"]] = t

    # 技术节点写入
    for code, t in tech_by_code.items():
        nodes.append({
            "id": f"tech:{code}", "type": "technology",
            "label": t["name"], "level": t["level"],
            "code": code,
        })
    # 技术层级边 L3->L2, L2->L1（通过 parent_id 反查 code）
    id_to_code = {nid: t["code"] for nid, t in tech_all.items()}
    for nid, t in tech_all.items():
        if t["level"] in ("L1", "L4"):
            continue
        pid = t["parent_id"]
        pcode = id_to_code.get(pid) if pid else None
        if pcode and pcode in tech_by_code:
            edges.append({
                "source": f"tech:{t['code']}", "target": f"tech:{pcode}",
                "type": "belongs_to",
            })

    # ------------------------------------------------------------------
    # 3. 技能族节点 + 能力项节点 + 能力项→技能族 边 + 技能族→L2 桥边
    # ------------------------------------------------------------------
    for fam, items in family_map.items():
        nodes.append({
            "id": f"fam:{fam}", "type": "capability_family",
            "label": fam, "family": fam, "cap_count": len(items),
        })
        l2 = SKILL_FAMILY_TO_L2.get(fam)
        if l2 and l2 in tech_by_code:
            edges.append({
                "source": f"fam:{fam}", "target": f"tech:{l2}",
                "type": "supports_domain",
            })
        for cap_name, cnt in items:
            nodes.append({
                "id": f"cap:{cap_name}", "type": "capability",
                "label": cap_name, "family": fam, "mention_count": int(cnt or 0),
            })
            edges.append({
                "source": f"cap:{cap_name}", "target": f"fam:{fam}",
                "type": "belongs_to_family",
            })

    # ------------------------------------------------------------------
    # 4. 岗位节点 + 岗位→能力项 边（解析技能标签列）
    # ------------------------------------------------------------------
    job_has_tag = 0
    job_cap_edge_count = 0
    for r in jdata:
        occ_id = str(r[jidx["occ_id"]]).strip() if r[jidx["occ_id"]] else ""
        title = str(r[jidx["岗位"]] or "").strip()
        if not occ_id:
            continue
        nodes.append({
            "id": f"job:{occ_id}", "type": "job", "label": title,
            "title": title,
            "level": r[jidx["能力等级"]],
            "career_dir": r[jidx["职业方向"]],
            "career_kind": r[jidx["职业种类"]],
            "edu": r[jidx["学历(标准化)"]],
            "exp": r[jidx["经验(标准化)"]],
            "chain": r[jidx["产业链层级"]],
            "company": r[jidx["公司"]],
            "company_norm": _norm_org_name(r[jidx["公司"]]),
            "linked_company": r[jidx["关联公司"]],
            "linked_norm": _norm_org_name(r[jidx["关联公司"]]),
        })
        tag = r[jidx["技能标签"]]
        if tag:
            parts = [p.strip() for p in str(tag).split(";") if p.strip()]
            if parts:
                job_has_tag += 1
                for p in parts:
                    edges.append({
                        "source": f"job:{occ_id}", "target": f"cap:{p}",
                        "type": "requires_capability",
                    })
                    job_cap_edge_count += 1

    # ------------------------------------------------------------------
    # 5. 企业节点 + 企业→岗位 边（公司名归一匹配）
    # ------------------------------------------------------------------
    owb = openpyxl.load_workbook(ORG_XLSX, read_only=True, data_only=True)
    ows = owb["Sheet1"]
    org_rows = list(ows.iter_rows(values_only=True))
    oheader = org_rows[0]
    odata = org_rows[1:]
    oidx = {h: i for i, h in enumerate(oheader)}
    owb.close()

    # 建 岗位 公司名/别名 -> 岗位 occ_id 索引
    company_to_jobs = defaultdict(set)
    for r in jdata:
        occ_id = str(r[jidx["occ_id"]]).strip() if r[jidx["occ_id"]] else ""
        if not occ_id:
            continue
        for key in (r[jidx["公司"]], r[jidx["关联公司"]]):
            for nk in _alias_keys(key):
                company_to_jobs[nk].add(occ_id)

    org_matched = 0
    org_seq = 0
    for r in odata:
        name = str(r[oidx["企业名称"]] or "").strip()
        if not name:
            continue
        org_seq += 1
        nodes.append({
            "id": f"org:{org_seq:04d}", "type": "organization", "label": name,
            "chain": r[oidx["产业链(12类标准)"]],
            "tier": r[oidx["层级"]],
            "segment": r[oidx["细分领域"]],
            "job_count": r[oidx["在聘岗位数量"]],
        })
        org_keys = _alias_keys(name) | _alias_keys(r[oidx["英文名/别名"]])
        matched = set()
        for k in org_keys:
            matched |= company_to_jobs.get(k, set())
        if matched:
            org_matched += 1
            for occ_id in matched:
                edges.append({
                    "source": f"org:{org_seq:04d}", "target": f"job:{occ_id}",
                    "type": "posts_job",
                })

    # ------------------------------------------------------------------
    # 6. 统计 + 产出
    # ------------------------------------------------------------------
    ntype = Counter(n["type"] for n in nodes)
    etype = Counter(e["type"] for e in edges)
    stats = {
        "node_count": {k: v for k, v in ntype.items()},
        "node_total": len(nodes),
        "edge_count": {k: v for k, v in etype.items()},
        "edge_total": len(edges),
        "job_total": len(jdata),
        "job_with_skill_tag": job_has_tag,
        "job_skill_tag_rate": round(job_has_tag / len(jdata), 4),
        "capability_count": len(family_map),
        "capability_item_count": sum(len(v) for v in family_map.values()),
        "skill_family_count": len(family_map),
        "tech_l1": sum(1 for t in tech_by_code.values() if t["level"] == "L1"),
        "tech_l2": sum(1 for t in tech_by_code.values() if t["level"] == "L2"),
        "tech_l3": sum(1 for t in tech_by_code.values() if t["level"] == "L3"),
        "tech_l4_not_in_graph": l4_count,
        "org_total": len(odata),
        "org_matched_job": org_matched,
    }

    def write_json(path, data):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=1)

    write_json(os.path.join(OUT_DIR, "nodes.json"), nodes)
    write_json(os.path.join(OUT_DIR, "edges.json"), edges)
    write_json(os.path.join(OUT_DIR, "stats.json"), stats)

    # CSV
    node_fields = ["id", "type", "label", "level", "family", "mention_count",
                   "title", "career_dir", "career_kind", "chain", "segment"]
    with open(os.path.join(OUT_DIR, "nodes.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=node_fields, extrasaction="ignore")
        w.writeheader()
        for n in nodes:
            w.writerow(n)
    with open(os.path.join(OUT_DIR, "edges.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["source", "target", "type"])
        for e in edges:
            w.writerow([e["source"], e["target"], e["type"]])

    print("=== 图谱构建完成 ===")
    print("节点:", dict(ntype))
    print("边:", dict(etype))
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    print("产出目录:", OUT_DIR)


if __name__ == "__main__":
    main()
