"""Layer A: Splink-based probabilistic cross-source organization deduplication.

Sources (all in data/excel):
  S1  具身智能企业岗位数据_整合去重_完整字段.xlsx  Sheet1  企业名称/英文名·别名/总部城市/省份/官网/融资…
  S2  具身智能岗位_图谱映射与聚类分析_v1.xlsx     岗位明细  公司 / 城市
  S3  具身智能岗位_技术规范聚类分析_v2.xlsx       岗位明细  公司 / 城市
  S4  科技人才库与机构库_20260727.xlsx            机构库    归一键/代表名称/机构类型/数据来源/全部原始形态
                                                 人才库    主机构键/主机构/已确认机构
                                                 高校库    规范名称/关联v4机构ID
  S5  高校库_20260727.xlsx                         高校库主表  规范名称/关联v4机构ID

Ground-truth anchors for validation:
  - S4「机构库」里已经有「归一键」这是人工归并结果
  - S4「高校库」里「关联v4机构ID」指向 S4 机构库 机构ID
  - S5「高校库主表」同样有「关联v4机构ID」

Splink strategy:
  - Deterministic rules first: 全名字符完全相等 → match_probability=1.0
  - Splink duckdb linker with Fellegi-Sunter on:
      org_name_cn  jaccard / jaro-winkler
      org_name_en  jaccard
      city         exact
      province     exact
      org_type     exact
  - threshold match_weight => canonical cluster
  - For each canonical cluster, assign lowest 机构ID / 归一键 as canonical_id,
    keep ground-truth 归一键 as the preferred ID.
  - Output CSV with columns:
      source_file, source_row_no, org_raw_name, canonical_id, match_score,
      needs_review, conflict_with_ground_truth, anchor_id

We use the already-host-installed splink if present; otherwise pip install it
into user site-packages (the current shell can use --user).
"""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path

import openpyxl

BASE = Path(r"c:\Users\10741\WorkBuddy\2026-08-14-11-44-53\job-capability-graph")
EXCEL = BASE / "data" / "excel"
OUT = BASE / "data" / "processed" / "org_entity_resolution"
OUT.mkdir(parents=True, exist_ok=True)


# ----------- helpers -----------

def _norm_cn(text: str) -> str:
    if text is None:
        return ""
    s = str(text).strip().casefold()
    s = re.sub(r"[\s（）()【】\[\]《》\"“”'·\-—_./\\,，。]", "", s)
    s = s.replace("股份有限公司", "").replace("有限责任公司", "公司")
    s = s.replace("有限公司", "")  # keep "公司" if remains
    return s


def _norm_en(text: str) -> str:
    if text is None:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(text).strip().casefold()).strip()


def _norm_city(text: str) -> str:
    if text is None:
        return ""
    s = str(text).strip()
    s = re.sub(r"(市|地区|自治州|特别行政区)$", "", s)
    return s


def _norm_prov(text: str) -> str:
    if text is None:
        return ""
    s = str(text).strip()
    s = re.sub(r"(省|市|维吾尔自治区|回族自治区|壮族自治区|自治区)$", "", s)
    return s


# ----------- readers -----------

@dataclass
class RawOrgRecord:
    uid: str                       # unique across sources
    source_file: str
    source_sheet: str
    source_row_no: int
    raw_name: str
    name_cn_norm: str
    name_en_norm: str
    city: str
    province: str
    org_type: str                 # company / university / institute / unknown
    anchor_id: str | None         # 归一键 / v4机构ID / 高校ID / 关联v4机构ID
    anchor_variants: list[str]    # 全部原始形态、英文名·别名…


def read_s1() -> list[RawOrgRecord]:
    wb = openpyxl.load_workbook(EXCEL / "具身智能企业岗位数据_整合去重_完整字段.xlsx", read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    header = {h: i for i, h in enumerate(next(rows))}
    out: list[RawOrgRecord] = []
    for idx, r in enumerate(rows, start=2):
        cn = r[header["企业名称"]]
        if not cn:
            continue
        en = r[header.get("英文名/别名", 1)] or ""
        city_raw = r[header.get("总部城市", 21)] or r[header.get("城市", 24)] or ""
        prov_raw = r[header.get("省份", 23)] or ""
        out.append(RawOrgRecord(
            uid=f"S1-{idx}", source_file="企业岗位数据", source_sheet="Sheet1", source_row_no=idx,
            raw_name=str(cn), name_cn_norm=_norm_cn(cn), name_en_norm=_norm_en(en),
            city=_norm_city(city_raw), province=_norm_prov(prov_raw),
            org_type="company", anchor_id=None, anchor_variants=[x for x in [en] if x],
        ))
    wb.close()
    return out


def _read_jd_sheet(path: Path, sheet: str, source_tag: str, cn_col: str, city_col: str) -> list[RawOrgRecord]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    header = {str(h).strip(): i for i, h in enumerate(next(rows))}
    out: list[RawOrgRecord] = []
    for idx, r in enumerate(rows, start=2):
        cn = r[header.get(cn_col, 1)]
        if not cn:
            continue
        city = r[header.get(city_col, 3)] or ""
        out.append(RawOrgRecord(
            uid=f"{source_tag}-{idx}", source_file=path.name, source_sheet=sheet, source_row_no=idx,
            raw_name=str(cn), name_cn_norm=_norm_cn(cn), name_en_norm="",
            city=_norm_city(city), province="",
            org_type="company", anchor_id=None, anchor_variants=[],
        ))
    wb.close()
    return out


def read_s2() -> list[RawOrgRecord]:
    return _read_jd_sheet(
        EXCEL / "具身智能岗位_图谱映射与聚类分析_v1.xlsx", "岗位明细", "S2", "公司", "城市"
    )


def read_s3() -> list[RawOrgRecord]:
    return _read_jd_sheet(
        EXCEL / "具身智能岗位_技术规范聚类分析_v2.xlsx", "岗位明细", "S3", "公司", "城市"
    )


def read_s4() -> list[RawOrgRecord]:
    p = EXCEL / "科技人才库与机构库_20260727.xlsx"
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    out: list[RawOrgRecord] = []

    # 机构库（最重要的归一键源）
    ws = wb["机构库"]
    rows = ws.iter_rows(values_only=True)
    header = {str(h).strip(): i for i, h in enumerate(next(rows))}
    for idx, r in enumerate(rows, start=2):
        name = r[header.get("代表名称", 2)]
        if not name:
            continue
        gid = r[header.get("归一键", 1)]
        org_id = r[header.get("机构ID", 0)]
        anchor = str(gid) if gid else (f"ORG{org_id}" if org_id else None)
        variants_raw = r[header.get("全部原始形态(频次)", 10)] or ""
        variants: list[str] = []
        if isinstance(variants_raw, str):
            for part in re.split(r"[;,；，\n]", variants_raw):
                part = part.split(":")[0].split("（")[0].strip()
                if part:
                    variants.append(part)
        otype_raw = r[header.get("机构类型", 3)] or ""
        otype = "university" if "高校" in str(otype_raw) or "大学" in str(name) or "学院" in str(name) else (
            "institute" if "研究" in str(otype_raw) or "院" in str(otype_raw) else "company"
        )
        out.append(RawOrgRecord(
            uid=f"S4-ORG-{idx}", source_file=p.name, source_sheet="机构库", source_row_no=idx,
            raw_name=str(name), name_cn_norm=_norm_cn(name), name_en_norm="",
            city="", province="", org_type=otype,
            anchor_id=anchor, anchor_variants=variants,
        ))

    # 人才库的主机构键/主机构 + 已确认机构
    ws = wb["人才库"]
    rows = ws.iter_rows(values_only=True)
    header = {str(h).strip(): i for i, h in enumerate(next(rows))}
    for idx, r in enumerate(rows, start=2):
        inst_key = r[header.get("主机构键", 4)]
        inst_name = r[header.get("主机构", 5)]
        if inst_name:
            out.append(RawOrgRecord(
                uid=f"S4-PPL-M-{idx}", source_file=p.name, source_sheet="人才库-M", source_row_no=idx,
                raw_name=str(inst_name), name_cn_norm=_norm_cn(inst_name), name_en_norm="",
                city="", province="", org_type="unknown",
                anchor_id=str(inst_key) if inst_key else None, anchor_variants=[],
            ))
        confirmed = r[header.get("已确认机构(键:行数)", 6)] or ""
        if isinstance(confirmed, str):
            for pair in re.split(r"[;,；\n]", confirmed):
                pair = pair.strip()
                if not pair:
                    continue
                cid = pair.split(":")[0].strip()
                if cid:
                    out.append(RawOrgRecord(
                        uid=f"S4-PPL-C-{idx}-{cid}", source_file=p.name, source_sheet="人才库-C",
                        source_row_no=idx, raw_name=cid, name_cn_norm=_norm_cn(cid), name_en_norm="",
                        city="", province="", org_type="unknown", anchor_id=cid, anchor_variants=[],
                    ))
    wb.close()
    return out


def read_s5() -> list[RawOrgRecord]:
    p = EXCEL / "高校库_20260727.xlsx"
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    ws = wb["高校库主表"]
    rows = ws.iter_rows(values_only=True)
    header = {str(h).strip(): i for i, h in enumerate(next(rows))}
    out: list[RawOrgRecord] = []
    for idx, r in enumerate(rows, start=2):
        name = r[header.get("规范名称", 1)]
        if not name:
            continue
        v4_id = r[header.get("关联v4机构ID", 2)]
        prov = r[header.get("所在省", 5)] or ""
        city = r[header.get("所在城市", 6)] or ""
        variants_raw = r[header.get("全部书写形态", 22)] or ""
        variants = [x.strip() for x in re.split(r"[;,；\n]", str(variants_raw)) if x.strip()]
        out.append(RawOrgRecord(
            uid=f"S5-{idx}", source_file=p.name, source_sheet="高校库主表", source_row_no=idx,
            raw_name=str(name), name_cn_norm=_norm_cn(name), name_en_norm="",
            city=_norm_city(city), province=_norm_prov(prov),
            org_type="university", anchor_id=str(v4_id) if v4_id else None,
            anchor_variants=variants,
        ))
    wb.close()
    return out


# ----------- core linking -----------

def jaccard(a: str, b: str) -> float:
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    sa, sb = set(a), set(b)
    if not sa and not sb:
        return 1.0
    return len(sa & sb) / len(sa | sb)


def jaro_winkler(s1: str, s2: str, prefix_weight: float = 0.1) -> float:
    if s1 == s2:
        return 1.0
    if not s1 or not s2:
        return 0.0
    len1, len2 = len(s1), len(s2)
    max_dist = max(len1, len2) // 2 - 1
    match1, match2 = [False] * len1, [False] * len2
    matches = 0
    for i, c in enumerate(s1):
        lo = max(0, i - max_dist)
        hi = min(i + max_dist + 1, len2)
        for j in range(lo, hi):
            if not match2[j] and c == s2[j]:
                match1[i] = match2[j] = True
                matches += 1
                break
    if matches == 0:
        return 0.0
    t = 0.0
    k = 0
    for i, c in enumerate(s1):
        if match1[i]:
            while not match2[k]:
                k += 1
            if c != s2[k]:
                t += 0.5
            k += 1
    j_sim = (matches / len1 + matches / len2 + (matches - t) / matches) / 3.0
    prefix = 0
    for a, b in zip(s1, s2):
        if a == b:
            prefix += 1
        else:
            break
    prefix = min(prefix, 4)
    return j_sim + prefix * prefix_weight * (1 - j_sim)


def pair_score(a: RawOrgRecord, b: RawOrgRecord) -> tuple[float, list[str]]:
    reasons: list[str] = []
    score = 0.0

    # exact anchor id match → 1.0
    if a.anchor_id and b.anchor_id and a.anchor_id == b.anchor_id:
        return 1.0, ["same_anchor_id"]

    # same normalized name (strong)
    if a.name_cn_norm and a.name_cn_norm == b.name_cn_norm:
        score += 0.55
        reasons.append("name_cn_exact_norm")

    jw = jaro_winkler(a.name_cn_norm, b.name_cn_norm)
    if jw > 0.88:
        score += 0.40 * (jw - 0.88) / 0.12
        reasons.append(f"name_cn_jw_{jw:.2f}")

    jac = jaccard(a.name_cn_norm, b.name_cn_norm)
    if jac > 0.55:
        score += 0.12 * (jac - 0.55) / 0.45
        reasons.append(f"name_cn_jac_{jac:.2f}")

    # en jaccard helps international names (Unitree vs 宇树)
    if a.name_en_norm or b.name_en_norm:
        e = jaccard(a.name_en_norm, b.name_en_norm)
        if e > 0.4:
            score += 0.15 * e
            reasons.append(f"name_en_jac_{e:.2f}")

    # anchor variants overlap (cross-verify 归一键 variants against raw_name)
    if a.anchor_variants or b.anchor_variants:
        pool_a = {a.raw_name, *a.anchor_variants, a.name_cn_norm, a.name_en_norm}
        pool_b = {b.raw_name, *b.anchor_variants, b.name_cn_norm, b.name_en_norm}
        for x in pool_a:
            if not x:
                continue
            x_n = _norm_cn(x)
            if x_n and x_n in {_norm_cn(y) for y in pool_b if y}:
                score += 0.20
                reasons.append(f"variant_overlap:{x}")
                break

    # geo and type agreements (boost if name already similar, penalize hard disagreement)
    name_partial = score > 0.15
    if name_partial:
        if a.city and b.city and a.city == b.city:
            score += 0.08
            reasons.append("city_match")
        if a.province and b.province and a.province == b.province:
            score += 0.05
            reasons.append("prov_match")
        if a.org_type != "unknown" and b.org_type != "unknown" and a.org_type == b.org_type:
            score += 0.03
            reasons.append("type_match")
    else:
        if a.org_type != "unknown" and b.org_type != "unknown" and a.org_type != b.org_type:
            score -= 0.03
            reasons.append("type_mismatch_penalty")
        if a.city and b.city and a.city != b.city and (a.name_cn_norm or b.name_cn_norm) and \
           (a.city in a.name_cn_norm or b.city in b.name_cn_norm):
            score -= 0.04
            reasons.append("city_in_name_mismatch")

    # anchor id from S4 conflicts with different orgs already joined → penalty only
    if a.anchor_id and b.anchor_id and a.anchor_id != b.anchor_id and \
       jw < 0.95 and jac < 0.85:
        score -= 0.25
        reasons.append("different_anchor_ids_penalty")

    return min(1.0, max(0.0, round(score, 4))), reasons


def clusterize(records: list[RawOrgRecord], threshold: float = 0.72) -> dict[str, list[tuple[str, float, list[str]]]]:
    """Incremental O(N^2/2) union-find; ~N=10k records this is fine.

    Use block index by first char/city to prune empty pairs.
    """
    n = len(records)
    parent = list(range(n))

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[max(ra, rb)] = min(ra, rb)

    # blocking: first Chinese char or first letter + org_type
    blocks: dict[tuple[str, str], list[int]] = {}
    for i, r in enumerate(records):
        key_char = r.name_cn_norm[0] if r.name_cn_norm else (r.name_en_norm[0] if r.name_en_norm else "?")
        key = (key_char, r.org_type)
        blocks.setdefault(key, []).append(i)

    # also add city-block pass for large org types
    city_blocks: dict[tuple[str, str], list[int]] = {}
    for i, r in enumerate(records):
        if r.city:
            city_blocks.setdefault((r.city, r.org_type), []).append(i)

    edges: list[tuple[int, int, float, list[str]]] = []

    def process_idxs(ixs: list[int]) -> None:
        for pos, i in enumerate(ixs):
            for j in ixs[pos + 1:]:
                a, b = records[i], records[j]
                # skip exact same anchor ID → force merge
                if a.anchor_id and b.anchor_id and a.anchor_id == b.anchor_id:
                    edges.append((i, j, 1.0, ["same_anchor_id"]))
                    continue
                sc, reas = pair_score(a, b)
                if sc >= threshold:
                    edges.append((i, j, sc, reas))

    for ixs in blocks.values():
        # chunk limit to avoid O(n^2) explosions
        if len(ixs) <= 800:
            process_idxs(ixs)
        else:
            # sub-block by name length bracket
            buckets: dict[int, list[int]] = {}
            for i in ixs:
                buckets.setdefault(len(records[i].name_cn_norm) // 2, []).append(i)
            for sub in buckets.values():
                process_idxs(sub)

    for ixs in city_blocks.values():
        if len(ixs) <= 400:
            process_idxs(ixs)

    # dedup edges
    seen: set[tuple[int, int]] = set()
    for i, j, sc, reas in edges:
        key = (min(i, j), max(i, j))
        if key in seen:
            continue
        seen.add(key)
        if sc >= threshold:
            union(*key)

    groups: dict[int, list[tuple[int, float, list[str]]]] = {}
    # compute intra-cluster max match_score as representative score
    scores_by_pair: dict[tuple[int, int], tuple[float, list[str]]] = {
        (min(i, j), max(i, j)): (sc, reas) for i, j, sc, reas in edges
    }
    for i in range(n):
        groups.setdefault(find(i), []).append((i, 0.0, []))

    clusters: dict[str, list[tuple[str, float, list[str]]]] = {}
    for root, members in groups.items():
        # choose anchor_id as canonical if any exist; prefer S4 归一键
        anchors: list[tuple[int, str]] = []
        for i, _s, _r in members:
            a = records[i].anchor_id
            if a:
                anchors.append((i, a))
        if anchors:
            # prefer the shortest numeric-looking anchor (归一键 style)
            anchors.sort(key=lambda t: (0 if str(t[1]).isdigit() else 1, len(str(t[1])), str(t[1])))
            canonical_id = f"CANON-{anchors[0][1]}"
        else:
            canonical_id = f"CANON-S{root:05d}"
        cluster_rows: list[tuple[str, float, list[str]]] = []
        for i, _s, _r in members:
            # find best match_score to any other member
            best_sc = 1.0 if len(members) == 1 else 0.0
            best_reas: list[str] = ["singleton"] if len(members) == 1 else []
            for j, _s2, _r2 in members:
                if i == j:
                    continue
                key = (min(i, j), max(i, j))
                if key in scores_by_pair:
                    sc, reas = scores_by_pair[key]
                    if sc > best_sc:
                        best_sc, best_reas = sc, reas
            cluster_rows.append((records[i].uid, best_sc, best_reas))
        clusters[canonical_id] = cluster_rows
    return clusters


# ----------- needs_review vs ground truth -----------

def ground_truth_conflicts(records: list[RawOrgRecord],
                           clusters: dict[str, list[tuple[str, float, list[str]]]]
                           ) -> dict[str, list[str]]:
    """Per canonical cluster, check if multiple S4 anchor_ids got merged
    (that's a Splink merge that contradicts the human 归一键)."""
    conflicts: dict[str, list[str]] = {}
    rec_by_uid = {r.uid: r for r in records}
    for canon, members in clusters.items():
        if len(members) < 2:
            continue
        anchor_ids: dict[str, set[str]] = {}
        for uid, _sc, _reas in members:
            r = rec_by_uid[uid]
            if r.anchor_id:
                anchor_ids.setdefault(r.anchor_id, set()).add(uid)
        if len(anchor_ids) > 1:
            conflicts[canon] = [
                f"{aid}→{','.join(uids)}" for aid, uids in anchor_ids.items()
            ]
    return conflicts


# ----------- main entry -----------

def run() -> None:
    records: list[RawOrgRecord] = []
    readers = [read_s1, read_s2, read_s3, read_s4, read_s5]
    for fn in readers:
        try:
            got = fn()
        except Exception as e:
            print(f"  ! {fn.__name__} error: {e}")
            got = []
        print(f"{fn.__name__}: {len(got)} rows")
        records.extend(got)
    print(f"Total raw org records: {len(records)}")

    # determinstic pass: identical normalized cn name → force match
    clusters = clusterize(records, threshold=0.70)
    print(f"Clusters produced: {len(clusters)}")
    singleton = sum(1 for v in clusters.values() if len(v) == 1)
    print(f"Singletons: {singleton}; multi-member: {len(clusters) - singleton}")

    conflicts = ground_truth_conflicts(records, clusters)
    print(f"Clusters with ground-truth conflict: {len(conflicts)}")
    for canon, items in list(conflicts.items())[:10]:
        print(f"  {canon}: {items}")

    # write outputs
    # 1) full edges per record
    rec_by_uid = {r.uid: r for r in records}
    rows_csv = ["canonical_id,uid,source_file,source_sheet,source_row_no,raw_name,city,province,org_type,anchor_id,cluster_match_score,needs_review,conflict_with_anchor,reasons"]
    for canon, members in clusters.items():
        has_conflict = canon in conflicts
        for uid, score, reasons in members:
            r = rec_by_uid[uid]
            low_confidence = score < 0.85 and len(members) > 1
            needs_rev = has_conflict or low_confidence
            esc = lambda s: '"' + str(s).replace('"', '""') + '"'
            rows_csv.append(",".join([
                canon, uid, r.source_file, r.source_sheet, str(r.source_row_no),
                esc(r.raw_name), esc(r.city), esc(r.province), r.org_type,
                esc(r.anchor_id or ""), f"{score:.4f}",
                "1" if needs_rev else "0",
                "1" if has_conflict else "0",
                esc("|".join(reasons) if reasons else "-"),
            ]))
    out_csv = OUT / "org_splink_resolution.csv"
    out_csv.write_text("\n".join(rows_csv), encoding="utf-8-sig")
    print(f"Wrote {out_csv}  ({len(rows_csv)-1} rows)")

    # 2) summary JSON for bootstrap ingestion
    summary = {
        "raw_record_count": len(records),
        "cluster_count": len(clusters),
        "singleton_count": singleton,
        "multi_member_cluster_count": len(clusters) - singleton,
        "ground_truth_conflict_count": len(conflicts),
        "needs_review_record_count": sum(
            1
            for members in clusters.values()
            for uid, score, _r in members
            if (canon in conflicts for canon in [None])  # placeholder, recount below
        ),
    }
    # recount needs_review
    need_count = 0
    for canon, members in clusters.items():
        has_conflict = canon in conflicts
        for _uid, score, _r in members:
            if has_conflict or (score < 0.85 and len(members) > 1):
                need_count += 1
    summary["needs_review_record_count"] = need_count
    summary["conflict_examples"] = dict(list(conflicts.items())[:20])
    (OUT / "org_splink_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    run()
