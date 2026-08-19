"""窗口 C：词表治理链路与歧义规则 v2 的回归。

覆盖三处容易悄悄退化的行为：
1. 变更集渲染器必须挡住重复词与跨节点同词形（否则匹配结果会依赖别名 id 顺序）；
2. 下线是「置可匹配为否」而不是删行——谱系与来源留痕必须保留；
3. 歧义规则 v2 的语境窗口是「命中词所在句」而不是整段，且语境词按权重累计。
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from app.modules.job.parsing_service import (
    AMBIGUITY_RULE_DEFINITIONS,
    MARKER_WEIGHT_THRESHOLD,
    JobParsingService,
    sentence_around,
)
from tools.build_taxonomy_workbook import apply_changeset

REPO_ROOT = Path(__file__).resolve().parents[2]
BASE_WORKBOOK = REPO_ROOT / "data/source/20260810/core/技术词主数据_20260727.xlsx"
CHANGESET = REPO_ROOT / "data/governance/taxonomy_v1_2_changeset.json"


def _changeset() -> dict:
    return json.loads(CHANGESET.read_text(encoding="utf-8"))


@pytest.mark.skipif(not BASE_WORKBOOK.exists(), reason="缺少基线工作簿")
def test_changeset_renders_v1_2_with_retirements_preserved_as_rows() -> None:
    headers, tables = apply_changeset(BASE_WORKBOOK, _changeset())

    l4_by_key = {(row["挂载L3编码"], row["技术词"]): row for row in tables["L4技术词"]}
    # 下线保留行：谱系、来源留痕都还在，只是不参与匹配。
    retired = l4_by_key[("T1.01.11", "大模型")]
    assert retired["可匹配"] == "否"
    assert retired["治理动作"].startswith("retired:")
    assert retired["L2编码"] == "T1.01"

    # 新增词与新增 L3 都挂上了。
    assert l4_by_key[("T5.01.01", "ROS2")]["可匹配"] == "是"
    l3_codes = {row["L3编码"] for row in tables["L3技术点"]}
    assert {"T5.02.02", "T5.03.03", "T1.03.14"} <= l3_codes

    # 「碎片词」在 v1.1 就不参与匹配，升版后不能被误开成可匹配。
    fragments = [row for row in tables["L4技术词"] if row["L4类型"] == "碎片词"]
    assert fragments and all(row["可匹配"] == "否" for row in fragments)

    # 同一词形只能挂一个节点，否则命中归属会依赖别名 id 顺序。
    matchable: dict[str, str] = {}
    for row in tables["L4技术词"]:
        if row["可匹配"] == "是":
            term = str(row["技术词"]).strip().casefold()
            assert matchable.setdefault(term, row["挂载L3编码"]) == row["挂载L3编码"]

    assert "可匹配" in headers["L4技术词"]


@pytest.mark.skipif(not BASE_WORKBOOK.exists(), reason="缺少基线工作簿")
def test_changeset_rejects_duplicate_new_term(tmp_path: Path) -> None:
    changeset = _changeset()
    # 「ROS」在 v1.1 已存在，重复新增必须报错而不是静默产生两行。
    changeset["new_terms"].append({"term": "ROS", "l3_code": "T5.01.01", "type": "组合词"})
    with pytest.raises(SystemExit, match="重复"):
        apply_changeset(BASE_WORKBOOK, changeset)


@pytest.mark.skipif(not BASE_WORKBOOK.exists(), reason="缺少基线工作簿")
def test_rendered_workbook_roundtrips_through_openpyxl(tmp_path: Path) -> None:
    from tools.build_taxonomy_workbook import write_workbook

    headers, tables = apply_changeset(BASE_WORKBOOK, _changeset())
    out = tmp_path / "v1_2.xlsx"
    write_workbook(out, headers, tables)
    workbook = openpyxl.load_workbook(out, read_only=True)
    assert set(workbook.sheetnames) == {"L1技术域", "L2技术类", "L3技术点", "L4技术词"}
    assert workbook["L3技术点"].max_row == len(tables["L3技术点"]) + 1
    assert workbook["L4技术词"].max_row == len(tables["L4技术词"]) + 1


def test_sentence_around_stops_at_sentence_boundary() -> None:
    text = "熟悉机器人整机集成。负责控制系统的运维与巡检。要求本科以上学历。"
    start = text.index("控制系统")
    window = sentence_around(text, start, start + len("控制系统"))
    # 「机器人」在上一句，不应进入语境窗口——这正是 v1 放行大量误报的原因。
    assert "机器人" not in window
    assert "运维与巡检" in window


def test_marker_weight_needs_two_weak_markers_or_one_strong() -> None:
    markers = AMBIGUITY_RULE_DEFINITIONS[("控制系统", "T1.03.12")][1]
    payload = [{"term": term, "weight": weight} for term, weight in markers]
    weigh = JobParsingService._marker_weight

    # 「机器人」是被降权的泛触发词，单独出现不足以放行。
    assert weigh("负责机器人控制系统的日常维护", payload) < MARKER_WEIGHT_THRESHOLD
    # 一个强词即可放行。
    assert weigh("负责伺服控制系统调试", payload) >= MARKER_WEIGHT_THRESHOLD
    # 两个弱词累计也可以放行。
    assert weigh("嵌入式实时控制系统开发", payload) >= MARKER_WEIGHT_THRESHOLD


def test_marker_weight_accepts_legacy_string_markers() -> None:
    assert JobParsingService._marker_weight("具身智能方向", ["具身", "vla"]) == 1.0


def test_occupation_gate_separates_non_embodied_roles() -> None:
    """职能门判的是「非具身智能技术岗」，不是「非技术岗」。"""
    from app.modules.extraction.occupation import classify_occupation

    for title in ("灵巧手销售经理", "上海 整机产品经理", "上海 smt工艺工程师", "高级云运维工程师"):
        decision = classify_occupation(title)
        assert not decision.is_embodied_technical
        assert decision.reason_code.startswith("non_embodied_occupation:")
        assert decision.matched_term

    technical = ("slam算法工程师", "嵌入式软件开发工程师", "机器人运控工程师")
    for title in technical:
        decision = classify_occupation(title)
        assert decision.is_embodied_technical
        assert decision.reason_code is None


def test_occupation_gate_is_deterministic_across_multi_category_hits() -> None:
    """同时命中多个类别时取字典序最小的类别，保证同输入同结果（可重放）。"""
    from app.modules.extraction.occupation import classify_occupation

    title = "销售项目经理"  # commercial(销售) 与 product_ops(项目经理) 同时命中
    assert classify_occupation(title).category == "commercial"
    assert classify_occupation(title) == classify_occupation(title)


def test_language_gate_keeps_english_jds_out_of_the_feature_space() -> None:
    """英文 JD 在中文特征管线里只能与彼此匹配，会聚成一个假岗位。"""
    from app.modules.extraction.occupation import chinese_character_ratio, is_supported_language

    chinese = "负责具身智能机器人运动控制算法的设计与实现，参与整机联调。"
    english = "Design and implement whole-body control algorithms for humanoid robots."
    assert is_supported_language(chinese)
    assert not is_supported_language(english)
    assert chinese_character_ratio("") == 0.0
    # 中英混排但以中文为主的 JD 仍然支持。
    assert is_supported_language("负责 ROS2 与 MoveIt 的机器人运动规划开发，熟悉 C++。")
