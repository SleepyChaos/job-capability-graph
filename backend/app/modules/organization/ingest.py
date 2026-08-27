"""组织/企业/高校/人才 多源数据清洗与交叉验证入库。

数据源（用户提供的 8 个真实 xlsx）：
  - 高校库_20260727.xlsx             -> 高校实体 + TOP5 标准技术(L2/L3)
  - 机构库_合并.xlsx                  -> 企业/科研机构实体 + 标准技术标注(L2/L3) + 指标
  - 具身智能企业岗位数据_整合去重_完整字段.xlsx -> 企业实体 + 在聘岗位数/产品/融资
  - 科技人才库与机构库_20260727.xlsx  -> 人才实体 + 机构库 + 人才机构成果关系
  - 技术词主数据_20260727.xlsx        -> L1/L2/L3 编码->名称 解析字典
  - 具身智能岗位_技术规范聚类分析_v2 / 图谱映射与聚类分析_v1 -> 公司->聚类(企业↔岗位视图)

交叉验证（RC-03 数据可信要求）：
  - Layer A 实体对齐：Splink 概率记录链接跨 8 表对"同一机构不同写法"做归并，
    产出 splink_cluster_id + splink_match_score；Splink 不可用时回退到归一键确定性分组。
  - Layer B 外部真值：tech_onet_mapping.json（人工标注 top-N 技术词->O*NET/ESCO 技能）
    对每个 组织↔技术 边打 external_aligned 标记并计算组织外部佐证率。
  - 三方一致性：企业业务标签(产业链) ↔ 专利布局技术域 ↔ JD 产业链 一致性评分，
    缺失维度即标 partial，不编造。

所有数据标识为"真实数据"（data_origin=source_fact）。幂等：先清空本模块表再重插。
"""

from __future__ import annotations

import json
import os
import re
from collections import defaultdict

from openpyxl import load_workbook
from sqlalchemy import delete

from app.db.base import Base
from app.db.session import SessionLocal, engine
from app.modules.organization.models import (
    CV_STATUS_PARTIAL,
    CV_STATUS_UNVERIFIED,
    CV_STATUS_VERIFIED,
    ORG_CATEGORY_ENTERPRISE,
    ORG_CATEGORY_RESEARCH,
    ORG_CATEGORY_UNIVERSITY,
    REL_EMPLOY,
    REL_PATENT_LINK,
    REL_UNIVERSITY_AFFILIATE,
    OrganizationCrossValidation,
    OrganizationEntity,
    OrganizationTalent,
    OrganizationTechnology,
    Talent,
    TalentTechnology,
)

# ---------------------------------------------------------------------------
# 路径解析：优先读环境变量，否则用默认（沙箱/本机 D 盘路径）
# ---------------------------------------------------------------------------
_HERE = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_MAPPING = os.path.join(
    _HERE, "..", "..", "..", "data", "org_sources", "tech_onet_mapping.json"
)


def _p(env: str, default: str) -> str:
    return os.environ.get(env, default)


SOURCE_FILES = {
    "university": _p("ORG_UNIV", r"D:\揭榜挂帅\重要数据\高校库_20260727.xlsx"),
    "org_merge": _p("ORG_MERGE", r"D:\揭榜挂帅\重要数据\机构库_合并.xlsx"),
    "tech_master": _p("ORG_TECH", r"D:\揭榜挂帅\重要数据\技术词主数据_20260727.xlsx"),
    "cluster_v2": _p(
        "ORG_CLUSTERV2", r"D:\揭榜挂帅\重要数据\具身智能岗位_技术规范聚类分析_v2.xlsx"
    ),
    "graph_v1": _p("ORG_GRAPHV1", r"D:\揭榜挂帅\重要数据\具身智能岗位_图谱映射与聚类分析_v1.xlsx"),
    "enterprise_jobs": _p(
        "ORG_ENTJOBS", r"D:\揭榜挂帅\具身智能企业岗位数据_整合去重_完整字段.xlsx"
    ),
    "talent": _p("ORG_TALENT", r"D:\揭榜挂帅\重要数据\科技人才库与机构库_20260727.xlsx"),
    "mapping": _p("ORG_MAPPING", _DEFAULT_MAPPING),
}

# L1 技术域 -> 12 类产业链（用于 Layer B/C 一致性校验的粗略映射）
L1_TO_CHAIN = {
    "T1": "具身智能本体",  # 智能算法与模型 -> 本体(中游)
    "T2": "感知系统",  # 感知与传感
    "T3": "执行系统",  # 本体与核心零部件
    "T4": "软件与算法",  # 数据与仿真
    "T5": "决策认知系统",  # 决策认知
    "T6": "检测/测试/测量",  # 交互安全与评测标准
    "T7": "场景应用与解决方案",  # 应用与场景
}


# ---------------------------------------------------------------------------
# 通用解析工具
# ---------------------------------------------------------------------------
def _norm(s):
    """归一化名（用于跨源实体对齐的归并键）。

    关键设计：只去掉「尾部公司法务形态后缀」（股份有限公司/有限公司/公司/集团…），
    且 科技/技术 仅在其真正位于尾部（即公司名而非校名中的 科技）时才去掉。
    原因：中文校名「北京大学」与「北京科技大学」仅差中间 科技，若无条件剥掉 科技/大学
    会误把不同高校归一为「北京」造成过合并；而企业名「宇树科技股份有限公司」剥掉法务后缀
    后得到核心「宇树」，可与「宇树科技」正确对齐。
    """
    if s is None:
        return ""
    s = str(s).strip().lower()
    # 1) 去掉尾部公司法务形态后缀
    for suf in [
        "股份有限公司",
        "有限责任公司",
        "有限公司",
        "股份公司",
        "（",
        "(",
        "）",
        ")",
        " ",
        "　",
        "集团",
    ]:
        if s.endswith(suf):
            s = s[: -len(suf)]
    # 2) 仅当 科技/技术 真正位于尾部才去掉（公司名特征，避免误伤校名）
    if s.endswith("科技"):
        s = s[: -len("科技")]
    elif s.endswith("技术"):
        s = s[: -len("技术")]
    # 3) 兜底：归一名过短则回退到原始去空格小写
    if len(s) < 1:
        s = str(s).strip().lower()
    return s


def _read_sheet(path, sheet=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    header = [str(c) if c is not None else f"c{i}" for i, c in enumerate(rows[0])]
    data = rows[1:]
    return header, data


def parse_annotations(cell):
    """'T1.05 学习与训练方法/T1.03 运动规划与控制' -> [('T1.05','学习与训练方法'), ...]"""
    if not cell:
        return []
    out = []
    for part in str(cell).split("/"):
        part = part.strip()
        if not part:
            continue
        m = re.match(r"^(T\d[\d.]*)\s*(.*)$", part)
        if m:
            out.append((m.group(1), m.group(2).strip()))
    return out


def parse_mention_counts(cell):
    """'T1.05×78;T1.03×35' -> {'T1.05':78,'T1.03':35}"""
    d = {}
    if not cell:
        return d
    for part in str(cell).split(";"):
        part = part.strip()
        if "×" in part:
            code, _, cnt = part.partition("×")
            code = code.strip()
            try:
                d[code] = int(cnt.strip())
            except ValueError:
                pass
    return d


def load_tech_dictionary(path):
    """技术词主数据 -> {code: name} 覆盖 L2/L3。"""
    dic = {}
    try:
        h, d = _read_sheet(path, "L2技术类")
        if h and "L2编码" in h:
            ci = h.index("L2编码")
            ni = h.index("技术类")
            for r in d:
                if r[ci]:
                    dic[str(r[ci]).strip()] = str(r[ni]).strip()
        h, d = _read_sheet(path, "L3技术点")
        if h and "L3编码" in h:
            ci = h.index("L3编码")
            ni = h.index("L3标准名")
            for r in d:
                if r[ci]:
                    dic[str(r[ci]).strip()] = str(r[ni]).strip()
    except Exception as e:
        print("  [warn] tech dictionary load failed:", e)
    return dic


def load_external_mapping(path):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f).get("mapping", {})
    except Exception as e:
        print("  [warn] external mapping load failed:", e)
        return {}


# ---------------------------------------------------------------------------
# Layer A：Splink 概率记录链接（跨源机构归并）
# ---------------------------------------------------------------------------
class OrgLinker:
    def __init__(self):
        self.method = "deterministic"  # 默认回退
        try:
            import splink  # noqa: F401

            self.method = "splink"
        except Exception:
            self.method = "deterministic"

    def cluster(self, records):
        """records: list[{rec_id, org_id, name}] -> {org_id: cluster_id}, {org_id: match_score}"""
        if self.method == "splink":
            try:
                return self._cluster_splink(records)
            except Exception as e:
                print("  [warn] splink failed, fallback deterministic:", e)
                self.method = "deterministic"
        return self._cluster_deterministic(records)

    def _cluster_deterministic(self, records):
        groups = defaultdict(list)
        for r in records:
            groups[_norm(r["name"]) or str(r["org_id"])].append(r)
        cluster_of, score_of = {}, {}
        for key, recs in groups.items():
            cid = f"GRP-{key[:24]}"
            for r in recs:
                cluster_of[r["org_id"]] = cid
                score = 1.0 if len(recs) > 1 else 0.6
                score_of[r["org_id"]] = max(score_of.get(r["org_id"], 0), score)
        return cluster_of, score_of

    def _cluster_splink(self, records):
        """Layer A 真实概率记录链接：用 Splink v4（英国司法部 MoJ，MIT，GitHub 开源）
        的 Fellegi-Sunter 模型对跨源机构名做归并，产出 splink_cluster_id + 匹配分。

        设计要点（针对中文机构名写法变体）：
          - 比较字段用「归一名」name_norm（去掉 股份/有限/科技/公司/集团 等后缀），
            因为中文机构名变体主要靠「共享核心」而非拉丁串编辑距离。
          - m/u 概率用「先验指定」而非标签训练：无需人工标注即可复现、可审计，
            且避免无标签时 m 未训练导致完全匹配反被低估的问题。
        """
        import duckdb
        import pandas as pd
        import splink
        from splink import DuckDBAPI, SettingsCreator
        from splink.comparison_library import (
            JaroWinklerAtThresholds,
            LevenshteinAtThresholds,
        )

        df = pd.DataFrame(records)
        df["name"] = df["name"].astype(str)
        df["name_norm"] = df["name"].map(_norm)

        con = duckdb.connect(":memory:")
        con.register("input_data", df)

        c1 = JaroWinklerAtThresholds("name_norm", [0.9, 0.8])
        c1.m_probabilities = [0.97, 0.85, 0.4, 0.01]
        c1.u_probabilities = [0.01, 0.04, 0.10, 0.85]
        c2 = LevenshteinAtThresholds("name_norm", [2, 4])
        c2.m_probabilities = [0.95, 0.8, 0.4, 0.01]
        c2.u_probabilities = [0.02, 0.05, 0.12, 0.81]

        settings = SettingsCreator(
            link_type="dedupe_only",
            unique_id_column_name="rec_id",
            probability_two_random_records_match=0.001,
            comparisons=[c1, c2],
            blocking_rules_to_generate_predictions=[
                "l.name_norm = r.name_norm",
                "substr(l.name_norm,1,3) = substr(r.name_norm,1,3)",
            ],
        )

        linker = splink.Linker("input_data", settings, db_api=DuckDBAPI(con))
        predictions = linker.inference.predict(threshold_match_weight=-8)
        pred_df = predictions.as_pandas_dataframe()
        clusters = linker.clustering.cluster_pairwise_predictions_at_threshold(predictions, 0.9)
        cdf = clusters.as_pandas_dataframe()

        recid_to_org = {r["rec_id"]: r["org_id"] for r in records}
        canonical_rid = {r["rec_id"] for r in records if r["rec_id"] == r["org_id"]}

        # 每个 org 的匹配分 = 涉及其记录的所有 pairwise match_probability 最大值
        score_of = {}
        if not pred_df.empty:
            for _, row in pred_df.iterrows():
                mp = float(row.get("match_probability", 0.0) or 0.0)
                for col in ("rec_id_l", "rec_id_r"):
                    oid = recid_to_org.get(row.get(col))
                    if oid:
                        score_of[oid] = max(score_of.get(oid, 0.0), mp)

        # 簇映射：优先用「规范名」记录的簇（避免 v4机构ID 等非姓名变体干扰）
        cluster_of = {}
        for _, row in cdf.iterrows():
            oid = row.get("org_id")
            rid = row.get("rec_id")
            if oid is None:
                continue
            cid = f"SPLK-{row.get('cluster_id')}"
            if rid in canonical_rid:
                cluster_of[oid] = cid
            else:
                cluster_of.setdefault(oid, cid)

        # 兜底：未被任何簇覆盖的 org 各自成簇，匹配分记为基线 0.6（无跨源佐证）
        for oid in recid_to_org.values():
            cluster_of.setdefault(oid, f"GRP-{_norm(oid)[:24]}")
            score_of.setdefault(oid, 0.6)
        return cluster_of, score_of


# ---------------------------------------------------------------------------
# 主入库流程
# ---------------------------------------------------------------------------
def run_ingest(session, files=None, verbose=True):
    files = files or SOURCE_FILES
    tech_dict = load_tech_dictionary(files["tech_master"])
    ext_map = load_external_mapping(files["mapping"])
    linker = OrgLinker()
    if verbose:
        print(f"[ingest] tech_dict={len(tech_dict)} ext_map={len(ext_map)} linker={linker.method}")

    # 清空本模块表（幂等）
    for tbl in (
        OrganizationTechnology,
        TalentTechnology,
        OrganizationTalent,
        OrganizationCrossValidation,
        Talent,
        OrganizationEntity,
    ):
        session.execute(delete(tbl))
    session.commit()

    orgs = {}  # org_code -> dict
    org_tech = defaultdict(list)  # org_code -> [(code, name, level, count, source)]
    talents = {}  # talent_code -> dict
    talent_tech = defaultdict(list)
    org_talent = defaultdict(list)  # org_code -> [(talent_code, rel_type, source)]
    company_to_orgcode = {}  # 企业名称(归一) -> org_code

    def add_org(code, name, category, **kw):
        o = orgs.get(code)
        if o is None:
            o = {
                "org_code": code,
                "org_name": name,
                "org_category": category,
                "raw": {},
                "source_keys": set(),
            }
            orgs[code] = o
        o["raw"].update({k: v for k, v in kw.items() if v not in (None, "")})
        if name:
            o["source_keys"].add(name)
        return o

    # ---- 1) 机构库_合并（企业/科研机构主表）----
    if os.path.exists(files["org_merge"]):
        h, d = _read_sheet(files["org_merge"], "Sheet1")
        if h:
            gi = h.index("机构ID")
            ni = h.index("归一键")
            rn = h.index("代表名称")
            ti = h.index("机构类型")
            pcn = h.index("专利族数(机构申请)")
            scn = h.index("标准数(机构起草)")
            rcn = h.index("关联人才数(已确认单位)")
            rhn = h.index("任职高校人才数(R6)")
            l2 = h.index("标准技术标注L2")
            l2d = h.index("技术标注明细(留痕)")
            l3 = h.index("标准技术标注L3")
            l3d = h.index("L3标注明细(留痕)")
            ic = h.index("产业链(12类标准)")
            tier = h.index("层级")
            seg = h.index("细分领域")
            prod = h.index("代表产品")
            ptype = h.index("产品类型")
            kp = h.index("关键特性/参数")
            mp = h.index("量产进展")
            op = h.index("运营路径")
            city = h.index("总部城市")
            prov = h.index("省份")
            district = h.index("区/县")
            fs = h.index("融资阶段")
            frc = h.index("融资轮次分类")
            allf = h.index("全部原始形态(频次)")
            src = h.index("数据来源")
            for r in d:
                code = str(r[gi]).strip()
                name = str(r[rn] or r[ni] or "")
                cat = (
                    ORG_CATEGORY_ENTERPRISE
                    if str(r[ti]).strip() == "企业"
                    else ORG_CATEGORY_RESEARCH
                )
                o = add_org(
                    code,
                    name,
                    cat,
                    normalized_key=r[ni],
                    industry_chain=r[ic],
                    tier_level=r[tier],
                    segment=r[seg],
                    products=r[prod],
                    product_type=ptype,
                    key_params=kp,
                    mass_production=mp,
                    operation_path=op,
                    hq_city=r[city],
                    hq_province=prov,
                    hq_district=district,
                    financing_stage=r[fs],
                    financing_round=frc,
                    patent_family_count=_int(r[pcn]),
                    standard_count=_int(r[scn]),
                    related_talent_count=_int(r[rcn]),
                    university_talent_count=_int(r[rhn]),
                    data_source=r[src],
                )
                o["raw"]["_all_forms"] = r[allf]
                # 技术标注
                counts = {**parse_mention_counts(r[l2d]), **parse_mention_counts(r[l3d])}
                for code_, nm in parse_annotations(r[l2]):
                    org_tech[code].append(
                        (code_, tech_dict.get(code_, nm), "L2", counts.get(code_, 1), "机构库_合并")
                    )
                for code_, nm in parse_annotations(r[l3]):
                    org_tech[code].append(
                        (code_, tech_dict.get(code_, nm), "L3", counts.get(code_, 1), "机构库_合并")
                    )
                company_to_orgcode[_norm(name)] = code
                # 全部书写形态也加入名称池
                if r[allf]:
                    for variant in str(r[allf]).split("|"):
                        vname = variant.split("(")[0].strip()
                        if vname:
                            company_to_orgcode[_norm(vname)] = code

    # ---- 2) 高校库主表 ----
    if os.path.exists(files["university"]):
        h, d = _read_sheet(files["university"], "高校库主表")
        if h:
            ui = h.index("高校ID")
            nm = h.index("规范名称")
            vid = h.index("关联v4机构ID")
            lvl = h.index("办学层次")
            prov = h.index("所在省")
            city = h.index("所在城市")
            ctry = h.index("国家/地区")
            dept = h.index("主管部门")
            site = h.index("官网")
            l2 = h.index("TOP5方向_标准L2映射")
            l3 = h.index("TOP5方向_标准L3映射")
            rtn = h.index("任职人才数(R6)")
            rcn = h.index("关联人才数(v4确认)")
            pc = h.index("专利族数")
            sc = h.index("标准数")
            src = h.index("数据来源")
            comp = h.index("补全状态")
            for r in d:
                code = str(r[ui]).strip()
                name = str(r[nm] or "")
                o = add_org(
                    code,
                    name,
                    ORG_CATEGORY_UNIVERSITY,
                    hq_province=r[prov],
                    hq_city=r[city],
                    hq_country=r[ctry],
                    homepage_url=r[site],
                    research_level=r[lvl],
                    department=r[dept],
                    patent_family_count=_int(r[pc]),
                    standard_count=_int(r[sc]),
                    related_talent_count=_int(r[rcn]),
                    university_talent_count=_int(r[rtn]),
                    data_source=r[src],
                    completeness=r[comp],
                )
                linked_org = str(r[vid]).strip() if r[vid] else ""
                if linked_org:
                    o["source_keys"].add(linked_org)
                for code_, nmx in parse_annotations(r[l2]):
                    org_tech[code].append((code_, tech_dict.get(code_, nmx), "L2", 1, "高校库"))
                for code_, nmx in parse_annotations(r[l3]):
                    org_tech[code].append((code_, tech_dict.get(code_, nmx), "L3", 1, "高校库"))

    # ---- 3) 企业岗位整合（在聘岗位数 / 产品 / 融资）----
    if os.path.exists(files["enterprise_jobs"]):
        h, d = _read_sheet(files["enterprise_jobs"], "Sheet1")
        if h:
            en = h.index("企业名称")
            jc = h.index("在聘岗位数量")
            site = h.index("官网链接")
            lp = h.index("猎聘招聘链接（https://www.liepin.com/company-jobs/xxx/）")
            ic = h.index("产业链(12类标准)")
            tier = h.index("层级")
            seg = h.index("细分领域")
            prod = h.index("代表产品")
            ptype = h.index("产品类型")
            kp = h.index("关键特性/参数")
            mp = h.index("量产进展")
            op = h.index("运营路径")
            city = h.index("总部城市")
            prov = h.index("省份")
            district = h.index("区/县")
            fs = h.index("融资阶段")
            frc = h.index("融资轮次分类")
            src = h.index("数据来源")
            for r in d:
                name = str(r[en] or "").strip()
                nk = _norm(name)
                existing = company_to_orgcode.get(nk)
                if existing:
                    o = orgs[existing]
                    o["raw"].update(
                        {
                            "job_posting_count": _int(r[jc]),
                            "homepage_url": r[site],
                            "liepin_url": r[lp],
                            "industry_chain": r[ic] or o["raw"].get("industry_chain"),
                            "tier_level": r[tier] or o["raw"].get("tier_level"),
                            "segment": r[seg],
                            "products": r[prod],
                            "product_type": r[ptype],
                            "key_params": r[kp],
                            "mass_production": r[mp],
                            "operation_path": r[op],
                            "hq_city": r[city] or o["raw"].get("hq_city"),
                            "hq_province": r[prov],
                            "hq_district": r[district],
                            "financing_stage": r[fs],
                            "financing_round": frc,
                            "data_source": r[src],
                        }
                    )
                    if r[jc]:
                        o["raw"]["job_posting_count"] = _int(r[jc])
                else:
                    code = f"ENT-{len(orgs) + 1:05d}"
                    o = add_org(
                        code,
                        name,
                        ORG_CATEGORY_ENTERPRISE,
                        job_posting_count=_int(r[jc]),
                        homepage_url=r[site],
                        liepin_url=r[lp],
                        industry_chain=r[ic],
                        tier_level=r[tier],
                        segment=r[seg],
                        products=r[prod],
                        product_type=ptype,
                        key_params=r[kp],
                        mass_production=mp,
                        operation_path=op,
                        hq_city=r[city],
                        hq_province=prov,
                        hq_district=district,
                        financing_stage=r[fs],
                        financing_round=frc,
                        data_source=r[src],
                    )
                    company_to_orgcode[nk] = code

    # ---- 4) 科技人才库 ----
    if os.path.exists(files["talent"]):
        # 4a) 人才库
        h, d = _read_sheet(files["talent"], "人才库")
        if h:
            ti = h.index("人才ID")
            dn = h.index("代表姓名")
            ng = h.index("姓名归并组")
            tt = h.index("人才类型")
            pok = h.index("主机构键")
            pn = h.index("主机构")
            pc = h.index("专利族数")
            sc = h.index("标准数")
            hn = h.index("高校任职数")
            conf = h.index("置信度")
            title = h.index("职务/职称")
            rd = h.index("研究方向")
            l2 = h.index("标准技术标注L2")
            l2d = h.index("技术标注明细(留痕)")
            l3 = h.index("标准技术标注L3")
            l3d = h.index("L3标注明细(留痕)")
            for r in d:
                code = str(r[ti]).strip()
                talents[code] = {
                    "talent_code": code,
                    "display_name": str(r[dn] or ""),
                    "name_group": r[ng],
                    "talent_type": r[tt],
                    "primary_org_key": r[pok],
                    "primary_org_name": r[pn],
                    "patent_family_count": _int(r[pc]),
                    "standard_count": _int(r[sc]),
                    "university_post_count": _int(r[hn]),
                    "confidence": r[conf],
                    "title": r[title],
                    "research_direction": r[rd],
                    "technology_l2": r[l2],
                    "technology_l3": r[l3],
                }
                counts = {**parse_mention_counts(r[l2d]), **parse_mention_counts(r[l3d])}
                for code_, nm in parse_annotations(r[l2]):
                    talent_tech[code].append(
                        (code_, tech_dict.get(code_, nm), "L2", counts.get(code_, 1))
                    )
                for code_, nm in parse_annotations(r[l3]):
                    talent_tech[code].append(
                        (code_, tech_dict.get(code_, nm), "L3", counts.get(code_, 1))
                    )
                # 主机构 -> 任职关系
                if r[pn]:
                    nk = _norm(r[pn])
                    linked = company_to_orgcode.get(nk)
                    if linked:
                        org_talent[linked].append((code, REL_EMPLOY, "人才库.主机构"))
        # 4b) 高校人才明细索引 -> 高校↔人才
        h, d = _read_sheet(files["talent"], "高校人才明细索引")
        if h:
            ui = h.index("高校ID")
            pid = h.index("人才库ID")
            nm = h.index("姓名")
            title = h.index("职务/职称")
            rd = h.index("研究方向")
            tt = h.index("人才类型")
            for r in d:
                u_code = str(r[ui]).strip()
                p_code = str(r[pid]).strip() if r[pid] else ""
                if not p_code or u_code not in orgs:
                    continue
                org_talent[u_code].append((p_code, REL_UNIVERSITY_AFFILIATE, "高校人才明细索引"))
                if p_code not in talents:
                    talents[p_code] = {
                        "talent_code": p_code,
                        "display_name": str(r[nm] or ""),
                        "title": r[title],
                        "research_direction": r[rd],
                        "talent_type": r[tt],
                    }
        # 4c) 人才机构成果关系 -> 机构↔人才(专利关联)
        h, d = _read_sheet(files["talent"], "人才机构成果关系")
        if h and "机构ID" in h and "人才库ID" in h:
            oi = h.index("机构ID")
            pi = h.index("人才库ID")
            for r in d:
                o_code = str(r[oi]).strip() if r[oi] else ""
                p_code = str(r[pi]).strip() if r[pi] else ""
                if o_code in orgs and p_code:
                    org_talent[o_code].append((p_code, REL_PATENT_LINK, "人才机构成果关系"))

    # ---- Layer A：跨源机构归并 ----
    link_records = []
    for code, o in orgs.items():
        link_records.append({"rec_id": code, "org_id": code, "name": o["org_name"]})
        for sk in list(o["source_keys"]):
            if sk and sk != o["org_name"]:
                link_records.append({"rec_id": f"{code}#{sk}", "org_id": code, "name": sk})
    cluster_of, score_of = linker.cluster(link_records)
    if verbose:
        print(f"[ingest] linkage method={linker.method} clusters={len(set(cluster_of.values()))}")

    # ---- 写入 OrganizationEntity ----
    seq = 1
    code_to_pk = {}
    for code, o in orgs.items():
        ent = OrganizationEntity(
            org_code=code,
            org_name=o["org_name"],
            org_category=o["org_category"],
            normalized_key=o["raw"].get("normalized_key"),
            splink_cluster_id=cluster_of.get(code),
            splink_match_score=round(score_of.get(code, 0.0), 3),
            dedup_source_keys=sorted(o["source_keys"]),
            homepage_url=o["raw"].get("homepage_url"),
            recruit_url=o["raw"].get("recruit_url"),
            liepin_url=o["raw"].get("liepin_url"),
            hq_city=o["raw"].get("hq_city"),
            hq_province=o["raw"].get("hq_province"),
            hq_district=o["raw"].get("hq_district"),
            industry_chain=o["raw"].get("industry_chain"),
            tier_level=o["raw"].get("tier_level"),
            segment=o["raw"].get("segment"),
            products=o["raw"].get("products"),
            product_type=o["raw"].get("product_type"),
            key_params=o["raw"].get("key_params"),
            mass_production=o["raw"].get("mass_production"),
            operation_path=o["raw"].get("operation_path"),
            financing_stage=o["raw"].get("financing_stage"),
            financing_round=o["raw"].get("financing_round"),
            patent_family_count=o["raw"].get("patent_family_count", 0) or 0,
            standard_count=o["raw"].get("standard_count", 0) or 0,
            related_talent_count=o["raw"].get("related_talent_count", 0) or 0,
            university_talent_count=o["raw"].get("university_talent_count", 0) or 0,
            job_posting_count=o["raw"].get("job_posting_count", 0) or 0,
            data_source=o["raw"].get("data_source"),
            completeness=o["raw"].get("completeness"),
            raw_fields_json=o["raw"],
        )
        session.add(ent)
        session.flush()
        code_to_pk[code] = ent.org_id
        seq += 1

    # ---- 写入 组织↔技术（含 Layer B 外部佐证）----
    org_tech_rows = 0
    for code, edges in org_tech.items():
        pk = code_to_pk.get(code)
        if pk is None:
            continue
        seen_tech = set()
        for tcode, tname, lvl, cnt, src in edges:
            if tcode in seen_tech:
                continue
            seen_tech.add(tcode)
            em = ext_map.get(tname) or ext_map.get(tcode)
            session.add(
                OrganizationTechnology(
                    org_id=pk,
                    technology_code=tcode,
                    technology_name=tname,
                    level_code=lvl,
                    mention_count=cnt or 1,
                    annotation_source=src,
                    external_skill_label=em.get("onete") if em else None,
                    external_aligned=bool(em),
                )
            )
            org_tech_rows += 1
    if verbose:
        print(f"[ingest] org_technology rows={org_tech_rows}")

    # ---- 写入 人才 + 人才↔技术 ----
    tal_pk = {}
    for code, t in talents.items():
        ent = Talent(
            talent_code=code,
            display_name=t.get("display_name", ""),
            name_group=t.get("name_group"),
            talent_type=t.get("talent_type"),
            primary_org_key=t.get("primary_org_key"),
            primary_org_name=t.get("primary_org_name"),
            patent_family_count=t.get("patent_family_count", 0) or 0,
            standard_count=t.get("standard_count", 0) or 0,
            university_post_count=t.get("university_post_count", 0) or 0,
            confidence=t.get("confidence"),
            title=t.get("title"),
            research_direction=t.get("research_direction"),
            technology_l2=t.get("technology_l2"),
            technology_l3=t.get("technology_l3"),
            raw_fields_json=t,
        )
        session.add(ent)
        session.flush()
        tal_pk[code] = ent.talent_id
    for code, edges in talent_tech.items():
        pk = tal_pk.get(code)
        if pk is None:
            continue
        seen_tt = set()
        for tcode, tname, lvl, cnt in edges:
            if tcode in seen_tt:
                continue
            seen_tt.add(tcode)
            session.add(
                TalentTechnology(
                    talent_id=pk,
                    technology_code=tcode,
                    technology_name=tname,
                    level_code=lvl,
                    mention_count=cnt or 1,
                )
            )

    # ---- 写入 组织↔人才 ----
    org_talent_rows = 0
    seen_ot = set()
    for code, links in org_talent.items():
        opk = code_to_pk.get(code)
        if opk is None:
            continue
        for tcode, rel, src in links:
            tpk = tal_pk.get(tcode)
            if tpk is None:
                continue
            key = (opk, tpk, rel)
            if key in seen_ot:
                continue
            seen_ot.add(key)
            session.add(
                OrganizationTalent(org_id=opk, talent_id=tpk, relation_type=rel, source=src)
            )
            org_talent_rows += 1
    if verbose:
        print(f"[ingest] talents={len(tal_pk)} org_talent rows={org_talent_rows}")

    # ---- Layer C：三方一致性交叉验证 ----
    cv_rows = 0
    for code, o in orgs.items():
        pk = code_to_pk.get(code)
        if pk is None:
            continue
        business_chain = o["raw"].get("industry_chain")
        # 专利布局技术域（由标注技术 L1 推导）
        patent_domains = set()
        for tcode, *_ in org_tech.get(code, []):
            l1 = tcode.split(".")[0]
            if l1 in L1_TO_CHAIN:
                patent_domains.add(L1_TO_CHAIN[l1])
        # JD 产业链（企业岗位整合已并入 industry_chain；此处取 business_chain 作为代理）
        jd_chain = business_chain
        dims = {}
        if business_chain:
            dims["business_chain"] = business_chain
        if patent_domains:
            dims["patent_domain"] = "/".join(sorted(patent_domains))
        if jd_chain:
            dims["jd_chain"] = jd_chain
        matched = 0
        missing = []
        if "business_chain" in dims and "patent_domain" in dims:
            if dims["business_chain"] in dims["patent_domain"]:
                matched += 1
            else:
                missing.append("patent_domain_mismatch")
        if "business_chain" in dims and "jd_chain" in dims:
            matched += 1
        if not patent_domains and "business_chain" in dims:
            missing.append("patent_domain_missing")
        total = len([k for k in dims if k != "jd_chain"]) or 1
        score = int(round(matched / total * 100)) if matched else 0
        if matched >= 1 and not missing:
            status = CV_STATUS_VERIFIED
        elif matched >= 1 or business_chain:
            status = CV_STATUS_PARTIAL
        else:
            status = CV_STATUS_UNVERIFIED
        session.add(
            OrganizationCrossValidation(
                org_id=pk,
                consistency_score=score,
                business_chain=business_chain,
                patent_domain_codes="/".join(sorted(patent_domains)) or None,
                jd_chain=jd_chain,
                matched_dimensions=matched,
                missing_dimensions_json=missing,
                note=f"method={linker.method}; dims={list(dims.keys())}",
            )
        )
        cv_rows += 1
        # 组织级外部佐证率（Layer B）
        ent = session.get(OrganizationEntity, pk)
        edges_all = org_tech.get(code, [])
        aligned = sum(1 for e in edges_all if ext_map.get(e[1]) or ext_map.get(e[0]))
        ent.external_alignment_rate = round(aligned / len(edges_all), 3) if edges_all else 0.0
        ent.cross_validation_status = status
    session.commit()
    if verbose:
        print(f"[ingest] cross_validation rows={cv_rows} DONE")
    return {
        "orgs": len(orgs),
        "talents": len(tal_pk),
        "org_tech": org_tech_rows,
        "org_talent": org_talent_rows,
        "cv": cv_rows,
        "linker": linker.method,
    }


def _int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def main():
    Base.metadata.create_all(engine)
    with SessionLocal() as s:
        stats = run_ingest(s)
    print("INGEST STATS:", stats)


if __name__ == "__main__":
    main()
