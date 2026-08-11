from __future__ import annotations

import hashlib
import json
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.modules.ingestion.models import FileImportRun, RawFileAsset, SpreadsheetRow
from tools.profile_source_workbooks import sha256_file


@dataclass(frozen=True)
class StageWorkbookResult:
    file_asset_id: int
    file_import_run_id: int
    import_run_code: str
    staged_row_count: int
    skipped_existing_row_count: int


def json_value(value: Any) -> Any:
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def hash_json(payload: Any) -> str:
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(serialized.encode()).hexdigest()


class WorkbookStagingService:
    def __init__(self, session: Session):
        self.session = session

    def stage(
        self,
        path: Path,
        *,
        storage_object_key: str,
        importer_code: str,
        mapping_code: str,
        mapping_version: str,
        access_classification: str,
        external_key_fields: dict[str, str] | None = None,
    ) -> StageWorkbookResult:
        path = path.resolve()
        file_hash = sha256_file(path)
        asset = self._get_or_create_asset(path, storage_object_key, file_hash)
        schema_hash = self._schema_hash(path)
        run, run_created = self._get_or_create_run(
            asset,
            importer_code=importer_code,
            mapping_code=mapping_code,
            mapping_version=mapping_version,
            schema_hash=schema_hash,
        )
        if not run_created and run.import_status_code == "success":
            skipped_count = self.session.scalar(
                select(func.count())
                .select_from(SpreadsheetRow)
                .where(SpreadsheetRow.file_asset_id == asset.file_asset_id)
            )
            return StageWorkbookResult(
                file_asset_id=asset.file_asset_id,
                file_import_run_id=run.file_import_run_id,
                import_run_code=run.import_run_code,
                staged_row_count=0,
                skipped_existing_row_count=skipped_count or 0,
            )
        run.import_status_code = "running"
        run.started_at = run.started_at or datetime.now(UTC).replace(tzinfo=None)
        self.session.flush()

        staged = 0
        skipped = 0
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            for sheet in workbook.worksheets:
                rows = sheet.iter_rows(values_only=True)
                header_values = next(rows, None)
                if header_values is None:
                    continue
                headers = self._headers(header_values)
                header_hash = hash_json(headers)
                external_field = (external_key_fields or {}).get(sheet.title)
                for row_number, row_values in enumerate(rows, start=2):
                    payload = {
                        header: json_value(row_values[index]) if index < len(row_values) else None
                        for index, header in enumerate(headers)
                    }
                    if not any(value not in (None, "") for value in payload.values()):
                        continue
                    exists = self.session.scalar(
                        select(SpreadsheetRow.spreadsheet_row_id).where(
                            SpreadsheetRow.file_asset_id == asset.file_asset_id,
                            SpreadsheetRow.sheet_name == sheet.title,
                            SpreadsheetRow.source_row_number == row_number,
                        )
                    )
                    if exists is not None:
                        skipped += 1
                        continue
                    external_key = payload.get(external_field) if external_field else None
                    self.session.add(
                        SpreadsheetRow(
                            file_asset_id=asset.file_asset_id,
                            sheet_name=sheet.title,
                            source_row_number=row_number,
                            external_record_key=(
                                str(external_key) if external_key not in (None, "") else None
                            ),
                            header_schema_hash=header_hash,
                            row_content_hash=hash_json(payload),
                            row_payload_json=payload,
                            access_classification_code=access_classification,
                        )
                    )
                    staged += 1
                self.session.flush()
        finally:
            workbook.close()

        run.total_row_count = staged + skipped
        run.success_row_count = staged
        run.skipped_row_count = skipped
        run.import_status_code = "success"
        run.completed_at = datetime.now(UTC).replace(tzinfo=None)
        self.session.commit()
        return StageWorkbookResult(
            file_asset_id=asset.file_asset_id,
            file_import_run_id=run.file_import_run_id,
            import_run_code=run.import_run_code,
            staged_row_count=staged,
            skipped_existing_row_count=skipped,
        )

    def _get_or_create_asset(
        self, path: Path, storage_object_key: str, file_hash: str
    ) -> RawFileAsset:
        asset = self.session.scalar(
            select(RawFileAsset).where(
                RawFileAsset.sha256_hash == file_hash,
                RawFileAsset.asset_type_code == "xlsx",
            )
        )
        if asset is not None:
            return asset
        asset = RawFileAsset(
            asset_code=f"xlsx_{file_hash[:24]}",
            asset_type_code="xlsx",
            storage_object_key=storage_object_key,
            original_file_name=path.name,
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_size_bytes=path.stat().st_size,
            sha256_hash=file_hash,
            virus_scan_status_code="not_scanned",
        )
        self.session.add(asset)
        self.session.flush()
        return asset

    def _get_or_create_run(
        self,
        asset: RawFileAsset,
        *,
        importer_code: str,
        mapping_code: str,
        mapping_version: str,
        schema_hash: str,
    ) -> tuple[FileImportRun, bool]:
        idempotency_key = hash_json(
            [asset.sha256_hash, importer_code, mapping_code, mapping_version, schema_hash]
        )
        run = self.session.scalar(
            select(FileImportRun).where(FileImportRun.idempotency_key == idempotency_key)
        )
        if run is not None:
            return run, False
        run = FileImportRun(
            import_run_code=f"imp_{idempotency_key[:24]}",
            file_asset_id=asset.file_asset_id,
            importer_code=importer_code,
            mapping_code=mapping_code,
            mapping_version=mapping_version,
            source_schema_hash=schema_hash,
            idempotency_key=idempotency_key,
        )
        self.session.add(run)
        self.session.flush()
        return run, True

    @staticmethod
    def _headers(values: tuple[Any, ...]) -> list[str]:
        return [
            str(value).strip() if value not in (None, "") else f"__column_{index + 1}"
            for index, value in enumerate(values)
        ]

    def _schema_hash(self, path: Path) -> str:
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            payload = []
            for sheet in workbook.worksheets:
                first_row = next(sheet.iter_rows(values_only=True), ())
                payload.append({"sheet": sheet.title, "headers": self._headers(first_row)})
            return hash_json(payload)
        finally:
            workbook.close()
